# import os
# import re
# import logging
# from io import BytesIO
# from datetime import datetime
# from dateutil.relativedelta import relativedelta

# import pandas as pd
# from openpyxl import Workbook, load_workbook
# from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
# from openpyxl.utils import get_column_letter
# from openpyxl.utils.dataframe import dataframe_to_rows
# from dotenv import load_dotenv
# import ibm_boto3
# from ibm_botocore.client import Config

# # -----------------------------------------------------------------------------
# # CONFIG / CONSTANTS
# # -----------------------------------------------------------------------------
# load_dotenv()
# logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(levelname)s] %(message)s')
# logger = logging.getLogger(__name__)

# # Validate required environment variables
# required = {
#     'COS_API_KEY': os.getenv('COS_API_KEY'),
#     'COS_SERVICE_INSTANCE_CRN': os.getenv('COS_SERVICE_INSTANCE_CRN'),
#     'COS_ENDPOINT': os.getenv('COS_ENDPOINT'),
#     'COS_BUCKET_NAME': os.getenv('COS_BUCKET_NAME'),
# }
# missing = [k for k, v in required.items() if not v]
# if missing:
#     logger.error(f"Missing required environment variables: {', '.join(missing)}")
#     raise SystemExit(1)

# COS_API_KEY     = required['COS_API_KEY']
# COS_CRN         = required['COS_SERVICE_INSTANCE_CRN']
# COS_ENDPOINT    = required['COS_ENDPOINT']
# BUCKET          = required['COS_BUCKET_NAME']

# # Folder paths in COS
# KRA_FOLDER = "Milestone/"
# TRACKER_FOLDER = "Wave City Club/"

# # Dynamic KRA file path
# WCC_KRA_KEY = None

# # Dynamic tracker paths
# TRACKER_PATHS = {}  # Maps month names to tracker file paths
# LOADED_TRACKERS = {}  # Cache loaded workbooks

# # Dynamic months and years
# MONTHS = []
# MONTH_YEARS = {}  # Maps month name to year
# TARGET_END_MONTH = None
# TARGET_END_YEAR = None

# # Quarterly definitions (for reference)
# QUARTERLY_GROUPS = [
#     ['June', 'July', 'August'],           # Q1
#     ['September', 'October', 'November'], # Q2
#     ['December', 'January', 'February'],  # Q3
#     ['March', 'April', 'May']             # Q4
# ]

# # Month shift logic: Report Month -> Tracker Month
# MONTH_SHIFT = {
#     'June': 'July',
#     'July': 'August',
#     'August': 'September',
#     'September': 'October',
#     'October': 'November',
#     'November': 'December',
#     'December': 'January',
#     'January': 'February',
#     'February': 'March',
#     'March': 'April',
#     'April': 'May',
#     'May': 'June'
# }

# # Block mapping from KRA to tracker sheets
# BLOCK_MAPPING = {
#     'Block 1 (B1) Banquet Hall': 'B1 Banket Hall & Finedine ',
#     'Fine Dine': 'B1 Banket Hall & Finedine ',
#     'Block 5 (B5) Admin + Member Lounge+Creche+Av Room + Surveillance Room +Toilets': 'B5',
#     'Block 6 (B6) Toilets': 'B6',
#     'Block 7(B7) Indoor Sports': 'B7',
#     'Block 9 (B9) Spa & Saloon': 'B9',
#     'Block 8 (B8) Squash Court': 'B8',
#     'Block 2 & 3 (B2 & B3) Cafe & Bar': 'B2 & B3',
#     'Block 4 (B4) Indoor Swimming Pool Changing Room & Toilets': 'B4',
#     'Block 11 (B11) Guest House': 'B11',
#     'Block 10 (B10) Gym': 'B10'
# }

# # KRA Column structure (0-indexed)
# KRA_COLUMNS = {
#     'BLOCK': 0,          # Column A: Block name
#     'MONTH1_ACTIVITY': 1,  # Column B: First month activity
#     'MONTH1_TARGET': 2,    # Column C: First month target %
#     'MONTH2_ACTIVITY': 3,  # Column D: Second month activity
#     'MONTH2_TARGET': 4,    # Column E: Second month target %
#     'MONTH3_ACTIVITY': 5,  # Column F: Third month activity
#     'MONTH3_TARGET': 6,    # Column G: Third month target %
# }

# # Tracker column positions (0-indexed)
# TRACKER_COLUMNS = {
#     'ACTIVITY_NAME': 6,    # Column G: Activity Name
#     'PERCENT_COMPLETE': 12  # Column M: % Complete
# }

# # -----------------------------------------------------------------------------
# # COS HELPERS
# # -----------------------------------------------------------------------------

# def init_cos():
#     return ibm_boto3.client(
#         's3',
#         ibm_api_key_id=COS_API_KEY,
#         ibm_service_instance_id=COS_CRN,
#         config=Config(signature_version='oauth'),
#         endpoint_url=COS_ENDPOINT,
#     )

# def download_file_bytes(cos, key):
#     if not key:
#         raise ValueError("File key cannot be None or empty")
#     obj = cos.get_object(Bucket=BUCKET, Key=key)
#     return obj['Body'].read()

# def list_files_in_folder(cos, folder_prefix):
#     """List all files in a specific folder (prefix) in the COS bucket"""
#     try:
#         response = cos.list_objects_v2(Bucket=BUCKET, Prefix=folder_prefix)
#         files = []
#         if 'Contents' in response:
#             for obj in response['Contents']:
#                 if not obj['Key'].endswith('/'):
#                     files.append(obj['Key'])
#         return files
#     except Exception as e:
#         logger.error(f"Error listing files in folder {folder_prefix}: {e}")
#         return []

# def extract_date_from_filename(filename):
#     """Extract date from filename in format (dd-mm-yyyy)"""
#     pattern = r'\((\d{2}-\d{2}-\d{4})\)'
#     match = re.search(pattern, filename)
#     if match:
#         date_str = match.group(1)
#         try:
#             return datetime.strptime(date_str, '%d-%m-%Y')
#         except ValueError:
#             logger.warning(f"Could not parse date {date_str} from filename {filename}")
#             return None
#     return None

# def get_month_name(month_num):
#     """Convert month number to month name"""
#     months = {
#         1: "January", 2: "February", 3: "March", 4: "April",
#         5: "May", 6: "June", 7: "July", 8: "August", 
#         9: "September", 10: "October", 11: "November", 12: "December"
#     }
#     return months.get(month_num, "Unknown")

# def get_month_number(month_name):
#     """Convert month name to month number"""
#     months = {
#         "January": 1, "February": 2, "March": 3, "April": 4,
#         "May": 5, "June": 6, "July": 7, "August": 8, 
#         "September": 9, "October": 10, "November": 11, "December": 12
#     }
#     return months.get(month_name, 1)

# def get_latest_kra_file(cos):
#     """Get the latest KRA milestone file from the Milestone/ folder"""
#     global WCC_KRA_KEY
    
#     logger.info("=== FINDING LATEST KRA MILESTONE FILE ===")
    
#     # List all files in Milestone folder
#     all_files = list_files_in_folder(cos, KRA_FOLDER)
#     logger.info(f"Found {len(all_files)} files in {KRA_FOLDER} folder")
    
#     # Pattern to match KRA milestone files
#     kra_pattern = r'KRA Milestones for.*\.xlsx$'
    
#     matching_files = []
    
#     for file_path in all_files:
#         filename = os.path.basename(file_path)
#         if re.search(kra_pattern, filename, re.IGNORECASE):
#             logger.info(f"Found KRA file: {filename}")
            
#             # Extract date from filename
#             file_date = extract_date_from_filename(filename)
#             if file_date:
#                 matching_files.append((file_path, file_date))
#                 logger.info(f"  Extracted date: {file_date.strftime('%d-%m-%Y')}")
#             else:
#                 logger.warning(f"  Could not extract date from {filename}")
#                 matching_files.append((file_path, datetime.min))  # Fallback for files without dates
    
#     if matching_files:
#         # Sort by date (newest first)
#         matching_files.sort(key=lambda x: x[1], reverse=True)
#         latest_file = matching_files[0][0]
#         WCC_KRA_KEY = latest_file
#         logger.info(f"✅ Latest KRA file: {os.path.basename(WCC_KRA_KEY)} (dated {matching_files[0][1].strftime('%d-%m-%Y')})")
#     else:
#         logger.error(f"❌ No KRA milestone files found in {KRA_FOLDER} folder!")
#         WCC_KRA_KEY = None
    
#     return WCC_KRA_KEY

# def extract_months_from_kra_filename(filename):
#     """
#     Extract quarter months from KRA filename
#     Format: "KRA Milestones for [Month] [Month] [Month] [Year].xlsx"
#     Example: "KRA Milestones for December January February 2026.xlsx"
#     The year in the filename is the END year of the quarter
#     """
#     months_pattern = r'KRA\s+Milestones\s+for\s+((?:January|February|March|April|May|June|July|August|September|October|November|December)(?:\s+(?:January|February|March|April|May|June|July|August|September|October|November|December))*)\s+(\d{4})'
    
#     match = re.search(months_pattern, filename, re.IGNORECASE)
#     if match:
#         months_str = match.group(1)
#         end_year = int(match.group(2))
        
#         # Extract individual months in order
#         month_names = re.findall(r'January|February|March|April|May|June|July|August|September|October|November|December', 
#                                  months_str, re.IGNORECASE)
        
#         logger.info(f"Extracted from KRA filename '{filename}':")
#         logger.info(f"  Months: {month_names}, End Year: {end_year}")
        
#         return [m.capitalize() for m in month_names], end_year
    
#     logger.warning(f"Could not parse KRA filename: {filename}")
#     return None, None

# def get_tracker_for_month(cos, report_month, month_year):
#     """
#     Get tracker file for a specific report month using the month-shift logic.
#     Uses the LATEST tracker available that matches the required month/year.
    
#     Args:
#         cos: COS client
#         report_month: The month shown in the report (e.g., "June")
#         month_year: Year of the report month
    
#     Returns:
#         Path to tracker file or None
#     """
#     # Get the tracker month based on shift logic
#     tracker_month_name = MONTH_SHIFT.get(report_month)
    
#     if not tracker_month_name:
#         logger.warning(f"No month shift mapping found for {report_month}")
#         return None
    
#     # Calculate the tracker year (handle year rollover)
#     tracker_month_num = get_month_number(tracker_month_name)
#     report_month_num = get_month_number(report_month)
    
#     tracker_year = month_year
#     # If tracker month is earlier in the year than report month, it's next year
#     if tracker_month_num < report_month_num:
#         tracker_year += 1
    
#     logger.info(f"Looking for tracker for Report: {report_month} {month_year} → Tracker: {tracker_month_name} {tracker_year}")
    
#     # List all files in Wave City Club folder
#     all_files = list_files_in_folder(cos, TRACKER_FOLDER)
#     logger.info(f"Total files in {TRACKER_FOLDER}: {len(all_files)}")
    
#     # Pattern to match Structure Work tracker files
#     tracker_pattern = r'Structure.*Work.*Tracker.*Wave.*City.*Club.*\.xlsx$'
    
#     matching_files = []
    
#     for file_path in all_files:
#         filename = os.path.basename(file_path)
#         if re.search(tracker_pattern, filename, re.IGNORECASE):
#             # Extract date from filename
#             file_date = extract_date_from_filename(filename)
            
#             logger.info(f"  Found tracker: {filename}")
            
#             if file_date:
#                 logger.info(f"    Extracted date: {file_date.strftime('%d-%m-%Y')} (Month: {file_date.month}, Year: {file_date.year})")
#                 logger.info(f"    Looking for: Month={tracker_month_num}, Year={tracker_year}")
                
#                 # Check if this file matches the month and year we're looking for
#                 if file_date.month == tracker_month_num and file_date.year == tracker_year:
#                     matching_files.append((file_path, file_date))
#                     logger.info(f"    ✅ MATCH!")
#                 else:
#                     logger.info(f"    ❌ Date mismatch (need {tracker_month_name} {tracker_year})")
#             else:
#                 logger.warning(f"    Could not extract date from {filename}")
    
#     if matching_files:
#         # Sort by date and get the LATEST matching tracker
#         matching_files.sort(key=lambda x: x[1], reverse=True)
#         latest_tracker = matching_files[0][0]
#         logger.info(f"✅ Selected tracker: {os.path.basename(latest_tracker)} (dated {matching_files[0][1].strftime('%d-%m-%Y')})")
#         return latest_tracker
#     else:
#         logger.warning(f"⚠️ No tracker found for {report_month} {month_year} (looking for {tracker_month_name} {tracker_year} dated files)")
#         return None

# def setup_quarterly_months(kra_filename):
#     """
#     Setup the quarterly months and years based on KRA filename.
    
#     The year in the KRA filename represents the END year of the quarter.
#     For cross-year quarters (e.g., Dec-Jan-Feb 2026):
#     - December belongs to 2025
#     - January belongs to 2026
#     - February belongs to 2026
#     """
#     global MONTHS, MONTH_YEARS, TARGET_END_MONTH, TARGET_END_YEAR
    
#     months, end_year = extract_months_from_kra_filename(kra_filename)
    
#     if not months or not end_year:
#         logger.error(f"Could not extract months from KRA filename: {kra_filename}")
#         return False
    
#     MONTHS = months
#     TARGET_END_MONTH = MONTHS[-1]
#     TARGET_END_YEAR = end_year
    
#     # Determine starting year for the quarter
#     first_month_num = get_month_number(MONTHS[0])
#     last_month_num = get_month_number(MONTHS[-1])
    
#     # For cross-year quarters (first month number > last month number),
#     # the first month belongs to the previous year
#     if first_month_num > last_month_num:
#         # Cross-year quarter (e.g., Dec 2025, Jan 2026, Feb 2026)
#         start_year = end_year - 1
#         logger.info(f"Cross-year quarter detected: {MONTHS[0]} {start_year} - {MONTHS[-1]} {end_year}")
#     else:
#         # Same-year quarter (e.g., Jun 2025, Jul 2025, Aug 2025)
#         start_year = end_year
#         logger.info(f"Same-year quarter: all months in {start_year}")
    
#     # Assign years to each month
#     current_year = start_year
#     for i, month in enumerate(MONTHS):
#         month_num = get_month_number(month)
        
#         # Year increments when month number decreases (year rollover)
#         if i > 0:
#             prev_month_num = get_month_number(MONTHS[i-1])
#             if month_num < prev_month_num:
#                 current_year += 1
        
#         MONTH_YEARS[month] = current_year
    
#     logger.info(f"Month-Year mapping:")
#     for month in MONTHS:
#         logger.info(f"  {month}: {MONTH_YEARS[month]}")
    
#     return True


# def find_club_sheet(wb):
#     """Find the sheet with 'Club' in its name"""
#     for sheet_name in wb.sheetnames:
#         if 'club' in sheet_name.lower():
#             logger.info(f"Found Club sheet: {sheet_name}")
#             return sheet_name
    
#     logger.warning("No sheet with 'Club' in name found. Available sheets: " + ", ".join(wb.sheetnames))
#     return None

# # -----------------------------------------------------------------------------
# # DATA EXTRACTION
# # -----------------------------------------------------------------------------

# def get_wcc_targets_from_kra(cos):
#     """
#     Extract targets from KRA milestone file with updated logic:
#     - Look for sheet with "Club" in its name
#     - For each activity, fetch both the target activity name AND target %
#     - Store both values for comparison with tracker data
#     """
#     global WCC_KRA_KEY
    
#     if not WCC_KRA_KEY:
#         raise ValueError("WCC_KRA_KEY is not set. Call get_latest_kra_file() first.")
    
#     logger.info(f"Reading KRA file: {WCC_KRA_KEY}")
    
#     # Download and load KRA file
#     kra_bytes = download_file_bytes(cos, WCC_KRA_KEY)
#     kra_wb = load_workbook(BytesIO(kra_bytes), data_only=True)
    
#     # Find the Club sheet
#     club_sheet_name = find_club_sheet(kra_wb)
#     if not club_sheet_name:
#         # Fallback to active sheet
#         logger.warning("Using active sheet as fallback")
#         ws = kra_wb.active
#     else:
#         ws = kra_wb[club_sheet_name]
    
#     targets = {}
    
#     # Find header row (look for "Blocks" in column A)
#     header_row = None
#     for row_idx in range(1, min(10, ws.max_row + 1)):
#         cell_value = ws.cell(row=row_idx, column=1).value
#         if cell_value and 'block' in str(cell_value).lower():
#             header_row = row_idx
#             logger.info(f"Found header row at row {header_row}")
#             break
    
#     if not header_row:
#         logger.error("Could not find header row with 'Blocks'")
#         kra_wb.close()
#         return targets
    
#     # Start from row after header
#     data_start_row = header_row + 1
    
#     for row_idx in range(data_start_row, ws.max_row + 1):
#         row = list(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
        
#         # Extract block name from column A
#         block_name = row[KRA_COLUMNS['BLOCK']]
        
#         # Skip empty rows
#         if not block_name or pd.isna(block_name):
#             continue
        
#         # Stop if we hit another header row or non-block data
#         block_name_str = str(block_name).strip()
#         if any(month.lower() in block_name_str.lower() for month in ['january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december']):
#             logger.info(f"Stopping at row {row_idx}: Found header-like content '{block_name_str}'")
#             break
        
#         # Skip if block name is just "Activity" or other single-word non-block names
#         if block_name_str.lower() in ['activity', 'activities', 'milestone', 'milestones', 'target', 'targets', 'blocks']:
#             logger.info(f"Skipping non-block row: '{block_name_str}'")
#             continue
            
#         logger.info(f"Processing block: {block_name}")
        
#         targets[block_name] = {}
        
#         # Process each month
#         for i, month in enumerate(MONTHS):
#             month_year = MONTH_YEARS[month]
            
#             # Calculate column indices for this month
#             if i == 0:
#                 activity_col = KRA_COLUMNS['MONTH1_ACTIVITY']
#                 target_col = KRA_COLUMNS['MONTH1_TARGET']
#             elif i == 1:
#                 activity_col = KRA_COLUMNS['MONTH2_ACTIVITY']
#                 target_col = KRA_COLUMNS['MONTH2_TARGET']
#             elif i == 2:
#                 activity_col = KRA_COLUMNS['MONTH3_ACTIVITY']
#                 target_col = KRA_COLUMNS['MONTH3_TARGET']
#             else:
#                 logger.warning(f"More than 3 months found, skipping month {month}")
#                 continue
            
#             # Extract activity name and target %
#             activity_name = row[activity_col] if activity_col < len(row) else None
#             target_percent = row[target_col] if target_col < len(row) else None
            
#             # Validate target_percent is a number, not text
#             validated_target_percent = None
#             if target_percent and not pd.isna(target_percent):
#                 try:
#                     validated_target_percent = float(target_percent)
#                 except (ValueError, TypeError):
#                     logger.warning(f"Invalid target % value for {block_name} - {month}: '{target_percent}' (not a number)")
#                     validated_target_percent = None
            
#             # Store both activity and target % for this month
#             targets[block_name][month] = {
#                 'activity': activity_name if activity_name and not pd.isna(activity_name) else None,
#                 'target_percent': validated_target_percent
#             }
            
#             logger.info(f"  {month}: Activity='{activity_name}', Target %={validated_target_percent}")
    
#     kra_wb.close()
#     return targets

# def find_activity_completion_in_tracker(tracker_wb, sheet_name, target_activity):
#     """
#     Search for an activity in the tracker and return its % complete value.
    
#     Args:
#         tracker_wb: Loaded tracker workbook
#         sheet_name: Sheet name to search in
#         target_activity: Activity name to find
    
#     Returns:
#         Float value of % complete, or None if not found
#     """
#     if sheet_name not in tracker_wb.sheetnames:
#         logger.warning(f"Sheet '{sheet_name}' not found in tracker")
#         return None
    
#     ws = tracker_wb[sheet_name]
    
#     # Search through rows starting from row 3 (skip headers)
#     for row_idx in range(2, ws.max_row + 1):
#         row = list(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
        
#         # Get activity name from column G (index 6)
#         if len(row) > TRACKER_COLUMNS['ACTIVITY_NAME']:
#             activity_name = row[TRACKER_COLUMNS['ACTIVITY_NAME']]
            
#             if activity_name and str(activity_name).strip().lower() == str(target_activity).strip().lower():
#                 # Found the activity, get % complete from column M (index 12)
#                 if len(row) > TRACKER_COLUMNS['PERCENT_COMPLETE']:
#                     percent_complete = row[TRACKER_COLUMNS['PERCENT_COMPLETE']]
#                     try:
#                         return float(percent_complete) if percent_complete is not None else 0.0
#                     except (ValueError, TypeError):
#                         logger.warning(f"Could not convert % complete to float for activity '{activity_name}': {percent_complete}")
#                         return 0.0
    
#     logger.debug(f"Activity '{target_activity}' not found in sheet '{sheet_name}'")
#     return None

# def get_wcc_progress_from_tracker_all_months(cos, targets):
#     """
#     Extract progress data for all months with UPDATED LOGIC:
#     - Weightage = 100 for all blocks
#     - Weighted % = (% Complete / Weightage) * 100 for each month
#     - If tracker % > target %, display 100%
#     - Responsible and Delay columns shown once at the end
#     """
#     global TRACKER_PATHS, LOADED_TRACKERS
    
#     # Load all trackers into memory
#     for month in MONTHS:
#         tracker_path = TRACKER_PATHS.get(month)
#         if tracker_path and month not in LOADED_TRACKERS:
#             logger.info(f"Loading tracker for {month}: {tracker_path}")
#             tracker_bytes = download_file_bytes(cos, tracker_path)
#             LOADED_TRACKERS[month] = load_workbook(BytesIO(tracker_bytes), data_only=True)
    
#     # Build dataframe
#     data = []
#     milestone_counter = 1
    
#     for block_name, month_targets in targets.items():
#         # Get the tracker sheet name for this block
#         tracker_sheet = BLOCK_MAPPING.get(block_name)
        
#         if not tracker_sheet:
#             logger.warning(f"No tracker sheet mapping found for block: {block_name}")
#             continue
        
#         row_data = {
#             'Milestone': f'Milestone-{milestone_counter:02d}',
#             'Block': block_name
#         }
        
#         # Determine the target to complete by the end of quarter
#         last_month = MONTHS[-1]
#         last_month_target_info = month_targets.get(last_month, {})
#         target_to_complete = last_month_target_info.get('activity', 'N/A')
#         row_data[f'Target to be complete by {last_month}-{MONTH_YEARS[last_month]}'] = target_to_complete
        
#         # Process each month
#         for month in MONTHS:
#             month_year = MONTH_YEARS[month]
            
#             # Get target info for this month
#             target_info = month_targets.get(month, {})
#             target_activity = target_info.get('activity')
#             target_percent = target_info.get('target_percent')
            
#             # Column names for this month
#             target_col = f'Target - {month}-{month_year}'
#             status_col = f'% work done- {month} Status'
#             achieved_col = f'Achieved- {month} {month_year}'
#             weightage_col = f'Weightage- {month}'
#             weighted_pct_col = f'Weighted %- {month}'
            
#             row_data[target_col] = target_activity if target_activity else 'No target'
            
#             # Check if tracker is available for this month
#             if month not in LOADED_TRACKERS:
#                 # Leave blank if tracker not available
#                 row_data[status_col] = ''
#                 row_data[achieved_col] = ''
#                 row_data[weightage_col] = ''
#                 row_data[weighted_pct_col] = ''
#                 continue
            
#             # If no target activity, mark as N/A
#             if not target_activity:
#                 row_data[status_col] = 'N/A'
#                 row_data[achieved_col] = 'No target for this month'
#                 row_data[weightage_col] = ''
#                 row_data[weighted_pct_col] = ''
#                 continue

#             # Check if target is explicitly "No target"
#             if str(target_activity).strip().lower() == 'no target':
#                 row_data[status_col] = '100%'
#                 row_data[achieved_col] = 'No target specified'
                
#                 # Calculate Weightage
#                 row_data[weightage_col] = 100
#                 row_data[weighted_pct_col] = '100%'
#                 continue
            
#             # Fetch completed % from tracker
#             tracker_wb = LOADED_TRACKERS[month]
#             completed_percent = find_activity_completion_in_tracker(tracker_wb, tracker_sheet, target_activity)
            
#             # Determine status based on comparison
#             if completed_percent is None:
#                 # Activity not found in tracker
#                 row_data[status_col] = '0%'
#                 row_data[achieved_col] = 'Activity not found in tracker'
#                 status_percent = 0.0
#             elif (target_percent is None or target_percent == 0) and completed_percent > 0:
#                  # No target or 0 target, but tracker has progress -> 100%
#                  row_data[status_col] = '100%'
#                  row_data[achieved_col] = f'Target achieved (Tracker: {completed_percent*100:.0f}%)'
#                  status_percent = 1.0
#             elif target_percent is not None and completed_percent > target_percent:
#                 # Tracker % > Target % -> display 100%
#                 row_data[status_col] = '100%'
#                 row_data[achieved_col] = f'Target exceeded ({completed_percent*100:.0f}% > {target_percent*100:.0f}%)'
#                 status_percent = 1.0
#             elif target_percent is not None and completed_percent == target_percent:
#                 # Target % matches completed %
#                 row_data[status_col] = '100%'
#                 row_data[achieved_col] = f'Target achieved ({completed_percent*100:.0f}% complete)'
#                 status_percent = 1.0
#             else:
#                 # Display completed % from tracker
#                 status_percent = completed_percent
#                 row_data[status_col] = f'{completed_percent*100:.0f}%'
                
#                 if target_percent is not None and completed_percent >= target_percent:
#                     row_data[achieved_col] = f'{completed_percent*100:.0f}% completed (Target: {target_percent*100:.0f}%)'
#                 elif target_percent is not None:
#                     row_data[achieved_col] = f'{completed_percent*100:.0f}% completed (Target: {target_percent*100:.0f}%)'
#                 else:
#                     row_data[achieved_col] = f'{completed_percent*100:.0f}% completed'
            
#             # Calculate Weightage and Weighted % for this month
#             weightage = 100  # Always 100
#             weighted_pct = (status_percent * 100 / weightage) * 100 if weightage > 0 else 0
#             row_data[weightage_col] = weightage
#             row_data[weighted_pct_col] = f'{weighted_pct:.0f}%'
        
#         # Add Responsible and Delay columns once at the end
#         row_data['Responsible'] = ''
#         row_data['Delay Reason'] = ''
        
#         data.append(row_data)
#         milestone_counter += 1
    
#     # Add summary row with average Weighted % for each month
#     summary_row = {'Milestone': 'AVERAGE WEIGHTED %', 'Block': ''}
#     summary_row[f'Target to be complete by {MONTHS[-1]}-{MONTH_YEARS[MONTHS[-1]]}'] = ''
    
#     for month in MONTHS:
#         month_year = MONTH_YEARS[month]
        
#         # Calculate average of Weighted % for this month
#         weighted_values = []
#         for row in data:
#             weighted_val = row.get(f'Weighted %- {month}', '')
#             if weighted_val and weighted_val != '':
#                 try:
#                     val = float(str(weighted_val).replace('%', ''))
#                     weighted_values.append(val)
#                 except (ValueError, TypeError):
#                     pass
        
#         # Calculate average
#         if weighted_values:
#             avg_weighted = sum(weighted_values) / len(weighted_values)
#             summary_row[f'Weighted %- {month}'] = f'{avg_weighted:.1f}%'
#         else:
#             summary_row[f'Weighted %- {month}'] = ''
        
#         # Leave other columns blank for summary row
#         summary_row[f'Target - {month}-{month_year}'] = ''
#         summary_row[f'% work done- {month} Status'] = ''
#         summary_row[f'Achieved- {month} {month_year}'] = ''
#         summary_row[f'Weightage- {month}'] = ''
    
#     summary_row['Responsible'] = ''
#     summary_row['Delay Reason'] = ''
    
#     data.append(summary_row)
    
#     # Create DataFrame
#     df = pd.DataFrame(data)
    
#     # Reorder columns to match format
#     column_order = ['Milestone', 'Block', f'Target to be complete by {MONTHS[-1]}-{MONTH_YEARS[MONTHS[-1]]}']
    
#     for month in MONTHS:
#         month_year = MONTH_YEARS[month]
#         column_order.extend([
#             f'Target - {month}-{month_year}',
#             f'% work done- {month} Status',
#             f'Achieved- {month} {month_year}',
#             f'Weightage- {month}',
#             f'Weighted %- {month}'
#         ])
    
#     column_order.extend(['Responsible', 'Delay Reason'])
    
#     df = df[column_order]
    
#     return df

# def apply_manual_overrides(df):
#     """
#     Apply any manual overrides if needed
#     This function can be customized based on business rules
#     """
#     logger.info("Applying manual overrides (if any)...")
#     return df

# # -----------------------------------------------------------------------------
# # REPORT GENERATION
# # -----------------------------------------------------------------------------

# def write_wcc_excel_report_consolidated(df, filename):
#     """Generate Excel report with proper formatting matching the format file"""
    
#     logger.info(f'Generating consolidated report: {filename}')
    
#     wb = Workbook()
#     ws = wb.active
#     ws.title = 'WCC Progress'
    
#     # Define styles
#     header_font = Font(name='Calibri', size=11, bold=True)
#     normal_font = Font(name='Calibri', size=10)
#     date_font = Font(name='Calibri', size=10, bold=True)
#     summary_font = Font(name='Calibri', size=11, bold=True)
    
#     center = Alignment(horizontal='center', vertical='center', wrap_text=True)
#     left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
#     border = Border(
#         left=Side(style='thin'),
#         right=Side(style='thin'),
#         top=Side(style='thin'),
#         bottom=Side(style='thin')
#     )
    
#     light_grey_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
#     light_blue_fill = PatternFill(start_color='B4C7E7', end_color='B4C7E7', fill_type='solid')
#     summary_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    
#     # Title row
#     ws.merge_cells('A1:T1')
#     ws['A1'] = 'Wave City Club- Progress Against Milestones'
#     ws['A1'].font = Font(name='Calibri', size=14, bold=True)
#     ws['A1'].alignment = center
    
#     # Date row
#     current_date = datetime.now().strftime('%d-%m-%Y')
#     ws.merge_cells('A2:T2')
#     ws['A2'] = f'Report Generated on: {current_date}'
#     ws['A2'].font = date_font
#     ws['A2'].alignment = center
    
#     # Month info row
#     month_info = ', '.join([f'{m} {MONTH_YEARS[m]}' for m in MONTHS])
#     ws.merge_cells('A3:T3')
#     ws['A3'] = f'Months Covered: {month_info}'
#     ws['A3'].font = date_font
#     ws['A3'].alignment = center
    
#     # Empty row
#     ws.merge_cells('A4:T4')
    
#     # Write dataframe starting at row 5
#     for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=5):
#         for c_idx, value in enumerate(row, start=1):
#             cell = ws.cell(row=r_idx, column=c_idx, value=value)
    
#     # Style header row (row 5)
#     header_row = 5
#     for cell in ws[header_row]:
#         cell.font = header_font
#         cell.alignment = center
#         cell.border = border
#         cell.fill = light_grey_fill
    
#     # Style data rows
#     data_start = 6
#     summary_row_idx = ws.max_row
#     data_end = summary_row_idx - 1
    
#     for row_num in range(data_start, data_end + 1):
#         for col_num in range(1, len(df.columns) + 1):
#             cell = ws.cell(row=row_num, column=col_num)
#             cell.font = normal_font
#             cell.border = border
            
#             # Alignment based on column type
#             if col_num in [1, 2, 3] or 'Target' in str(ws.cell(row=header_row, column=col_num).value or ''):
#                 cell.alignment = left
#             else:
#                 cell.alignment = center
    
#     # Style summary row (last row)
#     for col_num in range(1, len(df.columns) + 1):
#         cell = ws.cell(row=summary_row_idx, column=col_num)
#         cell.font = summary_font
#         cell.border = border
#         cell.fill = summary_fill
#         cell.alignment = center
    
#     # Dynamic column width adjustment
#     for col_num in range(1, len(df.columns) + 1):
#         col_letter = get_column_letter(col_num)
        
#         max_length = 0
#         for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=col_num, max_col=col_num):
#             for cell in row:
#                 if cell.value:
#                     max_length = max(max_length, len(str(cell.value)))
        
#         calculated_width = min(max(max_length + 2, 8), 15)
#         ws.column_dimensions[col_letter].width = calculated_width
    
#     # Set row heights
#     ws.row_dimensions[1].height = 25
#     ws.row_dimensions[2].height = 20
#     ws.row_dimensions[3].height = 20
#     for i in range(5, ws.max_row + 1):
#         ws.row_dimensions[i].height = 25
#     ws.row_dimensions[summary_row_idx].height = 30  # Make summary row taller
    
#     wb.save(filename)
#     logger.info(f'Report saved to {filename}')

# def get_unique_filename(base_name):
#     """If file exists, append (1), (2), etc."""
#     if not os.path.exists(base_name):
#         return base_name

#     name, ext = os.path.splitext(base_name)
#     counter = 1
#     new_name = f"{name}({counter}){ext}"
#     while os.path.exists(new_name):
#         counter += 1
#         new_name = f"{name}({counter}){ext}"
#     return new_name

# # -----------------------------------------------------------------------------
# # MAIN FUNCTION
# # -----------------------------------------------------------------------------

# def main():
#     """Main execution function for quarterly report generation"""
#     logger.info("=== STARTING WAVE CITY CLUB QUARTERLY REPORT GENERATION (UPDATED LOGIC) ===")
    
#     try:
#         # Initialize COS client
#         cos = init_cos()
        
#         # Step 1: Get latest KRA file from Milestone/ folder
#         logger.info("\n=== STEP 1: Finding Latest KRA File ===")
#         kra_file = get_latest_kra_file(cos)
        
#         if not kra_file:
#             logger.error("❌ Failed to find KRA file")
#             return
        
#         logger.info(f"✅ Using KRA file: {kra_file}")
        
#         # Step 2: Extract months from KRA filename and setup
#         logger.info("\n=== STEP 2: Setting Up Quarterly Months ===")
#         if not setup_quarterly_months(os.path.basename(kra_file)):
#             logger.error("❌ Failed to setup quarterly months")
#             return
        
#         logger.info(f"✅ Quarterly months: {MONTHS}")
#         logger.info(f"✅ Month-Year mapping: {MONTH_YEARS}")
        
#         # Step 3: Find appropriate trackers for each month from Wave City Club/ folder
#         logger.info("\n=== STEP 3: Finding Trackers for Each Month ===")
#         global TRACKER_PATHS
        
#         for month in MONTHS:
#             month_year = MONTH_YEARS[month]
#             tracker_path = get_tracker_for_month(cos, month, month_year)
            
#             if tracker_path:
#                 TRACKER_PATHS[month] = tracker_path
#                 logger.info(f"✅ {month} {month_year}: {tracker_path}")
#             else:
#                 logger.warning(f"⚠️ {month} {month_year}: No tracker found - column will be blank in report")
#                 TRACKER_PATHS[month] = None
        
#         # Continue even if no trackers found
#         logger.info(f"Found trackers for {sum(1 for v in TRACKER_PATHS.values() if v)} out of {len(MONTHS)} months")
        
#         # Step 4: Get targets from KRA file (Club sheet)
#         logger.info("\n=== STEP 4: Extracting Targets from KRA File ===")
#         targets = get_wcc_targets_from_kra(cos)
        
#         if not targets:
#             logger.error("❌ Failed to extract targets from KRA file")
#             return
        
#         logger.info(f"✅ Extracted targets for {len(targets)} blocks")
        
#         # Step 5: Extract progress data for all months with NEW LOGIC
#         logger.info("\n=== STEP 5: Extracting Progress Data from Trackers (NEW LOGIC) ===")
#         df = get_wcc_progress_from_tracker_all_months(cos, targets)
        
#         if df.empty:
#             logger.error("❌ Failed to generate progress data")
#             return
        
#         logger.info(f"✅ Generated progress data for {len(df)-1} milestones + 1 summary row")
        
#         # Step 5.5: Apply manual overrides
#         logger.info("\n=== STEP 5.5: Applying Manual Overrides ===")
#         df = apply_manual_overrides(df)
        
#         # Step 6: Generate Excel report
#         logger.info("\n=== STEP 6: Generating Excel Report ===")
#         current_date_for_filename = datetime.now().strftime('%d-%m-%Y')
        
#         # Create filename with quarter months
#         quarter_str = "_".join(MONTHS)
#         base_filename = f"Wave_City_Club_Milestone_Report_{quarter_str}_{current_date_for_filename}.xlsx"
#         filename = get_unique_filename(base_filename)
        
#         write_wcc_excel_report_consolidated(df, filename)
        
#         logger.info("\n=== WAVE CITY CLUB QUARTERLY REPORT GENERATION COMPLETE ===")
#         logger.info(f"✅ Report saved as: {filename}")
        
#         # Log summary
#         logger.info("\n=== REPORT SUMMARY ===")
#         logger.info(f"  Quarter Months: {', '.join([f'{m} {MONTH_YEARS[m]}' for m in MONTHS])}")
#         logger.info(f"  KRA File: {os.path.basename(kra_file)}")
#         logger.info(f"  Total Blocks: {len(targets)}")
#         logger.info(f"  Trackers Used:")
#         for month in MONTHS:
#             tracker = TRACKER_PATHS.get(month)
#             if tracker:
#                 tracker_date = extract_date_from_filename(os.path.basename(tracker))
#                 logger.info(f"    - {month} {MONTH_YEARS[month]}: {os.path.basename(tracker)} (dated {tracker_date.strftime('%d-%m-%Y') if tracker_date else 'Unknown'})")
#             else:
#                 logger.info(f"    - {month} {MONTH_YEARS[month]}: Not Available (column will be blank)")
#         logger.info(f"  Total Milestones: {len(df)-1} + 1 summary row")
        
#     except Exception as e:
#         logger.error(f"❌ Error in main execution: {e}")
#         import traceback
#         logger.error(traceback.format_exc())
#         raise

# if __name__ == "__main__":
#     main()


















































# # import os
# # import re
# # import logging
# # from io import BytesIO
# # from datetime import datetime
# # from dateutil.relativedelta import relativedelta

# # import pandas as pd
# # from openpyxl import Workbook, load_workbook
# # from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
# # from openpyxl.utils import get_column_letter
# # from openpyxl.utils.dataframe import dataframe_to_rows
# # from dotenv import load_dotenv
# # import ibm_boto3
# # from ibm_botocore.client import Config

# # # -----------------------------------------------------------------------------
# # # CONFIG / CONSTANTS
# # # -----------------------------------------------------------------------------
# # load_dotenv()
# # logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(levelname)s] %(message)s')
# # logger = logging.getLogger(__name__)

# # # Validate required environment variables
# # required = {
# #     'COS_API_KEY': os.getenv('COS_API_KEY'),
# #     'COS_SERVICE_INSTANCE_CRN': os.getenv('COS_SERVICE_INSTANCE_CRN'),
# #     'COS_ENDPOINT': os.getenv('COS_ENDPOINT'),
# #     'COS_BUCKET_NAME': os.getenv('COS_BUCKET_NAME'),
# # }
# # missing = [k for k, v in required.items() if not v]
# # if missing:
# #     logger.error(f"Missing required environment variables: {', '.join(missing)}")
# #     raise SystemExit(1)

# # COS_API_KEY     = required['COS_API_KEY']
# # COS_CRN         = required['COS_SERVICE_INSTANCE_CRN']
# # COS_ENDPOINT    = required['COS_ENDPOINT']
# # BUCKET          = required['COS_BUCKET_NAME']

# # # Folder paths in COS
# # KRA_FOLDER = "Milestone/"
# # TRACKER_FOLDER = "Wave City Club/"

# # # Dynamic KRA file path
# # WCC_KRA_KEY = None

# # # Dynamic tracker paths
# # TRACKER_PATHS = {}  # Maps month names to tracker file paths
# # LOADED_TRACKERS = {}  # Cache loaded workbooks

# # # Dynamic months and years
# # MONTHS = []
# # MONTH_YEARS = {}  # Maps month name to year
# # TARGET_END_MONTH = None
# # TARGET_END_YEAR = None

# # # Quarterly definitions (for reference)
# # QUARTERLY_GROUPS = [
# #     ['June', 'July', 'August'],           # Q1
# #     ['September', 'October', 'November'], # Q2
# #     ['December', 'January', 'February'],  # Q3
# #     ['March', 'April', 'May']             # Q4
# # ]

# # # Month shift logic: Report Month -> Tracker Month
# # MONTH_SHIFT = {
# #     'June': 'July',
# #     'July': 'August',
# #     'August': 'September',
# #     'September': 'October',
# #     'October': 'November',
# #     'November': 'December',
# #     'December': 'January',
# #     'January': 'February',
# #     'February': 'March',
# #     'March': 'April',
# #     'April': 'May',
# #     'May': 'June'
# # }

# # # Block mapping from KRA to tracker sheets
# # BLOCK_MAPPING = {
# #     'Block 1 (B1) Banquet Hall': 'B1 Banket Hall & Finedine ',
# #     'Fine Dine': 'B1 Banket Hall & Finedine ',
# #     'Block 5 (B5) Admin + Member Lounge+Creche+Av Room + Surveillance Room +Toilets': 'B5',
# #     'Block 6 (B6) Toilets': 'B6',
# #     'Block 7(B7) Indoor Sports': 'B7',
# #     'Block 9 (B9) Spa & Saloon': 'B9',
# #     'Block 8 (B8) Squash Court': 'B8',
# #     'Block 2 & 3 (B2 & B3) Cafe & Bar': 'B2 & B3',
# #     'Block 4 (B4) Indoor Swimming Pool Changing Room & Toilets': 'B4',
# #     'Block 11 (B11) Guest House': 'B11',
# #     'Block 10 (B10) Gym': 'B10'
# # }

# # # KRA Column structure (0-indexed)
# # KRA_COLUMNS = {
# #     'BLOCK': 0,          # Column A: Block name
# #     'MONTH1_ACTIVITY': 1,  # Column B: First month activity
# #     'MONTH1_TARGET': 2,    # Column C: First month target %
# #     'MONTH2_ACTIVITY': 3,  # Column D: Second month activity
# #     'MONTH2_TARGET': 4,    # Column E: Second month target %
# #     'MONTH3_ACTIVITY': 5,  # Column F: Third month activity
# #     'MONTH3_TARGET': 6,    # Column G: Third month target %
# # }

# # # Tracker column positions (0-indexed)
# # TRACKER_COLUMNS = {
# #     'ACTIVITY_NAME': 6,    # Column G: Activity Name
# #     'PERCENT_COMPLETE': 12  # Column M: % Complete
# # }

# # # -----------------------------------------------------------------------------
# # # COS HELPERS
# # # -----------------------------------------------------------------------------

# # def init_cos():
# #     return ibm_boto3.client(
# #         's3',
# #         ibm_api_key_id=COS_API_KEY,
# #         ibm_service_instance_id=COS_CRN,
# #         config=Config(signature_version='oauth'),
# #         endpoint_url=COS_ENDPOINT,
# #     )

# # def download_file_bytes(cos, key):
# #     if not key:
# #         raise ValueError("File key cannot be None or empty")
# #     obj = cos.get_object(Bucket=BUCKET, Key=key)
# #     return obj['Body'].read()

# # def list_files_in_folder(cos, folder_prefix):
# #     """List all files in a specific folder (prefix) in the COS bucket"""
# #     try:
# #         response = cos.list_objects_v2(Bucket=BUCKET, Prefix=folder_prefix)
# #         files = []
# #         if 'Contents' in response:
# #             for obj in response['Contents']:
# #                 if not obj['Key'].endswith('/'):
# #                     files.append(obj['Key'])
# #         return files
# #     except Exception as e:
# #         logger.error(f"Error listing files in folder {folder_prefix}: {e}")
# #         return []

# # def extract_date_from_filename(filename):
# #     """Extract date from filename in format (dd-mm-yyyy)"""
# #     pattern = r'\((\d{2}-\d{2}-\d{4})\)'
# #     match = re.search(pattern, filename)
# #     if match:
# #         date_str = match.group(1)
# #         try:
# #             return datetime.strptime(date_str, '%d-%m-%Y')
# #         except ValueError:
# #             logger.warning(f"Could not parse date {date_str} from filename {filename}")
# #             return None
# #     return None

# # def get_month_name(month_num):
# #     """Convert month number to month name"""
# #     months = {
# #         1: "January", 2: "February", 3: "March", 4: "April",
# #         5: "May", 6: "June", 7: "July", 8: "August", 
# #         9: "September", 10: "October", 11: "November", 12: "December"
# #     }
# #     return months.get(month_num, "Unknown")

# # def get_month_number(month_name):
# #     """Convert month name to month number"""
# #     months = {
# #         "January": 1, "February": 2, "March": 3, "April": 4,
# #         "May": 5, "June": 6, "July": 7, "August": 8, 
# #         "September": 9, "October": 10, "November": 11, "December": 12
# #     }
# #     return months.get(month_name, 1)

# # def get_latest_kra_file(cos):
# #     """Get the latest KRA milestone file from the Milestone/ folder"""
# #     global WCC_KRA_KEY
    
# #     logger.info("=== FINDING LATEST KRA MILESTONE FILE ===")
    
# #     # List all files in Milestone folder
# #     all_files = list_files_in_folder(cos, KRA_FOLDER)
# #     logger.info(f"Found {len(all_files)} files in {KRA_FOLDER} folder")
    
# #     # Pattern to match KRA milestone files
# #     kra_pattern = r'KRA Milestones for.*\.xlsx$'
    
# #     matching_files = []
    
# #     for file_path in all_files:
# #         filename = os.path.basename(file_path)
# #         if re.search(kra_pattern, filename, re.IGNORECASE):
# #             logger.info(f"Found KRA file: {filename}")
# #             matching_files.append(file_path)
    
# #     if matching_files:
# #         # Sort by filename to get the latest
# #         latest_file = sorted(matching_files)[-1]
# #         WCC_KRA_KEY = latest_file
# #         logger.info(f"✅ Latest KRA file: {WCC_KRA_KEY}")
# #     else:
# #         logger.error(f"❌ No KRA milestone files found in {KRA_FOLDER} folder!")
# #         WCC_KRA_KEY = None
    
# #     return WCC_KRA_KEY

# # def extract_months_from_kra_filename(filename):
# #     """Extract quarter months from KRA filename"""
# #     months_pattern = r'for\s+((?:January|February|March|April|May|June|July|August|September|October|November|December)(?:\s+(?:January|February|March|April|May|June|July|August|September|October|November|December))*)\s+(\d{4})'
    
# #     match = re.search(months_pattern, filename, re.IGNORECASE)
# #     if match:
# #         months_str = match.group(1)
# #         year = int(match.group(2))
        
# #         # Extract individual months
# #         month_names = re.findall(r'January|February|March|April|May|June|July|August|September|October|November|December', 
# #                                  months_str, re.IGNORECASE)
        
# #         return [m.capitalize() for m in month_names], year
    
# #     return None, None

# # def get_tracker_for_month(cos, report_month, month_year):
# #     """
# #     Get tracker file for a specific report month using the month-shift logic.
    
# #     Args:
# #         cos: COS client
# #         report_month: The month shown in the report (e.g., "June")
# #         month_year: Year of the report month
    
# #     Returns:
# #         Path to tracker file or None
# #     """
# #     # Get the tracker month based on shift logic
# #     tracker_month_name = MONTH_SHIFT.get(report_month)
    
# #     if not tracker_month_name:
# #         logger.warning(f"No month shift mapping found for {report_month}")
# #         return None
    
# #     # Calculate the tracker year (handle year rollover)
# #     tracker_month_num = get_month_number(tracker_month_name)
# #     report_month_num = get_month_number(report_month)
    
# #     tracker_year = month_year
# #     # If tracker month is earlier in the year than report month, it's next year
# #     if tracker_month_num < report_month_num:
# #         tracker_year += 1
    
# #     logger.info(f"Looking for tracker for Report: {report_month} {month_year} → Tracker: {tracker_month_name} {tracker_year}")
    
# #     # List all files in Wave City Club folder
# #     all_files = list_files_in_folder(cos, TRACKER_FOLDER)
    
# #     # Pattern to match Structure Work tracker files
# #     tracker_pattern = r'Structure.*Work.*Tracker.*Wave.*City.*Club.*\.xlsx$'
    
# #     matching_files = []
    
# #     for file_path in all_files:
# #         filename = os.path.basename(file_path)
# #         if re.search(tracker_pattern, filename, re.IGNORECASE):
# #             # Extract date from filename
# #             file_date = extract_date_from_filename(filename)
# #             if file_date and file_date.month == tracker_month_num and file_date.year == tracker_year:
# #                 matching_files.append((file_path, file_date))
# #                 logger.info(f"Found matching tracker: {filename} (dated {file_date.strftime('%d-%m-%Y')})")
    
# #     if matching_files:
# #         # Sort by date and get the latest
# #         matching_files.sort(key=lambda x: x[1], reverse=True)
# #         latest_tracker = matching_files[0][0]
# #         logger.info(f"✅ Selected tracker: {os.path.basename(latest_tracker)}")
# #         return latest_tracker
# #     else:
# #         logger.warning(f"⚠️ No tracker found for {report_month} {month_year} (looking for {tracker_month_name} {tracker_year} dated files)")
# #         return None

# # def setup_quarterly_months(kra_filename):
# #     """Setup the quarterly months and years based on KRA filename"""
# #     global MONTHS, MONTH_YEARS, TARGET_END_MONTH, TARGET_END_YEAR
    
# #     months, year = extract_months_from_kra_filename(kra_filename)
    
# #     if not months or not year:
# #         logger.error(f"Could not extract months from KRA filename: {kra_filename}")
# #         return False
    
# #     MONTHS = months
# #     TARGET_END_MONTH = MONTHS[-1]
# #     TARGET_END_YEAR = year
    
# #     # Assign years to each month (handle cross-year quarters)
# #     for i, month in enumerate(MONTHS):
# #         month_num = get_month_number(month)
        
# #         # Handle December-January-February quarter
# #         if i > 0:
# #             prev_month_num = get_month_number(MONTHS[i-1])
# #             # If current month number is less than previous, we've crossed into new year
# #             if month_num < prev_month_num:
# #                 year += 1
        
# #         MONTH_YEARS[month] = year
    
# #     return True

# # def find_club_sheet(wb):
# #     """Find the sheet with 'Club' in its name"""
# #     for sheet_name in wb.sheetnames:
# #         if 'club' in sheet_name.lower():
# #             logger.info(f"Found Club sheet: {sheet_name}")
# #             return sheet_name
    
# #     logger.warning("No sheet with 'Club' in name found. Available sheets: " + ", ".join(wb.sheetnames))
# #     return None

# # # -----------------------------------------------------------------------------
# # # DATA EXTRACTION
# # # -----------------------------------------------------------------------------

# # def get_wcc_targets_from_kra(cos):
# #     """
# #     Extract targets from KRA milestone file with updated logic:
# #     - Look for sheet with "Club" in its name
# #     - For each activity, fetch both the target activity name AND target %
# #     - Store both values for comparison with tracker data
# #     """
# #     global WCC_KRA_KEY
    
# #     if not WCC_KRA_KEY:
# #         raise ValueError("WCC_KRA_KEY is not set. Call get_latest_kra_file() first.")
    
# #     logger.info(f"Reading KRA file: {WCC_KRA_KEY}")
    
# #     # Download and load KRA file
# #     kra_bytes = download_file_bytes(cos, WCC_KRA_KEY)
# #     kra_wb = load_workbook(BytesIO(kra_bytes), data_only=True)
    
# #     # Find the Club sheet
# #     club_sheet_name = find_club_sheet(kra_wb)
# #     if not club_sheet_name:
# #         # Fallback to active sheet
# #         logger.warning("Using active sheet as fallback")
# #         ws = kra_wb.active
# #     else:
# #         ws = kra_wb[club_sheet_name]
    
# #     targets = {}
    
# #     # Find header row (look for "Blocks" in column A)
# #     header_row = None
# #     for row_idx in range(1, min(10, ws.max_row + 1)):
# #         cell_value = ws.cell(row=row_idx, column=1).values
# #         if cell_value and 'block' in str(cell_value).lower():
# #             header_row = row_idx
# #             logger.info(f"Found header row at row {header_row}")
# #             break
    
# #     if not header_row:
# #         logger.error("Could not find header row with 'Blocks'")
# #         kra_wb.close()
# #         return targets
    
# #     # Start from row after header
# #     data_start_row = header_row + 1
    
# #     for row_idx in range(data_start_row, ws.max_row + 1):
# #         row = list(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
        
# #         # Extract block name from column A
# #         block_name = row[KRA_COLUMNS['BLOCK']]
        
# #         # Skip empty rows
# #         if not block_name or pd.isna(block_name):
# #             continue
        
# #         # Stop if we hit another header row or non-block data
# #         block_name_str = str(block_name).strip()
# #         if any(month.lower() in block_name_str.lower() for month in ['january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december']):
# #             logger.info(f"Stopping at row {row_idx}: Found header-like content '{block_name_str}'")
# #             break
        
# #         # Skip if block name is just "Activity" or other single-word non-block names
# #         if block_name_str.lower() in ['activity', 'activities', 'milestone', 'milestones', 'target', 'targets', 'blocks']:
# #             logger.info(f"Skipping non-block row: '{block_name_str}'")
# #             continue
            
# #         logger.info(f"Processing block: {block_name}")
        
# #         targets[block_name] = {}
        
# #         # Process each month
# #         for i, month in enumerate(MONTHS):
# #             month_year = MONTH_YEARS[month]
            
# #             # Calculate column indices for this month
# #             if i == 0:
# #                 activity_col = KRA_COLUMNS['MONTH1_ACTIVITY']
# #                 target_col = KRA_COLUMNS['MONTH1_TARGET']
# #             elif i == 1:
# #                 activity_col = KRA_COLUMNS['MONTH2_ACTIVITY']
# #                 target_col = KRA_COLUMNS['MONTH2_TARGET']
# #             elif i == 2:
# #                 activity_col = KRA_COLUMNS['MONTH3_ACTIVITY']
# #                 target_col = KRA_COLUMNS['MONTH3_TARGET']
# #             else:
# #                 logger.warning(f"More than 3 months found, skipping month {month}")
# #                 continue
            
# #             # Extract activity name and target %
# #             activity_name = row[activity_col] if activity_col < len(row) else None
# #             target_percent = row[target_col] if target_col < len(row) else None
            
# #             # Validate target_percent is a number, not text
# #             validated_target_percent = None
# #             if target_percent and not pd.isna(target_percent):
# #                 try:
# #                     validated_target_percent = float(target_percent)
# #                 except (ValueError, TypeError):
# #                     logger.warning(f"Invalid target % value for {block_name} - {month}: '{target_percent}' (not a number)")
# #                     validated_target_percent = None
            
# #             # Store both activity and target % for this month
# #             targets[block_name][month] = {
# #                 'activity': activity_name if activity_name and not pd.isna(activity_name) else None,
# #                 'target_percent': validated_target_percent
# #             }
            
# #             logger.info(f"  {month}: Activity='{activity_name}', Target %={validated_target_percent}")
    
# #     kra_wb.close()
# #     return targets

# # def find_activity_completion_in_tracker(tracker_wb, sheet_name, target_activity):
# #     """
# #     Search for an activity in the tracker and return its % complete value.
    
# #     Args:
# #         tracker_wb: Loaded tracker workbook
# #         sheet_name: Sheet name to search in
# #         target_activity: Activity name to find
    
# #     Returns:
# #         Float value of % complete, or None if not found
# #     """
# #     if sheet_name not in tracker_wb.sheetnames:
# #         logger.warning(f"Sheet '{sheet_name}' not found in tracker")
# #         return None
    
# #     ws = tracker_wb[sheet_name]
    
# #     # Search through rows starting from row 3 (skip headers)
# #     for row_idx in range(2, ws.max_row + 1):
# #         row = list(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
        
# #         # Get activity name from column G (index 6)
# #         if len(row) > TRACKER_COLUMNS['ACTIVITY_NAME']:
# #             activity_name = row[TRACKER_COLUMNS['ACTIVITY_NAME']]
            
# #             if activity_name and str(activity_name).strip().lower() == str(target_activity).strip().lower():
# #                 # Found the activity, get % complete from column M (index 12)
# #                 if len(row) > TRACKER_COLUMNS['PERCENT_COMPLETE']:
# #                     percent_complete = row[TRACKER_COLUMNS['PERCENT_COMPLETE']]
# #                     try:
# #                         return float(percent_complete) if percent_complete is not None else 0.0
# #                     except (ValueError, TypeError):
# #                         logger.warning(f"Could not convert % complete to float for activity '{activity_name}': {percent_complete}")
# #                         return 0.0
    
# #     logger.debug(f"Activity '{target_activity}' not found in sheet '{sheet_name}'")
# #     return None

# # def get_wcc_progress_from_tracker_all_months(cos, targets):
# #     """
# #     Extract progress data for all months with UPDATED LOGIC:
# #     - Weightage = 100 for all blocks
# #     - Weighted % = (% Complete / Weightage) * 100 for each month
# #     - If tracker % > target %, display 100%
# #     - Responsible and Delay columns shown once at the end
# #     """
# #     global TRACKER_PATHS, LOADED_TRACKERS
    
# #     # Load all trackers into memory
# #     for month in MONTHS:
# #         tracker_path = TRACKER_PATHS.get(month)
# #         if tracker_path and month not in LOADED_TRACKERS:
# #             logger.info(f"Loading tracker for {month}: {tracker_path}")
# #             tracker_bytes = download_file_bytes(cos, tracker_path)
# #             LOADED_TRACKERS[month] = load_workbook(BytesIO(tracker_bytes), data_only=True)
    
# #     # Build dataframe
# #     data = []
# #     milestone_counter = 1
    
# #     for block_name, month_targets in targets.items():
# #         # Get the tracker sheet name for this block
# #         tracker_sheet = BLOCK_MAPPING.get(block_name)
        
# #         if not tracker_sheet:
# #             logger.warning(f"No tracker sheet mapping found for block: {block_name}")
# #             continue
        
# #         row_data = {
# #             'Milestone': f'Milestone-{milestone_counter:02d}',
# #             'Block': block_name
# #         }
        
# #         # Determine the target to complete by the end of quarter
# #         last_month = MONTHS[-1]
# #         last_month_target_info = month_targets.get(last_month, {})
# #         target_to_complete = last_month_target_info.get('activity', 'N/A')
# #         row_data[f'Target to be complete by {last_month}-{MONTH_YEARS[last_month]}'] = target_to_complete
        
# #         # Process each month
# #         for month in MONTHS:
# #             month_year = MONTH_YEARS[month]
            
# #             # Get target info for this month
# #             target_info = month_targets.get(month, {})
# #             target_activity = target_info.get('activity')
# #             target_percent = target_info.get('target_percent')
            
# #             # Column names for this month
# #             target_col = f'Target - {month}-{month_year}'
# #             status_col = f'% work done- {month} Status'
# #             achieved_col = f'Achieved- {month} {month_year}'
# #             weightage_col = f'Weightage- {month}'
# #             weighted_pct_col = f'Weighted %- {month}'
            
# #             row_data[target_col] = target_activity if target_activity else 'No target'
            
# #             # Check if tracker is available for this month
# #             if month not in LOADED_TRACKERS:
# #                 # Leave blank if tracker not available
# #                 row_data[status_col] = ''
# #                 row_data[achieved_col] = ''
# #                 row_data[weightage_col] = ''
# #                 row_data[weighted_pct_col] = ''
# #                 continue
            
# #             # If no target activity, mark as N/A
# #             if not target_activity:
# #                 row_data[status_col] = 'N/A'
# #                 row_data[achieved_col] = 'No target for this month'
# #                 row_data[weightage_col] = ''
# #                 row_data[weighted_pct_col] = ''
# #                 continue
            
# #             # Fetch completed % from tracker
# #             tracker_wb = LOADED_TRACKERS[month]
# #             completed_percent = find_activity_completion_in_tracker(tracker_wb, tracker_sheet, target_activity)
            
# #             # Determine status based on comparison
# #             if completed_percent is None:
# #                 # Activity not found in tracker
# #                 row_data[status_col] = '0%'
# #                 row_data[achieved_col] = 'Activity not found in tracker'
# #                 status_percent = 0.0
# #             elif target_percent is not None and completed_percent > target_percent:
# #                 # Tracker % > Target % -> display 100%
# #                 row_data[status_col] = '100%'
# #                 row_data[achieved_col] = f'Target exceeded ({completed_percent*100:.0f}% > {target_percent*100:.0f}%)'
# #                 status_percent = 1.0
# #             elif target_percent is not None and completed_percent == target_percent:
# #                 # Target % matches completed %
# #                 row_data[status_col] = '100%'
# #                 row_data[achieved_col] = f'Target achieved ({completed_percent*100:.0f}% complete)'
# #                 status_percent = 1.0
# #             else:
# #                 # Display completed % from tracker
# #                 status_percent = completed_percent
# #                 row_data[status_col] = f'{completed_percent*100:.0f}%'
                
# #                 if target_percent is not None and completed_percent >= target_percent:
# #                     row_data[achieved_col] = f'{completed_percent*100:.0f}% completed (Target: {target_percent*100:.0f}%)'
# #                 elif target_percent is not None:
# #                     row_data[achieved_col] = f'{completed_percent*100:.0f}% completed (Target: {target_percent*100:.0f}%)'
# #                 else:
# #                     row_data[achieved_col] = f'{completed_percent*100:.0f}% completed'
            
# #             # Calculate Weightage and Weighted % for this month
# #             weightage = 100  # Always 100
# #             weighted_pct = (status_percent * 100 / weightage) * 100 if weightage > 0 else 0
# #             row_data[weightage_col] = weightage
# #             row_data[weighted_pct_col] = f'{weighted_pct:.0f}%'
        
# #         # Add Responsible and Delay columns once at the end
# #         row_data['Responsible'] = ''
# #         row_data['Delay Reason'] = ''
        
# #         data.append(row_data)
# #         milestone_counter += 1
    
# #     # Add summary row with average Weighted % for each month
# #     summary_row = {'Milestone': 'AVERAGE WEIGHTED %', 'Block': ''}
# #     summary_row[f'Target to be complete by {MONTHS[-1]}-{MONTH_YEARS[MONTHS[-1]]}'] = ''
    
# #     for month in MONTHS:
# #         month_year = MONTH_YEARS[month]
        
# #         # Calculate average of Weighted % for this month
# #         weighted_values = []
# #         for row in data:
# #             weighted_val = row.get(f'Weighted %- {month}', '')
# #             if weighted_val and weighted_val != '':
# #                 try:
# #                     val = float(str(weighted_val).replace('%', ''))
# #                     weighted_values.append(val)
# #                 except (ValueError, TypeError):
# #                     pass
        
# #         # Calculate average
# #         if weighted_values:
# #             avg_weighted = sum(weighted_values) / len(weighted_values)
# #             summary_row[f'Weighted %- {month}'] = f'{avg_weighted:.1f}%'
# #         else:
# #             summary_row[f'Weighted %- {month}'] = ''
        
# #         # Leave other columns blank for summary row
# #         summary_row[f'Target - {month}-{month_year}'] = ''
# #         summary_row[f'% work done- {month} Status'] = ''
# #         summary_row[f'Achieved- {month} {month_year}'] = ''
# #         summary_row[f'Weightage- {month}'] = ''
    
# #     summary_row['Responsible'] = ''
# #     summary_row['Delay Reason'] = ''
    
# #     data.append(summary_row)
    
# #     # Create DataFrame
# #     df = pd.DataFrame(data)
    
# #     # Reorder columns to match format
# #     column_order = ['Milestone', 'Block', f'Target to be complete by {MONTHS[-1]}-{MONTH_YEARS[MONTHS[-1]]}']
    
# #     for month in MONTHS:
# #         month_year = MONTH_YEARS[month]
# #         column_order.extend([
# #             f'Target - {month}-{month_year}',
# #             f'% work done- {month} Status',
# #             f'Achieved- {month} {month_year}',
# #             f'Weightage- {month}',
# #             f'Weighted %- {month}'
# #         ])
    
# #     column_order.extend(['Responsible', 'Delay Reason'])
    
# #     df = df[column_order]
    
# #     return df

# # def apply_manual_overrides(df):
# #     """
# #     Apply any manual overrides if needed
# #     This function can be customized based on business rules
# #     """
# #     logger.info("Applying manual overrides (if any)...")
# #     return df

# # # -----------------------------------------------------------------------------
# # # REPORT GENERATION
# # # -----------------------------------------------------------------------------

# # def write_wcc_excel_report_consolidated(df, filename):
# #     """Generate Excel report with proper formatting matching the format file"""
    
# #     logger.info(f'Generating consolidated report: {filename}')
    
# #     wb = Workbook()
# #     ws = wb.active
# #     ws.title = 'WCC Progress'
    
# #     # Define styles
# #     header_font = Font(name='Calibri', size=11, bold=True)
# #     normal_font = Font(name='Calibri', size=10)
# #     date_font = Font(name='Calibri', size=10, bold=True)
# #     summary_font = Font(name='Calibri', size=11, bold=True)
    
# #     center = Alignment(horizontal='center', vertical='center', wrap_text=True)
# #     left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
# #     border = Border(
# #         left=Side(style='thin'),
# #         right=Side(style='thin'),
# #         top=Side(style='thin'),
# #         bottom=Side(style='thin')
# #     )
    
# #     light_grey_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
# #     light_blue_fill = PatternFill(start_color='B4C7E7', end_color='B4C7E7', fill_type='solid')
# #     summary_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    
# #     # Title row
# #     ws.merge_cells('A1:T1')
# #     ws['A1'] = 'Wave City Club- Progress Against Milestones'
# #     ws['A1'].font = Font(name='Calibri', size=14, bold=True)
# #     ws['A1'].alignment = center
    
# #     # Date row
# #     current_date = datetime.now().strftime('%d-%m-%Y')
# #     ws.merge_cells('A2:T2')
# #     ws['A2'] = f'Report Generated on: {current_date}'
# #     ws['A2'].font = date_font
# #     ws['A2'].alignment = center
    
# #     # Month info row
# #     month_info = ', '.join([f'{m} {MONTH_YEARS[m]}' for m in MONTHS])
# #     ws.merge_cells('A3:T3')
# #     ws['A3'] = f'Months Covered: {month_info}'
# #     ws['A3'].font = date_font
# #     ws['A3'].alignment = center
    
# #     # Empty row
# #     ws.merge_cells('A4:T4')
    
# #     # Write dataframe starting at row 5
# #     for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=5):
# #         for c_idx, value in enumerate(row, start=1):
# #             cell = ws.cell(row=r_idx, column=c_idx, value=value)
    
# #     # Style header row (row 5)
# #     header_row = 5
# #     for cell in ws[header_row]:
# #         cell.font = header_font
# #         cell.alignment = center
# #         cell.border = border
# #         cell.fill = light_grey_fill
    
# #     # Style data rows
# #     data_start = 6
# #     summary_row_idx = ws.max_row
# #     data_end = summary_row_idx - 1
    
# #     for row_num in range(data_start, data_end + 1):
# #         for col_num in range(1, len(df.columns) + 1):
# #             cell = ws.cell(row=row_num, column=col_num)
# #             cell.font = normal_font
# #             cell.border = border
            
# #             # Alignment based on column type
# #             if col_num in [1, 2, 3] or 'Target' in str(ws.cell(row=header_row, column=col_num).value or ''):
# #                 cell.alignment = left
# #             else:
# #                 cell.alignment = center
    
# #     # Style summary row (last row)
# #     for col_num in range(1, len(df.columns) + 1):
# #         cell = ws.cell(row=summary_row_idx, column=col_num)
# #         cell.font = summary_font
# #         cell.border = border
# #         cell.fill = summary_fill
# #         cell.alignment = center
    
# #     # Dynamic column width adjustment
# #     for col_num in range(1, len(df.columns) + 1):
# #         col_letter = get_column_letter(col_num)
        
# #         max_length = 0
# #         for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=col_num, max_col=col_num):
# #             for cell in row:
# #                 if cell.value:
# #                     max_length = max(max_length, len(str(cell.value)))
        
# #         calculated_width = min(max(max_length + 2, 8), 15)
# #         ws.column_dimensions[col_letter].width = calculated_width
    
# #     # Set row heights
# #     ws.row_dimensions[1].height = 25
# #     ws.row_dimensions[2].height = 20
# #     ws.row_dimensions[3].height = 20
# #     for i in range(5, ws.max_row + 1):
# #         ws.row_dimensions[i].height = 25
# #     ws.row_dimensions[summary_row_idx].height = 30  # Make summary row taller
    
# #     wb.save(filename)
# #     logger.info(f'Report saved to {filename}')

# # def get_unique_filename(base_name):
# #     """If file exists, append (1), (2), etc."""
# #     if not os.path.exists(base_name):
# #         return base_name

# #     name, ext = os.path.splitext(base_name)
# #     counter = 1
# #     new_name = f"{name}({counter}){ext}"
# #     while os.path.exists(new_name):
# #         counter += 1
# #         new_name = f"{name}({counter}){ext}"
# #     return new_name

# # # -----------------------------------------------------------------------------
# # # MAIN FUNCTION
# # # -----------------------------------------------------------------------------

# # def main():
# #     """Main execution function for quarterly report generation"""
# #     logger.info("=== STARTING WAVE CITY CLUB QUARTERLY REPORT GENERATION (UPDATED LOGIC) ===")
    
# #     try:
# #         # Initialize COS client
# #         cos = init_cos()
        
# #         # Step 1: Get latest KRA file from Milestone/ folder
# #         logger.info("\n=== STEP 1: Finding Latest KRA File ===")
# #         kra_file = get_latest_kra_file(cos)
        
# #         if not kra_file:
# #             logger.error("❌ Failed to find KRA file")
# #             return
        
# #         logger.info(f"✅ Using KRA file: {kra_file}")
        
# #         # Step 2: Extract months from KRA filename and setup
# #         logger.info("\n=== STEP 2: Setting Up Quarterly Months ===")
# #         if not setup_quarterly_months(os.path.basename(kra_file)):
# #             logger.error("❌ Failed to setup quarterly months")
# #             return
        
# #         logger.info(f"✅ Quarterly months: {MONTHS}")
# #         logger.info(f"✅ Month-Year mapping: {MONTH_YEARS}")
        
# #         # Step 3: Find appropriate trackers for each month from Wave City Club/ folder
# #         logger.info("\n=== STEP 3: Finding Trackers for Each Month ===")
# #         global TRACKER_PATHS
        
# #         for month in MONTHS:
# #             month_year = MONTH_YEARS[month]
# #             tracker_path = get_tracker_for_month(cos, month, month_year)
            
# #             if tracker_path:
# #                 TRACKER_PATHS[month] = tracker_path
# #                 logger.info(f"✅ {month} {month_year}: {tracker_path}")
# #             else:
# #                 logger.warning(f"⚠️ {month} {month_year}: No tracker found - column will be blank in report")
# #                 TRACKER_PATHS[month] = None
        
# #         # Continue even if no trackers found
# #         logger.info(f"Found trackers for {sum(1 for v in TRACKER_PATHS.values() if v)} out of {len(MONTHS)} months")
        
# #         # Step 4: Get targets from KRA file (Club sheet)
# #         logger.info("\n=== STEP 4: Extracting Targets from KRA File ===")
# #         targets = get_wcc_targets_from_kra(cos)
        
# #         if not targets:
# #             logger.error("❌ Failed to extract targets from KRA file")
# #             return
        
# #         logger.info(f"✅ Extracted targets for {len(targets)} blocks")
        
# #         # Step 5: Extract progress data for all months with NEW LOGIC
# #         logger.info("\n=== STEP 5: Extracting Progress Data from Trackers (NEW LOGIC) ===")
# #         df = get_wcc_progress_from_tracker_all_months(cos, targets)
        
# #         if df.empty:
# #             logger.error("❌ Failed to generate progress data")
# #             return
        
# #         logger.info(f"✅ Generated progress data for {len(df)-1} milestones + 1 summary row")
        
# #         # Step 5.5: Apply manual overrides
# #         logger.info("\n=== STEP 5.5: Applying Manual Overrides ===")
# #         df = apply_manual_overrides(df)
        
# #         # Step 6: Generate Excel report
# #         logger.info("\n=== STEP 6: Generating Excel Report ===")
# #         current_date_for_filename = datetime.now().strftime('%d-%m-%Y')
        
# #         # Create filename with quarter months
# #         quarter_str = "_".join(MONTHS)
# #         base_filename = f"Wave_City_Club_Milestone_Report_{quarter_str}_{current_date_for_filename}.xlsx"
# #         filename = get_unique_filename(base_filename)
        
# #         write_wcc_excel_report_consolidated(df, filename)
        
# #         logger.info("\n=== WAVE CITY CLUB QUARTERLY REPORT GENERATION COMPLETE ===")
# #         logger.info(f"✅ Report saved as: {filename}")
        
# #         # Log summary
# #         logger.info("\n=== REPORT SUMMARY ===")
# #         logger.info(f"  Quarter Months: {', '.join([f'{m} {MONTH_YEARS[m]}' for m in MONTHS])}")
# #         logger.info(f"  KRA File: {os.path.basename(kra_file)}")
# #         logger.info(f"  Total Blocks: {len(targets)}")
# #         logger.info(f"  Trackers Used:")
# #         for month in MONTHS:
# #             tracker = TRACKER_PATHS.get(month)
# #             if tracker:
# #                 tracker_date = extract_date_from_filename(os.path.basename(tracker))
# #                 logger.info(f"    - {month} {MONTH_YEARS[month]}: {os.path.basename(tracker)} (dated {tracker_date.strftime('%d-%m-%Y') if tracker_date else 'Unknown'})")
# #             else:
# #                 logger.info(f"    - {month} {MONTH_YEARS[month]}: Not Available (column will be blank)")
# #         logger.info(f"  Total Milestones: {len(df)-1} + 1 summary row")
        
# #     except Exception as e:
# #         logger.error(f"❌ Error in main execution: {e}")
# #         import traceback
# #         logger.error(traceback.format_exc())
# #         raise

# # if __name__ == "__main__":
# #     main()






















# # import os
# # import re
# # import logging
# # from io import BytesIO
# # from datetime import datetime
# # from dateutil.relativedelta import relativedelta

# # import pandas as pd
# # from openpyxl import Workbook, load_workbook
# # from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
# # from openpyxl.utils import get_column_letter
# # from openpyxl.utils.dataframe import dataframe_to_rows
# # from dotenv import load_dotenv
# # import ibm_boto3
# # from ibm_botocore.client import Config

# # # -----------------------------------------------------------------------------
# # # CONFIG / CONSTANTS
# # # -----------------------------------------------------------------------------
# # load_dotenv()
# # logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(levelname)s] %(message)s')
# # logger = logging.getLogger(__name__)

# # # Validate required environment variables
# # required = {
# #     'COS_API_KEY': os.getenv('COS_API_KEY'),
# #     'COS_SERVICE_INSTANCE_CRN': os.getenv('COS_SERVICE_INSTANCE_CRN'),
# #     'COS_ENDPOINT': os.getenv('COS_ENDPOINT'),
# #     'COS_BUCKET_NAME': os.getenv('COS_BUCKET_NAME'),
# # }
# # missing = [k for k, v in required.items() if not v]
# # if missing:
# #     logger.error(f"Missing required environment variables: {', '.join(missing)}")
# #     raise SystemExit(1)

# # COS_API_KEY     = required['COS_API_KEY']
# # COS_CRN         = required['COS_SERVICE_INSTANCE_CRN']
# # COS_ENDPOINT    = required['COS_ENDPOINT']
# # BUCKET          = required['COS_BUCKET_NAME']

# # # Folder paths in COS
# # KRA_FOLDER = "Milestone/"
# # TRACKER_FOLDER = "Wave City Club/"

# # # Dynamic KRA file path
# # WCC_KRA_KEY = None

# # # Dynamic tracker paths
# # TRACKER_PATHS = {}  # Maps month names to tracker file paths
# # LOADED_TRACKERS = {}  # Cache loaded workbooks

# # # Dynamic months and years
# # MONTHS = []
# # MONTH_YEARS = {}  # Maps month name to year
# # TARGET_END_MONTH = None
# # TARGET_END_YEAR = None

# # # Quarterly definitions (for reference)
# # QUARTERLY_GROUPS = [
# #     ['June', 'July', 'August'],           # Q1
# #     ['September', 'October', 'November'], # Q2
# #     ['December', 'January', 'February'],  # Q3
# #     ['March', 'April', 'May']             # Q4
# # ]

# # # Month shift logic: Report Month -> Tracker Month
# # MONTH_SHIFT = {
# #     'June': 'July',
# #     'July': 'August',
# #     'August': 'September',
# #     'September': 'October',
# #     'October': 'November',
# #     'November': 'December',
# #     'December': 'January',
# #     'January': 'February',
# #     'February': 'March',
# #     'March': 'April',
# #     'April': 'May',
# #     'May': 'June'
# # }

# # # Block mapping from KRA to tracker sheets
# # BLOCK_MAPPING = {
# #     'Block 1 (B1) Banquet Hall': 'B1 Banket Hall & Finedine ',
# #     'Fine Dine': 'B1 Banket Hall & Finedine ',
# #     'Block 5 (B5) Admin + Member Lounge+Creche+Av Room + Surveillance Room +Toilets': 'B5',
# #     'Block 6 (B6) Toilets': 'B6',
# #     'Block 7(B7) Indoor Sports': 'B7',
# #     'Block 9 (B9) Spa & Saloon': 'B9',
# #     'Block 8 (B8) Squash Court': 'B8',
# #     'Block 2 & 3 (B2 & B3) Cafe & Bar': 'B2 & B3',
# #     'Block 4 (B4) Indoor Swimming Pool Changing Room & Toilets': 'B4',
# #     'Block 11 (B11) Guest House': 'B11',
# #     'Block 10 (B10) Gym': 'B10'
# # }

# # # KRA Column structure (0-indexed)
# # KRA_COLUMNS = {
# #     'BLOCK': 0,          # Column A: Block name
# #     'MONTH1_ACTIVITY': 1,  # Column B: First month activity
# #     'MONTH1_TARGET': 2,    # Column C: First month target %
# #     'MONTH2_ACTIVITY': 3,  # Column D: Second month activity
# #     'MONTH2_TARGET': 4,    # Column E: Second month target %
# #     'MONTH3_ACTIVITY': 5,  # Column F: Third month activity
# #     'MONTH3_TARGET': 6,    # Column G: Third month target %
# # }

# # # Tracker column positions (0-indexed)
# # TRACKER_COLUMNS = {
# #     'ACTIVITY_NAME': 6,    # Column G: Activity Name
# #     'PERCENT_COMPLETE': 12  # Column M: % Complete
# # }

# # # -----------------------------------------------------------------------------
# # # COS HELPERS
# # # -----------------------------------------------------------------------------

# # def init_cos():
# #     return ibm_boto3.client(
# #         's3',
# #         ibm_api_key_id=COS_API_KEY,
# #         ibm_service_instance_id=COS_CRN,
# #         config=Config(signature_version='oauth'),
# #         endpoint_url=COS_ENDPOINT,
# #     )

# # def download_file_bytes(cos, key):
# #     if not key:
# #         raise ValueError("File key cannot be None or empty")
# #     obj = cos.get_object(Bucket=BUCKET, Key=key)
# #     return obj['Body'].read()

# # def list_files_in_folder(cos, folder_prefix):
# #     """List all files in a specific folder (prefix) in the COS bucket"""
# #     try:
# #         response = cos.list_objects_v2(Bucket=BUCKET, Prefix=folder_prefix)
# #         files = []
# #         if 'Contents' in response:
# #             for obj in response['Contents']:
# #                 if not obj['Key'].endswith('/'):
# #                     files.append(obj['Key'])
# #         return files
# #     except Exception as e:
# #         logger.error(f"Error listing files in folder {folder_prefix}: {e}")
# #         return []

# # def extract_date_from_filename(filename):
# #     """Extract date from filename in format (dd-mm-yyyy)"""
# #     pattern = r'\((\d{2}-\d{2}-\d{4})\)'
# #     match = re.search(pattern, filename)
# #     if match:
# #         date_str = match.group(1)
# #         try:
# #             return datetime.strptime(date_str, '%d-%m-%Y')
# #         except ValueError:
# #             logger.warning(f"Could not parse date {date_str} from filename {filename}")
# #             return None
# #     return None

# # def get_month_name(month_num):
# #     """Convert month number to month name"""
# #     months = {
# #         1: "January", 2: "February", 3: "March", 4: "April",
# #         5: "May", 6: "June", 7: "July", 8: "August", 
# #         9: "September", 10: "October", 11: "November", 12: "December"
# #     }
# #     return months.get(month_num, "Unknown")

# # def get_month_number(month_name):
# #     """Convert month name to month number"""
# #     months = {
# #         "January": 1, "February": 2, "March": 3, "April": 4,
# #         "May": 5, "June": 6, "July": 7, "August": 8, 
# #         "September": 9, "October": 10, "November": 11, "December": 12
# #     }
# #     return months.get(month_name, 1)

# # def get_latest_kra_file(cos):
# #     """Get the latest KRA milestone file from the Milestone/ folder"""
# #     global WCC_KRA_KEY
    
# #     logger.info("=== FINDING LATEST KRA MILESTONE FILE ===")
    
# #     # List all files in Milestone folder
# #     all_files = list_files_in_folder(cos, KRA_FOLDER)
# #     logger.info(f"Found {len(all_files)} files in {KRA_FOLDER} folder")
    
# #     # Pattern to match KRA milestone files
# #     kra_pattern = r'KRA Milestones for.*\.xlsx$'
    
# #     matching_files = []
    
# #     for file_path in all_files:
# #         filename = os.path.basename(file_path)
# #         if re.search(kra_pattern, filename, re.IGNORECASE):
# #             logger.info(f"Found KRA file: {filename}")
            
# #             # Extract date from filename
# #             file_date = extract_date_from_filename(filename)
# #             if file_date:
# #                 matching_files.append((file_path, file_date))
# #                 logger.info(f"  Extracted date: {file_date.strftime('%d-%m-%Y')}")
# #             else:
# #                 logger.warning(f"  Could not extract date from {filename}")
# #                 matching_files.append((file_path, datetime.min))  # Fallback for files without dates
    
# #     if matching_files:
# #         # Sort by date (newest first)
# #         matching_files.sort(key=lambda x: x[1], reverse=True)
# #         latest_file = matching_files[0][0]
# #         WCC_KRA_KEY = latest_file
# #         logger.info(f"✅ Latest KRA file: {os.path.basename(WCC_KRA_KEY)} (dated {matching_files[0][1].strftime('%d-%m-%Y')})")
# #     else:
# #         logger.error(f"❌ No KRA milestone files found in {KRA_FOLDER} folder!")
# #         WCC_KRA_KEY = None
    
# #     return WCC_KRA_KEY

# # def extract_months_from_kra_filename(filename):
# #     """
# #     Extract quarter months from KRA filename
# #     Format: "KRA Milestones for [Month] [Month] [Month] [Year].xlsx"
# #     Example: "KRA Milestones for December January February 2026.xlsx"
# #     The year in the filename is the END year of the quarter
# #     """
# #     months_pattern = r'KRA\s+Milestones\s+for\s+((?:January|February|March|April|May|June|July|August|September|October|November|December)(?:\s+(?:January|February|March|April|May|June|July|August|September|October|November|December))*)\s+(\d{4})'
    
# #     match = re.search(months_pattern, filename, re.IGNORECASE)
# #     if match:
# #         months_str = match.group(1)
# #         end_year = int(match.group(2))
        
# #         # Extract individual months in order
# #         month_names = re.findall(r'January|February|March|April|May|June|July|August|September|October|November|December', 
# #                                  months_str, re.IGNORECASE)
        
# #         logger.info(f"Extracted from KRA filename '{filename}':")
# #         logger.info(f"  Months: {month_names}, End Year: {end_year}")
        
# #         return [m.capitalize() for m in month_names], end_year
    
# #     logger.warning(f"Could not parse KRA filename: {filename}")
# #     return None, None

# # def get_tracker_for_month(cos, report_month, month_year):
# #     """
# #     Get tracker file for a specific report month using the month-shift logic.
# #     Uses the LATEST tracker available that matches the required month/year.
    
# #     Args:
# #         cos: COS client
# #         report_month: The month shown in the report (e.g., "June")
# #         month_year: Year of the report month
    
# #     Returns:
# #         Path to tracker file or None
# #     """
# #     # Get the tracker month based on shift logic
# #     tracker_month_name = MONTH_SHIFT.get(report_month)
    
# #     if not tracker_month_name:
# #         logger.warning(f"No month shift mapping found for {report_month}")
# #         return None
    
# #     # Calculate the tracker year (handle year rollover)
# #     tracker_month_num = get_month_number(tracker_month_name)
# #     report_month_num = get_month_number(report_month)
    
# #     tracker_year = month_year
# #     # If tracker month is earlier in the year than report month, it's next year
# #     if tracker_month_num < report_month_num:
# #         tracker_year += 1
    
# #     logger.info(f"Looking for tracker for Report: {report_month} {month_year} → Tracker: {tracker_month_name} {tracker_year}")
    
# #     # List all files in Wave City Club folder
# #     all_files = list_files_in_folder(cos, TRACKER_FOLDER)
# #     logger.info(f"Total files in {TRACKER_FOLDER}: {len(all_files)}")
    
# #     # Pattern to match Structure Work tracker files
# #     tracker_pattern = r'Structure.*Work.*Tracker.*Wave.*City.*Club.*\.xlsx$'
    
# #     matching_files = []
    
# #     for file_path in all_files:
# #         filename = os.path.basename(file_path)
# #         if re.search(tracker_pattern, filename, re.IGNORECASE):
# #             # Extract date from filename
# #             file_date = extract_date_from_filename(filename)
            
# #             logger.info(f"  Found tracker: {filename}")
            
# #             if file_date:
# #                 logger.info(f"    Extracted date: {file_date.strftime('%d-%m-%Y')} (Month: {file_date.month}, Year: {file_date.year})")
# #                 logger.info(f"    Looking for: Month={tracker_month_num}, Year={tracker_year}")
                
# #                 # Check if this file matches the month and year we're looking for
# #                 if file_date.month == tracker_month_num and file_date.year == tracker_year:
# #                     matching_files.append((file_path, file_date))
# #                     logger.info(f"    ✅ MATCH!")
# #                 else:
# #                     logger.info(f"    ❌ Date mismatch (need {tracker_month_name} {tracker_year})")
# #             else:
# #                 logger.warning(f"    Could not extract date from {filename}")
    
# #     if matching_files:
# #         # Sort by date and get the LATEST matching tracker
# #         matching_files.sort(key=lambda x: x[1], reverse=True)
# #         latest_tracker = matching_files[0][0]
# #         logger.info(f"✅ Selected tracker: {os.path.basename(latest_tracker)} (dated {matching_files[0][1].strftime('%d-%m-%Y')})")
# #         return latest_tracker
# #     else:
# #         logger.warning(f"⚠️ No tracker found for {report_month} {month_year} (looking for {tracker_month_name} {tracker_year} dated files)")
# #         return None

# # def setup_quarterly_months(kra_filename):
# #     """
# #     Setup the quarterly months and years based on KRA filename.
    
# #     The year in the KRA filename represents the END year of the quarter.
# #     For cross-year quarters (e.g., Dec-Jan-Feb 2026):
# #     - December belongs to 2025
# #     - January belongs to 2026
# #     - February belongs to 2026
# #     """
# #     global MONTHS, MONTH_YEARS, TARGET_END_MONTH, TARGET_END_YEAR
    
# #     months, end_year = extract_months_from_kra_filename(kra_filename)
    
# #     if not months or not end_year:
# #         logger.error(f"Could not extract months from KRA filename: {kra_filename}")
# #         return False
    
# #     MONTHS = months
# #     TARGET_END_MONTH = MONTHS[-1]
# #     TARGET_END_YEAR = end_year
    
# #     # Determine starting year for the quarter
# #     first_month_num = get_month_number(MONTHS[0])
# #     last_month_num = get_month_number(MONTHS[-1])
    
# #     # For cross-year quarters (first month number > last month number),
# #     # the first month belongs to the previous year
# #     if first_month_num > last_month_num:
# #         # Cross-year quarter (e.g., Dec 2025, Jan 2026, Feb 2026)
# #         start_year = end_year - 1
# #         logger.info(f"Cross-year quarter detected: {MONTHS[0]} {start_year} - {MONTHS[-1]} {end_year}")
# #     else:
# #         # Same-year quarter (e.g., Jun 2025, Jul 2025, Aug 2025)
# #         start_year = end_year
# #         logger.info(f"Same-year quarter: all months in {start_year}")
    
# #     # Assign years to each month
# #     current_year = start_year
# #     for i, month in enumerate(MONTHS):
# #         month_num = get_month_number(month)
        
# #         # Year increments when month number decreases (year rollover)
# #         if i > 0:
# #             prev_month_num = get_month_number(MONTHS[i-1])
# #             if month_num < prev_month_num:
# #                 current_year += 1
        
# #         MONTH_YEARS[month] = current_year
    
# #     logger.info(f"Month-Year mapping:")
# #     for month in MONTHS:
# #         logger.info(f"  {month}: {MONTH_YEARS[month]}")
    
# #     return True


# # def find_club_sheet(wb):
# #     """Find the sheet with 'Club' in its name"""
# #     for sheet_name in wb.sheetnames:
# #         if 'club' in sheet_name.lower():
# #             logger.info(f"Found Club sheet: {sheet_name}")
# #             return sheet_name
    
# #     logger.warning("No sheet with 'Club' in name found. Available sheets: " + ", ".join(wb.sheetnames))
# #     return None

# # # -----------------------------------------------------------------------------
# # # DATA EXTRACTION
# # # -----------------------------------------------------------------------------

# # def get_wcc_targets_from_kra(cos):
# #     """
# #     Extract targets from KRA milestone file with updated logic:
# #     - Look for sheet with "Club" in its name
# #     - For each activity, fetch both the target activity name AND target %
# #     - Store both values for comparison with tracker data
# #     """
# #     global WCC_KRA_KEY
    
# #     if not WCC_KRA_KEY:
# #         raise ValueError("WCC_KRA_KEY is not set. Call get_latest_kra_file() first.")
    
# #     logger.info(f"Reading KRA file: {WCC_KRA_KEY}")
    
# #     # Download and load KRA file
# #     kra_bytes = download_file_bytes(cos, WCC_KRA_KEY)
# #     kra_wb = load_workbook(BytesIO(kra_bytes), data_only=True)
    
# #     # Find the Club sheet
# #     club_sheet_name = find_club_sheet(kra_wb)
# #     if not club_sheet_name:
# #         # Fallback to active sheet
# #         logger.warning("Using active sheet as fallback")
# #         ws = kra_wb.active
# #     else:
# #         ws = kra_wb[club_sheet_name]
    
# #     targets = {}
    
# #     # Find header row (look for "Blocks" in column A)
# #     header_row = None
# #     for row_idx in range(1, min(10, ws.max_row + 1)):
# #         cell_value = ws.cell(row=row_idx, column=1).value
# #         if cell_value and 'block' in str(cell_value).lower():
# #             header_row = row_idx
# #             logger.info(f"Found header row at row {header_row}")
# #             break
    
# #     if not header_row:
# #         logger.error("Could not find header row with 'Blocks'")
# #         kra_wb.close()
# #         return targets
    
# #     # Start from row after header
# #     data_start_row = header_row + 1
    
# #     for row_idx in range(data_start_row, ws.max_row + 1):
# #         row = list(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
        
# #         # Extract block name from column A
# #         block_name = row[KRA_COLUMNS['BLOCK']]
        
# #         # Skip empty rows
# #         if not block_name or pd.isna(block_name):
# #             continue
        
# #         # Stop if we hit another header row or non-block data
# #         block_name_str = str(block_name).strip()
# #         if any(month.lower() in block_name_str.lower() for month in ['january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december']):
# #             logger.info(f"Stopping at row {row_idx}: Found header-like content '{block_name_str}'")
# #             break
        
# #         # Skip if block name is just "Activity" or other single-word non-block names
# #         if block_name_str.lower() in ['activity', 'activities', 'milestone', 'milestones', 'target', 'targets', 'blocks']:
# #             logger.info(f"Skipping non-block row: '{block_name_str}'")
# #             continue
            
# #         logger.info(f"Processing block: {block_name}")
        
# #         targets[block_name] = {}
        
# #         # Process each month
# #         for i, month in enumerate(MONTHS):
# #             month_year = MONTH_YEARS[month]
            
# #             # Calculate column indices for this month
# #             if i == 0:
# #                 activity_col = KRA_COLUMNS['MONTH1_ACTIVITY']
# #                 target_col = KRA_COLUMNS['MONTH1_TARGET']
# #             elif i == 1:
# #                 activity_col = KRA_COLUMNS['MONTH2_ACTIVITY']
# #                 target_col = KRA_COLUMNS['MONTH2_TARGET']
# #             elif i == 2:
# #                 activity_col = KRA_COLUMNS['MONTH3_ACTIVITY']
# #                 target_col = KRA_COLUMNS['MONTH3_TARGET']
# #             else:
# #                 logger.warning(f"More than 3 months found, skipping month {month}")
# #                 continue
            
# #             # Extract activity name and target %
# #             activity_name = row[activity_col] if activity_col < len(row) else None
# #             target_percent = row[target_col] if target_col < len(row) else None
            
# #             # Validate target_percent is a number, not text
# #             validated_target_percent = None
# #             if target_percent and not pd.isna(target_percent):
# #                 try:
# #                     validated_target_percent = float(target_percent)
# #                 except (ValueError, TypeError):
# #                     logger.warning(f"Invalid target % value for {block_name} - {month}: '{target_percent}' (not a number)")
# #                     validated_target_percent = None
            
# #             # Store both activity and target % for this month
# #             targets[block_name][month] = {
# #                 'activity': activity_name if activity_name and not pd.isna(activity_name) else None,
# #                 'target_percent': validated_target_percent
# #             }
            
# #             logger.info(f"  {month}: Activity='{activity_name}', Target %={validated_target_percent}")
    
# #     kra_wb.close()
# #     return targets

# # def find_activity_completion_in_tracker(tracker_wb, sheet_name, target_activity):
# #     """
# #     Search for an activity in the tracker and return its % complete value.
    
# #     Args:
# #         tracker_wb: Loaded tracker workbook
# #         sheet_name: Sheet name to search in
# #         target_activity: Activity name to find
    
# #     Returns:
# #         Float value of % complete, or None if not found
# #     """
# #     if sheet_name not in tracker_wb.sheetnames:
# #         logger.warning(f"Sheet '{sheet_name}' not found in tracker")
# #         return None
    
# #     ws = tracker_wb[sheet_name]
    
# #     # Search through rows starting from row 3 (skip headers)
# #     for row_idx in range(2, ws.max_row + 1):
# #         row = list(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
        
# #         # Get activity name from column G (index 6)
# #         if len(row) > TRACKER_COLUMNS['ACTIVITY_NAME']:
# #             activity_name = row[TRACKER_COLUMNS['ACTIVITY_NAME']]
            
# #             if activity_name and str(activity_name).strip().lower() == str(target_activity).strip().lower():
# #                 # Found the activity, get % complete from column M (index 12)
# #                 if len(row) > TRACKER_COLUMNS['PERCENT_COMPLETE']:
# #                     percent_complete = row[TRACKER_COLUMNS['PERCENT_COMPLETE']]
# #                     try:
# #                         return float(percent_complete) if percent_complete is not None else 0.0
# #                     except (ValueError, TypeError):
# #                         logger.warning(f"Could not convert % complete to float for activity '{activity_name}': {percent_complete}")
# #                         return 0.0
    
# #     logger.debug(f"Activity '{target_activity}' not found in sheet '{sheet_name}'")
# #     return None

# # def get_wcc_progress_from_tracker_all_months(cos, targets):
# #     """
# #     Extract progress data for all months with UPDATED LOGIC:
# #     - Weightage = 100 for all blocks
# #     - Weighted % = (% Complete / Weightage) * 100 for each month
# #     - If tracker % > target %, display 100%
# #     - Responsible and Delay columns shown once at the end
# #     """
# #     global TRACKER_PATHS, LOADED_TRACKERS
    
# #     # Load all trackers into memory
# #     for month in MONTHS:
# #         tracker_path = TRACKER_PATHS.get(month)
# #         if tracker_path and month not in LOADED_TRACKERS:
# #             logger.info(f"Loading tracker for {month}: {tracker_path}")
# #             tracker_bytes = download_file_bytes(cos, tracker_path)
# #             LOADED_TRACKERS[month] = load_workbook(BytesIO(tracker_bytes), data_only=True)
    
# #     # Build dataframe
# #     data = []
# #     milestone_counter = 1
    
# #     for block_name, month_targets in targets.items():
# #         # Get the tracker sheet name for this block
# #         tracker_sheet = BLOCK_MAPPING.get(block_name)
        
# #         if not tracker_sheet:
# #             logger.warning(f"No tracker sheet mapping found for block: {block_name}")
# #             continue
        
# #         row_data = {
# #             'Milestone': f'Milestone-{milestone_counter:02d}',
# #             'Block': block_name
# #         }
        
# #         # Determine the target to complete by the end of quarter
# #         last_month = MONTHS[-1]
# #         last_month_target_info = month_targets.get(last_month, {})
# #         target_to_complete = last_month_target_info.get('activity', 'N/A')
# #         row_data[f'Target to be complete by {last_month}-{MONTH_YEARS[last_month]}'] = target_to_complete
        
# #         # Process each month
# #         for month in MONTHS:
# #             month_year = MONTH_YEARS[month]
            
# #             # Get target info for this month
# #             target_info = month_targets.get(month, {})
# #             target_activity = target_info.get('activity')
# #             target_percent = target_info.get('target_percent')
            
# #             # Column names for this month
# #             target_col = f'Target - {month}-{month_year}'
# #             status_col = f'% work done- {month} Status'
# #             achieved_col = f'Achieved- {month} {month_year}'
# #             weightage_col = f'Weightage- {month}'
# #             weighted_pct_col = f'Weighted %- {month}'
            
# #             row_data[target_col] = target_activity if target_activity else 'No target'
            
# #             # Check if tracker is available for this month
# #             if month not in LOADED_TRACKERS:
# #                 # Leave blank if tracker not available
# #                 row_data[status_col] = ''
# #                 row_data[achieved_col] = ''
# #                 row_data[weightage_col] = ''
# #                 row_data[weighted_pct_col] = ''
# #                 continue
            
# #             # If no target activity, mark as N/A
# #             if not target_activity:
# #                 row_data[status_col] = 'N/A'
# #                 row_data[achieved_col] = 'No target for this month'
# #                 row_data[weightage_col] = ''
# #                 row_data[weighted_pct_col] = ''
# #                 continue
            
# #             # Fetch completed % from tracker
# #             tracker_wb = LOADED_TRACKERS[month]
# #             completed_percent = find_activity_completion_in_tracker(tracker_wb, tracker_sheet, target_activity)
            
# #             # Determine status based on comparison
# #             if completed_percent is None:
# #                 # Activity not found in tracker
# #                 row_data[status_col] = '0%'
# #                 row_data[achieved_col] = 'Activity not found in tracker'
# #                 status_percent = 0.0
# #             elif target_percent is not None and completed_percent > target_percent:
# #                 # Tracker % > Target % -> display 100%
# #                 row_data[status_col] = '100%'
# #                 row_data[achieved_col] = f'Target exceeded ({completed_percent*100:.0f}% > {target_percent*100:.0f}%)'
# #                 status_percent = 1.0
# #             elif target_percent is not None and completed_percent == target_percent:
# #                 # Target % matches completed %
# #                 row_data[status_col] = '100%'
# #                 row_data[achieved_col] = f'Target achieved ({completed_percent*100:.0f}% complete)'
# #                 status_percent = 1.0
# #             else:
# #                 # Display completed % from tracker
# #                 status_percent = completed_percent
# #                 row_data[status_col] = f'{completed_percent*100:.0f}%'
                
# #                 if target_percent is not None and completed_percent >= target_percent:
# #                     row_data[achieved_col] = f'{completed_percent*100:.0f}% completed (Target: {target_percent*100:.0f}%)'
# #                 elif target_percent is not None:
# #                     row_data[achieved_col] = f'{completed_percent*100:.0f}% completed (Target: {target_percent*100:.0f}%)'
# #                 else:
# #                     row_data[achieved_col] = f'{completed_percent*100:.0f}% completed'
            
# #             # Calculate Weightage and Weighted % for this month
# #             weightage = 100  # Always 100
# #             weighted_pct = (status_percent * 100 / weightage) * 100 if weightage > 0 else 0
# #             row_data[weightage_col] = weightage
# #             row_data[weighted_pct_col] = f'{weighted_pct:.0f}%'
        
# #         # Add Responsible and Delay columns once at the end
# #         row_data['Responsible'] = ''
# #         row_data['Delay Reason'] = ''
        
# #         data.append(row_data)
# #         milestone_counter += 1
    
# #     # Add summary row with average Weighted % for each month
# #     summary_row = {'Milestone': 'AVERAGE WEIGHTED %', 'Block': ''}
# #     summary_row[f'Target to be complete by {MONTHS[-1]}-{MONTH_YEARS[MONTHS[-1]]}'] = ''
    
# #     for month in MONTHS:
# #         month_year = MONTH_YEARS[month]
        
# #         # Calculate average of Weighted % for this month
# #         weighted_values = []
# #         for row in data:
# #             weighted_val = row.get(f'Weighted %- {month}', '')
# #             if weighted_val and weighted_val != '':
# #                 try:
# #                     val = float(str(weighted_val).replace('%', ''))
# #                     weighted_values.append(val)
# #                 except (ValueError, TypeError):
# #                     pass
        
# #         # Calculate average
# #         if weighted_values:
# #             avg_weighted = sum(weighted_values) / len(weighted_values)
# #             summary_row[f'Weighted %- {month}'] = f'{avg_weighted:.1f}%'
# #         else:
# #             summary_row[f'Weighted %- {month}'] = ''
        
# #         # Leave other columns blank for summary row
# #         summary_row[f'Target - {month}-{month_year}'] = ''
# #         summary_row[f'% work done- {month} Status'] = ''
# #         summary_row[f'Achieved- {month} {month_year}'] = ''
# #         summary_row[f'Weightage- {month}'] = ''
    
# #     summary_row['Responsible'] = ''
# #     summary_row['Delay Reason'] = ''
    
# #     data.append(summary_row)
    
# #     # Create DataFrame
# #     df = pd.DataFrame(data)
    
# #     # Reorder columns to match format
# #     column_order = ['Milestone', 'Block', f'Target to be complete by {MONTHS[-1]}-{MONTH_YEARS[MONTHS[-1]]}']
    
# #     for month in MONTHS:
# #         month_year = MONTH_YEARS[month]
# #         column_order.extend([
# #             f'Target - {month}-{month_year}',
# #             f'% work done- {month} Status',
# #             f'Achieved- {month} {month_year}',
# #             f'Weightage- {month}',
# #             f'Weighted %- {month}'
# #         ])
    
# #     column_order.extend(['Responsible', 'Delay Reason'])
    
# #     df = df[column_order]
    
# #     return df

# # def apply_manual_overrides(df):
# #     """
# #     Apply any manual overrides if needed
# #     This function can be customized based on business rules
# #     """
# #     logger.info("Applying manual overrides (if any)...")
# #     return df

# # # -----------------------------------------------------------------------------
# # # REPORT GENERATION
# # # -----------------------------------------------------------------------------

# # def write_wcc_excel_report_consolidated(df, filename):
# #     """Generate Excel report with proper formatting matching the format file"""
    
# #     logger.info(f'Generating consolidated report: {filename}')
    
# #     wb = Workbook()
# #     ws = wb.active
# #     ws.title = 'WCC Progress'
    
# #     # Define styles
# #     header_font = Font(name='Calibri', size=11, bold=True)
# #     normal_font = Font(name='Calibri', size=10)
# #     date_font = Font(name='Calibri', size=10, bold=True)
# #     summary_font = Font(name='Calibri', size=11, bold=True)
    
# #     center = Alignment(horizontal='center', vertical='center', wrap_text=True)
# #     left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
# #     border = Border(
# #         left=Side(style='thin'),
# #         right=Side(style='thin'),
# #         top=Side(style='thin'),
# #         bottom=Side(style='thin')
# #     )
    
# #     light_grey_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
# #     light_blue_fill = PatternFill(start_color='B4C7E7', end_color='B4C7E7', fill_type='solid')
# #     summary_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    
# #     # Title row
# #     ws.merge_cells('A1:T1')
# #     ws['A1'] = 'Wave City Club- Progress Against Milestones'
# #     ws['A1'].font = Font(name='Calibri', size=14, bold=True)
# #     ws['A1'].alignment = center
    
# #     # Date row
# #     current_date = datetime.now().strftime('%d-%m-%Y')
# #     ws.merge_cells('A2:T2')
# #     ws['A2'] = f'Report Generated on: {current_date}'
# #     ws['A2'].font = date_font
# #     ws['A2'].alignment = center
    
# #     # Month info row
# #     month_info = ', '.join([f'{m} {MONTH_YEARS[m]}' for m in MONTHS])
# #     ws.merge_cells('A3:T3')
# #     ws['A3'] = f'Months Covered: {month_info}'
# #     ws['A3'].font = date_font
# #     ws['A3'].alignment = center
    
# #     # Empty row
# #     ws.merge_cells('A4:T4')
    
# #     # Write dataframe starting at row 5
# #     for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=5):
# #         for c_idx, value in enumerate(row, start=1):
# #             cell = ws.cell(row=r_idx, column=c_idx, value=value)
    
# #     # Style header row (row 5)
# #     header_row = 5
# #     for cell in ws[header_row]:
# #         cell.font = header_font
# #         cell.alignment = center
# #         cell.border = border
# #         cell.fill = light_grey_fill
    
# #     # Style data rows
# #     data_start = 6
# #     summary_row_idx = ws.max_row
# #     data_end = summary_row_idx - 1
    
# #     for row_num in range(data_start, data_end + 1):
# #         for col_num in range(1, len(df.columns) + 1):
# #             cell = ws.cell(row=row_num, column=col_num)
# #             cell.font = normal_font
# #             cell.border = border
            
# #             # Alignment based on column type
# #             if col_num in [1, 2, 3] or 'Target' in str(ws.cell(row=header_row, column=col_num).value or ''):
# #                 cell.alignment = left
# #             else:
# #                 cell.alignment = center
    
# #     # Style summary row (last row)
# #     for col_num in range(1, len(df.columns) + 1):
# #         cell = ws.cell(row=summary_row_idx, column=col_num)
# #         cell.font = summary_font
# #         cell.border = border
# #         cell.fill = summary_fill
# #         cell.alignment = center
    
# #     # Dynamic column width adjustment
# #     for col_num in range(1, len(df.columns) + 1):
# #         col_letter = get_column_letter(col_num)
        
# #         max_length = 0
# #         for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=col_num, max_col=col_num):
# #             for cell in row:
# #                 if cell.value:
# #                     max_length = max(max_length, len(str(cell.value)))
        
# #         calculated_width = min(max(max_length + 2, 8), 15)
# #         ws.column_dimensions[col_letter].width = calculated_width
    
# #     # Set row heights
# #     ws.row_dimensions[1].height = 25
# #     ws.row_dimensions[2].height = 20
# #     ws.row_dimensions[3].height = 20
# #     for i in range(5, ws.max_row + 1):
# #         ws.row_dimensions[i].height = 25
# #     ws.row_dimensions[summary_row_idx].height = 30  # Make summary row taller
    
# #     wb.save(filename)
# #     logger.info(f'Report saved to {filename}')

# # def get_unique_filename(base_name):
# #     """If file exists, append (1), (2), etc."""
# #     if not os.path.exists(base_name):
# #         return base_name

# #     name, ext = os.path.splitext(base_name)
# #     counter = 1
# #     new_name = f"{name}({counter}){ext}"
# #     while os.path.exists(new_name):
# #         counter += 1
# #         new_name = f"{name}({counter}){ext}"
# #     return new_name

# # # -----------------------------------------------------------------------------
# # # MAIN FUNCTION
# # # -----------------------------------------------------------------------------

# # def main():
# #     """Main execution function for quarterly report generation"""
# #     logger.info("=== STARTING WAVE CITY CLUB QUARTERLY REPORT GENERATION (UPDATED LOGIC) ===")
    
# #     try:
# #         # Initialize COS client
# #         cos = init_cos()
        
# #         # Step 1: Get latest KRA file from Milestone/ folder
# #         logger.info("\n=== STEP 1: Finding Latest KRA File ===")
# #         kra_file = get_latest_kra_file(cos)
        
# #         if not kra_file:
# #             logger.error("❌ Failed to find KRA file")
# #             return
        
# #         logger.info(f"✅ Using KRA file: {kra_file}")
        
# #         # Step 2: Extract months from KRA filename and setup
# #         logger.info("\n=== STEP 2: Setting Up Quarterly Months ===")
# #         if not setup_quarterly_months(os.path.basename(kra_file)):
# #             logger.error("❌ Failed to setup quarterly months")
# #             return
        
# #         logger.info(f"✅ Quarterly months: {MONTHS}")
# #         logger.info(f"✅ Month-Year mapping: {MONTH_YEARS}")
        
# #         # Step 3: Find appropriate trackers for each month from Wave City Club/ folder
# #         logger.info("\n=== STEP 3: Finding Trackers for Each Month ===")
# #         global TRACKER_PATHS
        
# #         for month in MONTHS:
# #             month_year = MONTH_YEARS[month]
# #             tracker_path = get_tracker_for_month(cos, month, month_year)
            
# #             if tracker_path:
# #                 TRACKER_PATHS[month] = tracker_path
# #                 logger.info(f"✅ {month} {month_year}: {tracker_path}")
# #             else:
# #                 logger.warning(f"⚠️ {month} {month_year}: No tracker found - column will be blank in report")
# #                 TRACKER_PATHS[month] = None
        
# #         # Continue even if no trackers found
# #         logger.info(f"Found trackers for {sum(1 for v in TRACKER_PATHS.values() if v)} out of {len(MONTHS)} months")
        
# #         # Step 4: Get targets from KRA file (Club sheet)
# #         logger.info("\n=== STEP 4: Extracting Targets from KRA File ===")
# #         targets = get_wcc_targets_from_kra(cos)
        
# #         if not targets:
# #             logger.error("❌ Failed to extract targets from KRA file")
# #             return
        
# #         logger.info(f"✅ Extracted targets for {len(targets)} blocks")
        
# #         # Step 5: Extract progress data for all months with NEW LOGIC
# #         logger.info("\n=== STEP 5: Extracting Progress Data from Trackers (NEW LOGIC) ===")
# #         df = get_wcc_progress_from_tracker_all_months(cos, targets)
        
# #         if df.empty:
# #             logger.error("❌ Failed to generate progress data")
# #             return
        
# #         logger.info(f"✅ Generated progress data for {len(df)-1} milestones + 1 summary row")
        
# #         # Step 5.5: Apply manual overrides
# #         logger.info("\n=== STEP 5.5: Applying Manual Overrides ===")
# #         df = apply_manual_overrides(df)
        
# #         # Step 6: Generate Excel report
# #         logger.info("\n=== STEP 6: Generating Excel Report ===")
# #         current_date_for_filename = datetime.now().strftime('%d-%m-%Y')
        
# #         # Create filename with quarter months
# #         quarter_str = "_".join(MONTHS)
# #         base_filename = f"Wave_City_Club_Milestone_Report_{quarter_str}_{current_date_for_filename}.xlsx"
# #         filename = get_unique_filename(base_filename)
        
# #         write_wcc_excel_report_consolidated(df, filename)
        
# #         logger.info("\n=== WAVE CITY CLUB QUARTERLY REPORT GENERATION COMPLETE ===")
# #         logger.info(f"✅ Report saved as: {filename}")
        
# #         # Log summary
# #         logger.info("\n=== REPORT SUMMARY ===")
# #         logger.info(f"  Quarter Months: {', '.join([f'{m} {MONTH_YEARS[m]}' for m in MONTHS])}")
# #         logger.info(f"  KRA File: {os.path.basename(kra_file)}")
# #         logger.info(f"  Total Blocks: {len(targets)}")
# #         logger.info(f"  Trackers Used:")
# #         for month in MONTHS:
# #             tracker = TRACKER_PATHS.get(month)
# #             if tracker:
# #                 tracker_date = extract_date_from_filename(os.path.basename(tracker))
# #                 logger.info(f"    - {month} {MONTH_YEARS[month]}: {os.path.basename(tracker)} (dated {tracker_date.strftime('%d-%m-%Y') if tracker_date else 'Unknown'})")
# #             else:
# #                 logger.info(f"    - {month} {MONTH_YEARS[month]}: Not Available (column will be blank)")
# #         logger.info(f"  Total Milestones: {len(df)-1} + 1 summary row")
        
# #     except Exception as e:
# #         logger.error(f"❌ Error in main execution: {e}")
# #         import traceback
# #         logger.error(traceback.format_exc())
# #         raise

# # if __name__ == "__main__":
# #     main()
























import os
import sys
import re
import logging
from io import BytesIO
from datetime import datetime
from dateutil.relativedelta import relativedelta

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import ibm_boto3
from ibm_botocore.client import Config

CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR = os.path.dirname(CURRENT_DIR)
if ROOT_DIR not in sys.path:
    sys.path.insert(0, ROOT_DIR)

from env_loader import load_root_env

# -----------------------------------------------------------------------------
# CONFIG / CONSTANTS
# -----------------------------------------------------------------------------
load_root_env()
logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(levelname)s] %(message)s')
logger = logging.getLogger(__name__)

# Validate required environment variables
required = {
    'COS_API_KEY': os.getenv('COS_API_KEY'),
    'COS_SERVICE_INSTANCE_CRN': os.getenv('COS_SERVICE_INSTANCE_CRN'),
    'COS_ENDPOINT': os.getenv('COS_ENDPOINT'),
    'COS_BUCKET_NAME': os.getenv('COS_BUCKET_NAME'),
}
missing = [k for k, v in required.items() if not v]
if missing:
    logger.error(f"Missing required environment variables: {', '.join(missing)}")
    raise SystemExit(1)

COS_API_KEY     = required['COS_API_KEY']
COS_CRN         = required['COS_SERVICE_INSTANCE_CRN']
COS_ENDPOINT    = required['COS_ENDPOINT']
BUCKET          = required['COS_BUCKET_NAME']

# Folder paths in COS
KRA_FOLDER = "Milestone/"
TRACKER_FOLDER = "Wave City Club/"

# Dynamic KRA file path
WCC_KRA_KEY = None

# Dynamic tracker paths
TRACKER_PATHS = {}  # Maps month names to tracker file paths
LOADED_TRACKERS = {}  # Cache loaded workbooks

# Dynamic months and years
MONTHS = []
MONTH_YEARS = {}  # Maps month name to year
TARGET_END_MONTH = None
TARGET_END_YEAR = None

# Quarterly definitions (for reference)
QUARTERLY_GROUPS = [
    ['June', 'July', 'August'],           # Q1
    ['September', 'October', 'November'], # Q2
    ['December', 'January', 'February'],  # Q3
    ['March', 'April', 'May']             # Q4
]

# Month shift logic: Report Month -> Tracker Month
MONTH_SHIFT = {
    'June': 'July',
    'July': 'August',
    'August': 'September',
    'September': 'October',
    'October': 'November',
    'November': 'December',
    'December': 'January',
    'January': 'February',
    'February': 'March',
    'March': 'April',
    'April': 'May',
    'May': 'June'
}

# Block mapping from KRA to tracker sheets
BLOCK_MAPPING = {
    'Block 1 (B1) Banquet Hall': 'B1 Banket Hall & Finedine ',
    'Fine Dine': 'B1 Banket Hall & Finedine ',
    'Block 5 (B5) Admin + Member Lounge+Creche+Av Room + Surveillance Room +Toilets': 'B5',
    'Block 6 (B6) Toilets': 'B6',
    'Block 7(B7) Indoor Sports': 'B7',
    'Block 9 (B9) Spa & Saloon': 'B9',
    'Block 8 (B8) Squash Court': 'B8',
    'Block 2 & 3 (B2 & B3) Cafe & Bar': 'B2 & B3',
    'Block 4 (B4) Indoor Swimming Pool Changing Room & Toilets': 'B4',
    'Block 11 (B11) Guest House': 'B11',
    'Block 10 (B10) Gym': 'B10'
}

# KRA Column structure (0-indexed)
KRA_COLUMNS = {
    'BLOCK': 0,          # Column A: Block name
    'MONTH1_ACTIVITY': 1,  # Column B: First month activity
    'MONTH1_TARGET': 2,    # Column C: First month target %
    'MONTH2_ACTIVITY': 3,  # Column D: Second month activity
    'MONTH2_TARGET': 4,    # Column E: Second month target %
    'MONTH3_ACTIVITY': 5,  # Column F: Third month activity
    'MONTH3_TARGET': 6,    # Column G: Third month target %
}

# Tracker column positions (0-indexed)
TRACKER_COLUMNS = {
    'ACTIVITY_NAME': 6,    # Column G: Activity Name
    'PERCENT_COMPLETE': 12  # Column M: % Complete
}

# -----------------------------------------------------------------------------
# COS HELPERS
# -----------------------------------------------------------------------------

def init_cos():
    return ibm_boto3.client(
        's3',
        ibm_api_key_id=COS_API_KEY,
        ibm_service_instance_id=COS_CRN,
        config=Config(signature_version='oauth'),
        endpoint_url=COS_ENDPOINT,
    )

def download_file_bytes(cos, key):
    if not key:
        raise ValueError("File key cannot be None or empty")
    obj = cos.get_object(Bucket=BUCKET, Key=key)
    return obj['Body'].read()

def list_files_in_folder(cos, folder_prefix):
    """List all files in a specific folder (prefix) in the COS bucket"""
    try:
        response = cos.list_objects_v2(Bucket=BUCKET, Prefix=folder_prefix)
        files = []
        if 'Contents' in response:
            for obj in response['Contents']:
                if not obj['Key'].endswith('/'):
                    files.append(obj['Key'])
        return files
    except Exception as e:
        logger.error(f"Error listing files in folder {folder_prefix}: {e}")
        return []

def extract_date_from_filename(filename):
    """Extract date from filename in format (dd-mm-yyyy) or (d-m-yyyy)"""
    pattern = r'\((\d{1,2}-\d{1,2}-\d{4})\)'
    match = re.search(pattern, filename)
    if match:
        date_str = match.group(1)
        try:
            # Normalise to always parse as d-m-yyyy regardless of zero-padding
            parts = date_str.split('-')
            normalised = f"{int(parts[0]):02d}-{int(parts[1]):02d}-{parts[2]}"
            return datetime.strptime(normalised, '%d-%m-%Y')
        except ValueError:
            logger.warning(f"Could not parse date {date_str} from filename {filename}")
            return None
    return None

def get_month_name(month_num):
    """Convert month number to month name"""
    months = {
        1: "January", 2: "February", 3: "March", 4: "April",
        5: "May", 6: "June", 7: "July", 8: "August", 
        9: "September", 10: "October", 11: "November", 12: "December"
    }
    return months.get(month_num, "Unknown")

def get_month_number(month_name):
    """Convert month name to month number"""
    months = {
        "January": 1, "February": 2, "March": 3, "April": 4,
        "May": 5, "June": 6, "July": 7, "August": 8, 
        "September": 9, "October": 10, "November": 11, "December": 12
    }
    return months.get(month_name, 1)

def get_latest_kra_file(cos):
    """Get the latest KRA milestone file from the Milestone/ folder.
    Uses COS LastModified as the authoritative sort key so files without a
    (dd-mm-yyyy) date in their name are never silently ranked last.
    """
    global WCC_KRA_KEY

    logger.info("=== FINDING LATEST KRA MILESTONE FILE ===")

    try:
        response = cos.list_objects_v2(Bucket=BUCKET, Prefix=KRA_FOLDER)
    except Exception as e:
        logger.error(f"Error listing files in {KRA_FOLDER}: {e}")
        WCC_KRA_KEY = None
        return None

    if 'Contents' not in response:
        logger.error(f"No files found in {KRA_FOLDER}")
        WCC_KRA_KEY = None
        return None

    kra_pattern = r'KRA Milestones for.*\.xlsx$'
    matching_files = []

    for obj in response['Contents']:
        file_key = obj['Key']
        if file_key.endswith('/'):
            continue

        filename = os.path.basename(file_key)
        if not re.search(kra_pattern, filename, re.IGNORECASE):
            continue

        logger.info(f"Found KRA file: {filename}")

        file_date = extract_date_from_filename(filename)
        last_modified = obj['LastModified']

        if file_date:
            logger.info(f"  Filename date: {file_date.strftime('%d-%m-%Y')}")
        else:
            logger.warning(f"  No date in filename — will sort by COS LastModified ({last_modified})")

        matching_files.append({
            'path': file_key,
            'filename': filename,
            'file_date': file_date,
            'last_modified': last_modified,
        })

    if not matching_files:
        logger.error(f"❌ No KRA milestone files found in {KRA_FOLDER}")
        WCC_KRA_KEY = None
        return None

    # Sort by COS LastModified (always available, timezone-aware) — newest first.
    # Filename date is logged for visibility but not used for ranking.
    matching_files.sort(key=lambda f: f['last_modified'], reverse=True)
    latest = matching_files[0]
    WCC_KRA_KEY = latest['path']
    logger.info(f"✅ Latest KRA file: {latest['filename']} (LastModified: {latest['last_modified']})")
    return WCC_KRA_KEY

def extract_months_from_kra_filename(filename):
    """
    Extract quarter months from KRA filename
    Format: "KRA Milestones for [Month] [Month] [Month] [Year].xlsx"
    Example: "KRA Milestones for December January February 2026.xlsx"
    The year in the filename is the END year of the quarter
    """
    months_pattern = r'KRA\s+Milestones\s+for\s+((?:January|February|March|April|May|June|July|August|September|October|November|December)(?:\s+(?:January|February|March|April|May|June|July|August|September|October|November|December))*)\s+(\d{4})'
    
    match = re.search(months_pattern, filename, re.IGNORECASE)
    if match:
        months_str = match.group(1)
        end_year = int(match.group(2))
        
        # Extract individual months in order
        month_names = re.findall(r'January|February|March|April|May|June|July|August|September|October|November|December', 
                                 months_str, re.IGNORECASE)
        
        logger.info(f"Extracted from KRA filename '{filename}':")
        logger.info(f"  Months: {month_names}, End Year: {end_year}")
        
        return [m.capitalize() for m in month_names], end_year
    
    logger.warning(f"Could not parse KRA filename: {filename}")
    return None, None

def get_tracker_for_month(cos, report_month, month_year):
    """
    Get tracker file for a specific report month using the month-shift logic.
    Uses the LATEST tracker available that matches the required month/year.
    
    Args:
        cos: COS client
        report_month: The month shown in the report (e.g., "June")
        month_year: Year of the report month
    
    Returns:
        Path to tracker file or None
    """
    # Get the tracker month based on shift logic
    tracker_month_name = MONTH_SHIFT.get(report_month)
    
    if not tracker_month_name:
        logger.warning(f"No month shift mapping found for {report_month}")
        return None
    
    # Calculate the tracker year (handle year rollover)
    tracker_month_num = get_month_number(tracker_month_name)
    report_month_num = get_month_number(report_month)
    
    tracker_year = month_year
    # If tracker month is earlier in the year than report month, it's next year
    if tracker_month_num < report_month_num:
        tracker_year += 1
    
    logger.info(f"Looking for tracker for Report: {report_month} {month_year} → Tracker: {tracker_month_name} {tracker_year}")
    
    # List all files in Wave City Club folder
    all_files = list_files_in_folder(cos, TRACKER_FOLDER)
    logger.info(f"Total files in {TRACKER_FOLDER}: {len(all_files)}")
    
    # Pattern to match Structure Work tracker files
    tracker_pattern = r'Structure.*Work.*Tracker.*Wave.*City.*Club.*\.xlsx$'
    
    matching_files = []
    
    for file_path in all_files:
        filename = os.path.basename(file_path)
        if re.search(tracker_pattern, filename, re.IGNORECASE):
            # Extract date from filename
            file_date = extract_date_from_filename(filename)
            
            logger.info(f"  Found tracker: {filename}")
            
            if file_date:
                logger.info(f"    Extracted date: {file_date.strftime('%d-%m-%Y')} (Month: {file_date.month}, Year: {file_date.year})")
                logger.info(f"    Looking for: Month={tracker_month_num}, Year={tracker_year}")
                
                # Check if this file matches the month and year we're looking for
                if file_date.month == tracker_month_num and file_date.year == tracker_year:
                    matching_files.append((file_path, file_date))
                    logger.info(f"    ✅ MATCH!")
                else:
                    logger.info(f"    ❌ Date mismatch (need {tracker_month_name} {tracker_year})")
            else:
                logger.warning(f"    Could not extract date from {filename}")
    
    if matching_files:
        # Sort by date and get the LATEST matching tracker
        matching_files.sort(key=lambda x: x[1], reverse=True)
        latest_tracker = matching_files[0][0]
        logger.info(f"✅ Selected tracker: {os.path.basename(latest_tracker)} (dated {matching_files[0][1].strftime('%d-%m-%Y')})")
        return latest_tracker
    else:
        logger.warning(f"⚠️ No tracker found for {report_month} {month_year} (looking for {tracker_month_name} {tracker_year} dated files)")
        return None

def setup_quarterly_months(kra_filename):
    """
    Setup the quarterly months and years based on KRA filename.
    
    The year in the KRA filename represents the END year of the quarter.
    For cross-year quarters (e.g., Dec-Jan-Feb 2026):
    - December belongs to 2025
    - January belongs to 2026
    - February belongs to 2026
    """
    global MONTHS, MONTH_YEARS, TARGET_END_MONTH, TARGET_END_YEAR
    
    months, end_year = extract_months_from_kra_filename(kra_filename)
    
    if not months or not end_year:
        logger.error(f"Could not extract months from KRA filename: {kra_filename}")
        return False
    
    MONTHS = months
    TARGET_END_MONTH = MONTHS[-1]
    TARGET_END_YEAR = end_year
    
    # Determine starting year for the quarter
    first_month_num = get_month_number(MONTHS[0])
    last_month_num = get_month_number(MONTHS[-1])
    
    # For cross-year quarters (first month number > last month number),
    # the first month belongs to the previous year
    if first_month_num > last_month_num:
        # Cross-year quarter (e.g., Dec 2025, Jan 2026, Feb 2026)
        start_year = end_year - 1
        logger.info(f"Cross-year quarter detected: {MONTHS[0]} {start_year} - {MONTHS[-1]} {end_year}")
    else:
        # Same-year quarter (e.g., Jun 2025, Jul 2025, Aug 2025)
        start_year = end_year
        logger.info(f"Same-year quarter: all months in {start_year}")
    
    # Assign years to each month
    current_year = start_year
    for i, month in enumerate(MONTHS):
        month_num = get_month_number(month)
        
        # Year increments when month number decreases (year rollover)
        if i > 0:
            prev_month_num = get_month_number(MONTHS[i-1])
            if month_num < prev_month_num:
                current_year += 1
        
        MONTH_YEARS[month] = current_year
    
    logger.info(f"Month-Year mapping:")
    for month in MONTHS:
        logger.info(f"  {month}: {MONTH_YEARS[month]}")
    
    return True


def find_club_sheet(wb):
    """Find the sheet with 'Club' in its name"""
    for sheet_name in wb.sheetnames:
        if 'club' in sheet_name.lower():
            logger.info(f"Found Club sheet: {sheet_name}")
            return sheet_name
    
    logger.warning("No sheet with 'Club' in name found. Available sheets: " + ", ".join(wb.sheetnames))
    return None

# -----------------------------------------------------------------------------
# DATA EXTRACTION
# -----------------------------------------------------------------------------

def get_wcc_targets_from_kra(cos):
    """
    Extract targets from KRA milestone file with updated logic:
    - Look for sheet with "Club" in its name
    - For each activity, fetch both the target activity name AND target %
    - Store both values for comparison with tracker data
    """
    global WCC_KRA_KEY
    
    if not WCC_KRA_KEY:
        raise ValueError("WCC_KRA_KEY is not set. Call get_latest_kra_file() first.")
    
    logger.info(f"Reading KRA file: {WCC_KRA_KEY}")
    
    # Download and load KRA file
    kra_bytes = download_file_bytes(cos, WCC_KRA_KEY)
    kra_wb = load_workbook(BytesIO(kra_bytes), data_only=True)
    
    # Find the Club sheet
    club_sheet_name = find_club_sheet(kra_wb)
    if not club_sheet_name:
        # Fallback to active sheet
        logger.warning("Using active sheet as fallback")
        ws = kra_wb.active
    else:
        ws = kra_wb[club_sheet_name]
    
    targets = {}
    
    # Find header row (look for "Blocks" in column A)
    header_row = None
    for row_idx in range(1, min(10, ws.max_row + 1)):
        cell_value = ws.cell(row=row_idx, column=1).value
        if cell_value and 'block' in str(cell_value).lower():
            header_row = row_idx
            logger.info(f"Found header row at row {header_row}")
            break
    
    if not header_row:
        logger.error("Could not find header row with 'Blocks'")
        kra_wb.close()
        return targets
    
    # Start from row after header
    data_start_row = header_row + 1
    
    for row_idx in range(data_start_row, ws.max_row + 1):
        row = list(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
        
        # Extract block name from column A
        block_name = row[KRA_COLUMNS['BLOCK']]
        
        # Skip empty rows
        if not block_name or pd.isna(block_name):
            continue
        
        # Stop if we hit another header row or non-block data
        block_name_str = str(block_name).strip()
        if any(month.lower() in block_name_str.lower() for month in ['january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december']):
            logger.info(f"Stopping at row {row_idx}: Found header-like content '{block_name_str}'")
            break
        
        # Skip if block name is just "Activity" or other single-word non-block names
        if block_name_str.lower() in ['activity', 'activities', 'milestone', 'milestones', 'target', 'targets', 'blocks']:
            logger.info(f"Skipping non-block row: '{block_name_str}'")
            continue
            
        logger.info(f"Processing block: {block_name}")
        
        targets[block_name] = {}
        
        # Process each month
        for i, month in enumerate(MONTHS):
            month_year = MONTH_YEARS[month]
            
            # Calculate column indices for this month
            if i == 0:
                activity_col = KRA_COLUMNS['MONTH1_ACTIVITY']
                target_col = KRA_COLUMNS['MONTH1_TARGET']
            elif i == 1:
                activity_col = KRA_COLUMNS['MONTH2_ACTIVITY']
                target_col = KRA_COLUMNS['MONTH2_TARGET']
            elif i == 2:
                activity_col = KRA_COLUMNS['MONTH3_ACTIVITY']
                target_col = KRA_COLUMNS['MONTH3_TARGET']
            else:
                logger.warning(f"More than 3 months found, skipping month {month}")
                continue
            
            # Extract activity name and target %
            activity_name = row[activity_col] if activity_col < len(row) else None
            target_percent = row[target_col] if target_col < len(row) else None
            
            # Validate target_percent is a number, not text
            validated_target_percent = None
            if target_percent and not pd.isna(target_percent):
                try:
                    validated_target_percent = float(target_percent)
                except (ValueError, TypeError):
                    logger.warning(f"Invalid target % value for {block_name} - {month}: '{target_percent}' (not a number)")
                    validated_target_percent = None
            
            # Store both activity and target % for this month
            targets[block_name][month] = {
                'activity': activity_name if activity_name and not pd.isna(activity_name) else None,
                'target_percent': validated_target_percent
            }
            
            logger.info(f"  {month}: Activity='{activity_name}', Target %={validated_target_percent}")
    
    kra_wb.close()
    return targets

def find_activity_completion_in_tracker(tracker_wb, sheet_name, target_activity):
    """
    Search for an activity in the tracker and return its % complete value.
    
    Args:
        tracker_wb: Loaded tracker workbook
        sheet_name: Sheet name to search in
        target_activity: Activity name to find
    
    Returns:
        Float value of % complete, or None if not found
    """
    if sheet_name not in tracker_wb.sheetnames:
        logger.warning(f"Sheet '{sheet_name}' not found in tracker")
        return None
    
    ws = tracker_wb[sheet_name]
    
    # Search through rows starting from row 3 (skip headers)
    for row_idx in range(2, ws.max_row + 1):
        row = list(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
        
        # Get activity name from column G (index 6)
        if len(row) > TRACKER_COLUMNS['ACTIVITY_NAME']:
            activity_name = row[TRACKER_COLUMNS['ACTIVITY_NAME']]
            
            if activity_name and str(activity_name).strip().lower() == str(target_activity).strip().lower():
                # Found the activity, get % complete from column M (index 12)
                if len(row) > TRACKER_COLUMNS['PERCENT_COMPLETE']:
                    percent_complete = row[TRACKER_COLUMNS['PERCENT_COMPLETE']]
                    try:
                        return float(percent_complete) if percent_complete is not None else 0.0
                    except (ValueError, TypeError):
                        logger.warning(f"Could not convert % complete to float for activity '{activity_name}': {percent_complete}")
                        return 0.0
    
    logger.debug(f"Activity '{target_activity}' not found in sheet '{sheet_name}'")
    return None

def get_wcc_progress_from_tracker_all_months(cos, targets):
    """
    Extract progress data for all months with UPDATED LOGIC:
    - Weightage = 100 for all blocks
    - Weighted % = (% Complete / Weightage) * 100 for each month
    - If tracker % > target %, display 100%
    - Responsible and Delay columns shown once at the end
    """
    global TRACKER_PATHS, LOADED_TRACKERS
    
    # Load all trackers into memory
    for month in MONTHS:
        tracker_path = TRACKER_PATHS.get(month)
        if tracker_path and month not in LOADED_TRACKERS:
            logger.info(f"Loading tracker for {month}: {tracker_path}")
            tracker_bytes = download_file_bytes(cos, tracker_path)
            LOADED_TRACKERS[month] = load_workbook(BytesIO(tracker_bytes), data_only=True)
    
    # Build dataframe
    data = []
    milestone_counter = 1
    
    for block_name, month_targets in targets.items():
        # Get the tracker sheet name for this block
        tracker_sheet = BLOCK_MAPPING.get(block_name)
        
        if not tracker_sheet:
            logger.warning(f"No tracker sheet mapping found for block: {block_name}")
            continue
        
        row_data = {
            'Milestone': f'Milestone-{milestone_counter:02d}',
            'Block': block_name
        }
        
        # Determine the target to complete by the end of quarter
        last_month = MONTHS[-1]
        last_month_target_info = month_targets.get(last_month, {})
        target_to_complete = last_month_target_info.get('activity', 'N/A')
        row_data[f'Target to be complete by {last_month}-{MONTH_YEARS[last_month]}'] = target_to_complete
        
        # Process each month
        for month in MONTHS:
            month_year = MONTH_YEARS[month]
            
            # Get target info for this month
            target_info = month_targets.get(month, {})
            target_activity = target_info.get('activity')
            target_percent = target_info.get('target_percent')
            
            # Column names for this month
            target_col = f'Target - {month}-{month_year}'
            status_col = f'% work done- {month} Status'
            achieved_col = f'Achieved- {month} {month_year}'
            weightage_col = f'Weightage- {month}'
            weighted_pct_col = f'Weighted %- {month}'
            
            row_data[target_col] = target_activity if target_activity else 'No target'
            
            # Check if tracker is available for this month
            if month not in LOADED_TRACKERS:
                # Leave blank if tracker not available
                row_data[status_col] = ''
                row_data[achieved_col] = ''
                row_data[weightage_col] = ''
                row_data[weighted_pct_col] = ''
                continue
            
            # If no target activity, mark as N/A
            if not target_activity:
                row_data[status_col] = 'N/A'
                row_data[achieved_col] = 'No target for this month'
                row_data[weightage_col] = ''
                row_data[weighted_pct_col] = ''
                continue

            # Check if target is explicitly "No target"
            if str(target_activity).strip().lower() == 'no target':
                row_data[status_col] = '100%'
                row_data[achieved_col] = 'No target specified'
                
                # Calculate Weightage
                row_data[weightage_col] = 100
                row_data[weighted_pct_col] = '100%'
                continue
            
            # Fetch completed % from tracker
            tracker_wb = LOADED_TRACKERS[month]
            completed_percent = find_activity_completion_in_tracker(tracker_wb, tracker_sheet, target_activity)
            
            # Determine status based on comparison
            if completed_percent is None:
                # Activity not found in tracker
                row_data[status_col] = '0%'
                row_data[achieved_col] = 'Activity not found in tracker'
                status_percent = 0.0
            elif (target_percent is None or target_percent == 0) and completed_percent > 0:
                 # No target or 0 target, but tracker has progress -> 100%
                 row_data[status_col] = '100%'
                 row_data[achieved_col] = f'Target achieved (Tracker: {completed_percent*100:.0f}%)'
                 status_percent = 1.0
            elif target_percent is not None and completed_percent > target_percent:
                # Tracker % > Target % -> display 100%
                row_data[status_col] = '100%'
                row_data[achieved_col] = f'Target exceeded ({completed_percent*100:.0f}% > {target_percent*100:.0f}%)'
                status_percent = 1.0
            elif target_percent is not None and completed_percent == target_percent:
                # Target % matches completed %
                row_data[status_col] = '100%'
                row_data[achieved_col] = f'Target achieved ({completed_percent*100:.0f}% complete)'
                status_percent = 1.0
            else:
                # Display completed % from tracker
                status_percent = completed_percent
                row_data[status_col] = f'{completed_percent*100:.0f}%'
                
                if target_percent is not None and completed_percent >= target_percent:
                    row_data[achieved_col] = f'{completed_percent*100:.0f}% completed (Target: {target_percent*100:.0f}%)'
                elif target_percent is not None:
                    row_data[achieved_col] = f'{completed_percent*100:.0f}% completed (Target: {target_percent*100:.0f}%)'
                else:
                    row_data[achieved_col] = f'{completed_percent*100:.0f}% completed'
            
            # Calculate Weightage and Weighted % for this month
            weightage = 100  # Always 100
            weighted_pct = (status_percent * 100 / weightage) * 100 if weightage > 0 else 0
            row_data[weightage_col] = weightage
            row_data[weighted_pct_col] = f'{weighted_pct:.0f}%'
        
        # Add Responsible and Delay columns once at the end
        row_data['Responsible'] = ''
        row_data['Delay Reason'] = ''
        
        data.append(row_data)
        milestone_counter += 1
    
    # Add summary row with average Weighted % for each month
    summary_row = {'Milestone': 'AVERAGE WEIGHTED %', 'Block': ''}
    summary_row[f'Target to be complete by {MONTHS[-1]}-{MONTH_YEARS[MONTHS[-1]]}'] = ''
    
    for month in MONTHS:
        month_year = MONTH_YEARS[month]
        
        # Calculate average of Weighted % for this month
        weighted_values = []
        for row in data:
            weighted_val = row.get(f'Weighted %- {month}', '')
            if weighted_val and weighted_val != '':
                try:
                    val = float(str(weighted_val).replace('%', ''))
                    weighted_values.append(val)
                except (ValueError, TypeError):
                    pass
        
        # Calculate average
        if weighted_values:
            avg_weighted = sum(weighted_values) / len(weighted_values)
            summary_row[f'Weighted %- {month}'] = f'{avg_weighted:.1f}%'
        else:
            summary_row[f'Weighted %- {month}'] = ''
        
        # Leave other columns blank for summary row
        summary_row[f'Target - {month}-{month_year}'] = ''
        summary_row[f'% work done- {month} Status'] = ''
        summary_row[f'Achieved- {month} {month_year}'] = ''
        summary_row[f'Weightage- {month}'] = ''
    
    summary_row['Responsible'] = ''
    summary_row['Delay Reason'] = ''
    
    data.append(summary_row)
    
    # Create DataFrame
    df = pd.DataFrame(data)
    
    # Reorder columns to match format
    column_order = ['Milestone', 'Block', f'Target to be complete by {MONTHS[-1]}-{MONTH_YEARS[MONTHS[-1]]}']
    
    for month in MONTHS:
        month_year = MONTH_YEARS[month]
        column_order.extend([
            f'Target - {month}-{month_year}',
            f'% work done- {month} Status',
            f'Achieved- {month} {month_year}',
            f'Weightage- {month}',
            f'Weighted %- {month}'
        ])
    
    column_order.extend(['Responsible', 'Delay Reason'])
    
    df = df[column_order]
    
    return df

def apply_manual_overrides(df):
    """
    Apply any manual overrides if needed
    This function can be customized based on business rules
    """
    logger.info("Applying manual overrides (if any)...")
    return df

# -----------------------------------------------------------------------------
# REPORT GENERATION
# -----------------------------------------------------------------------------

def write_wcc_excel_report_consolidated(df, filename):
    """Generate Excel report with proper formatting matching the format file"""
    
    logger.info(f'Generating consolidated report: {filename}')
    
    wb = Workbook()
    ws = wb.active
    ws.title = 'WCC Progress'
    
    # Define styles
    header_font = Font(name='Calibri', size=11, bold=True)
    normal_font = Font(name='Calibri', size=10)
    date_font = Font(name='Calibri', size=10, bold=True)
    summary_font = Font(name='Calibri', size=11, bold=True)
    
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    light_grey_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    light_blue_fill = PatternFill(start_color='B4C7E7', end_color='B4C7E7', fill_type='solid')
    summary_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    
    # Title row
    ws.merge_cells('A1:T1')
    ws['A1'] = 'Wave City Club- Progress Against Milestones'
    ws['A1'].font = Font(name='Calibri', size=14, bold=True)
    ws['A1'].alignment = center
    
    # Date row
    current_date = datetime.now().strftime('%d-%m-%Y')
    ws.merge_cells('A2:T2')
    ws['A2'] = f'Report Generated on: {current_date}'
    ws['A2'].font = date_font
    ws['A2'].alignment = center
    
    # Month info row
    month_info = ', '.join([f'{m} {MONTH_YEARS[m]}' for m in MONTHS])
    ws.merge_cells('A3:T3')
    ws['A3'] = f'Months Covered: {month_info}'
    ws['A3'].font = date_font
    ws['A3'].alignment = center
    
    # Empty row
    ws.merge_cells('A4:T4')
    
    # Write dataframe starting at row 5
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=5):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
    
    # Style header row (row 5)
    header_row = 5
    for cell in ws[header_row]:
        cell.font = header_font
        cell.alignment = center
        cell.border = border
        cell.fill = light_grey_fill
    
    # Style data rows
    data_start = 6
    summary_row_idx = ws.max_row
    data_end = summary_row_idx - 1
    
    for row_num in range(data_start, data_end + 1):
        for col_num in range(1, len(df.columns) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.font = normal_font
            cell.border = border
            
            # Alignment based on column type
            if col_num in [1, 2, 3] or 'Target' in str(ws.cell(row=header_row, column=col_num).value or ''):
                cell.alignment = left
            else:
                cell.alignment = center
    
    # Style summary row (last row)
    for col_num in range(1, len(df.columns) + 1):
        cell = ws.cell(row=summary_row_idx, column=col_num)
        cell.font = summary_font
        cell.border = border
        cell.fill = summary_fill
        cell.alignment = center
    
    # Dynamic column width adjustment
    for col_num in range(1, len(df.columns) + 1):
        col_letter = get_column_letter(col_num)
        
        max_length = 0
        for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=col_num, max_col=col_num):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        
        calculated_width = min(max(max_length + 2, 8), 15)
        ws.column_dimensions[col_letter].width = calculated_width
    
    # Set row heights
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 20
    for i in range(5, ws.max_row + 1):
        ws.row_dimensions[i].height = 25
    ws.row_dimensions[summary_row_idx].height = 30  # Make summary row taller
    
    wb.save(filename)
    logger.info(f'Report saved to {filename}')

def get_unique_filename(base_name):
    """If file exists, append (1), (2), etc."""
    if not os.path.exists(base_name):
        return base_name

    name, ext = os.path.splitext(base_name)
    counter = 1
    new_name = f"{name}({counter}){ext}"
    while os.path.exists(new_name):
        counter += 1
        new_name = f"{name}({counter}){ext}"
    return new_name

# -----------------------------------------------------------------------------
# MAIN FUNCTION
# -----------------------------------------------------------------------------

def main():
    """Main execution function for quarterly report generation"""
    logger.info("=== STARTING WAVE CITY CLUB QUARTERLY REPORT GENERATION (UPDATED LOGIC) ===")
    
    try:
        # Initialize COS client
        cos = init_cos()
        
        # Step 1: Get latest KRA file from Milestone/ folder
        logger.info("\n=== STEP 1: Finding Latest KRA File ===")
        kra_file = get_latest_kra_file(cos)
        
        if not kra_file:
            logger.error("❌ Failed to find KRA file")
            return
        
        logger.info(f"✅ Using KRA file: {kra_file}")
        
        # Step 2: Extract months from KRA filename and setup
        logger.info("\n=== STEP 2: Setting Up Quarterly Months ===")
        if not setup_quarterly_months(os.path.basename(kra_file)):
            logger.error("❌ Failed to setup quarterly months")
            return
        
        logger.info(f"✅ Quarterly months: {MONTHS}")
        logger.info(f"✅ Month-Year mapping: {MONTH_YEARS}")
        
        # Step 3: Find appropriate trackers for each month from Wave City Club/ folder
        logger.info("\n=== STEP 3: Finding Trackers for Each Month ===")
        global TRACKER_PATHS
        
        for month in MONTHS:
            month_year = MONTH_YEARS[month]
            tracker_path = get_tracker_for_month(cos, month, month_year)
            
            if tracker_path:
                TRACKER_PATHS[month] = tracker_path
                logger.info(f"✅ {month} {month_year}: {tracker_path}")
            else:
                logger.warning(f"⚠️ {month} {month_year}: No tracker found - column will be blank in report")
                TRACKER_PATHS[month] = None
        
        # Continue even if no trackers found
        logger.info(f"Found trackers for {sum(1 for v in TRACKER_PATHS.values() if v)} out of {len(MONTHS)} months")
        
        # Step 4: Get targets from KRA file (Club sheet)
        logger.info("\n=== STEP 4: Extracting Targets from KRA File ===")
        targets = get_wcc_targets_from_kra(cos)
        
        if not targets:
            logger.error("❌ Failed to extract targets from KRA file")
            return
        
        logger.info(f"✅ Extracted targets for {len(targets)} blocks")
        
        # Step 5: Extract progress data for all months with NEW LOGIC
        logger.info("\n=== STEP 5: Extracting Progress Data from Trackers (NEW LOGIC) ===")
        df = get_wcc_progress_from_tracker_all_months(cos, targets)
        
        if df.empty:
            logger.error("❌ Failed to generate progress data")
            return
        
        logger.info(f"✅ Generated progress data for {len(df)-1} milestones + 1 summary row")
        
        # Step 5.5: Apply manual overrides
        logger.info("\n=== STEP 5.5: Applying Manual Overrides ===")
        df = apply_manual_overrides(df)
        
        # Step 6: Generate Excel report
        logger.info("\n=== STEP 6: Generating Excel Report ===")
        current_date_for_filename = datetime.now().strftime('%d-%m-%Y')
        
        # Create filename with quarter months
        quarter_str = "_".join(MONTHS)
        base_filename = f"Wave_City_Club_Milestone_Report_{quarter_str}_{current_date_for_filename}.xlsx"
        filename = get_unique_filename(base_filename)
        
        write_wcc_excel_report_consolidated(df, filename)
        
        logger.info("\n=== WAVE CITY CLUB QUARTERLY REPORT GENERATION COMPLETE ===")
        logger.info(f"✅ Report saved as: {filename}")
        
        # Log summary
        logger.info("\n=== REPORT SUMMARY ===")
        logger.info(f"  Quarter Months: {', '.join([f'{m} {MONTH_YEARS[m]}' for m in MONTHS])}")
        logger.info(f"  KRA File: {os.path.basename(kra_file)}")
        logger.info(f"  Total Blocks: {len(targets)}")
        logger.info(f"  Trackers Used:")
        for month in MONTHS:
            tracker = TRACKER_PATHS.get(month)
            if tracker:
                tracker_date = extract_date_from_filename(os.path.basename(tracker))
                logger.info(f"    - {month} {MONTH_YEARS[month]}: {os.path.basename(tracker)} (dated {tracker_date.strftime('%d-%m-%Y') if tracker_date else 'Unknown'})")
            else:
                logger.info(f"    - {month} {MONTH_YEARS[month]}: Not Available (column will be blank)")
        logger.info(f"  Total Milestones: {len(df)-1} + 1 summary row")
        
    except Exception as e:
        logger.error(f"❌ Error in main execution: {e}")
        import traceback
        logger.error(traceback.format_exc())
        raise

if __name__ == "__main__":
    main()


















































# import os
# import re
# import logging
# from io import BytesIO
# from datetime import datetime
# from dateutil.relativedelta import relativedelta

# import pandas as pd
# from openpyxl import Workbook, load_workbook
# from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
# from openpyxl.utils import get_column_letter
# from openpyxl.utils.dataframe import dataframe_to_rows
# from dotenv import load_dotenv
# import ibm_boto3
# from ibm_botocore.client import Config

# # -----------------------------------------------------------------------------
# # CONFIG / CONSTANTS
# # -----------------------------------------------------------------------------
# load_dotenv()
# logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(levelname)s] %(message)s')
# logger = logging.getLogger(__name__)

# # Validate required environment variables
# required = {
#     'COS_API_KEY': os.getenv('COS_API_KEY'),
#     'COS_SERVICE_INSTANCE_CRN': os.getenv('COS_SERVICE_INSTANCE_CRN'),
#     'COS_ENDPOINT': os.getenv('COS_ENDPOINT'),
#     'COS_BUCKET_NAME': os.getenv('COS_BUCKET_NAME'),
# }
# missing = [k for k, v in required.items() if not v]
# if missing:
#     logger.error(f"Missing required environment variables: {', '.join(missing)}")
#     raise SystemExit(1)

# COS_API_KEY     = required['COS_API_KEY']
# COS_CRN         = required['COS_SERVICE_INSTANCE_CRN']
# COS_ENDPOINT    = required['COS_ENDPOINT']
# BUCKET          = required['COS_BUCKET_NAME']

# # Folder paths in COS
# KRA_FOLDER = "Milestone/"
# TRACKER_FOLDER = "Wave City Club/"

# # Dynamic KRA file path
# WCC_KRA_KEY = None

# # Dynamic tracker paths
# TRACKER_PATHS = {}  # Maps month names to tracker file paths
# LOADED_TRACKERS = {}  # Cache loaded workbooks

# # Dynamic months and years
# MONTHS = []
# MONTH_YEARS = {}  # Maps month name to year
# TARGET_END_MONTH = None
# TARGET_END_YEAR = None

# # Quarterly definitions (for reference)
# QUARTERLY_GROUPS = [
#     ['June', 'July', 'August'],           # Q1
#     ['September', 'October', 'November'], # Q2
#     ['December', 'January', 'February'],  # Q3
#     ['March', 'April', 'May']             # Q4
# ]

# # Month shift logic: Report Month -> Tracker Month
# MONTH_SHIFT = {
#     'June': 'July',
#     'July': 'August',
#     'August': 'September',
#     'September': 'October',
#     'October': 'November',
#     'November': 'December',
#     'December': 'January',
#     'January': 'February',
#     'February': 'March',
#     'March': 'April',
#     'April': 'May',
#     'May': 'June'
# }

# # Block mapping from KRA to tracker sheets
# BLOCK_MAPPING = {
#     'Block 1 (B1) Banquet Hall': 'B1 Banket Hall & Finedine ',
#     'Fine Dine': 'B1 Banket Hall & Finedine ',
#     'Block 5 (B5) Admin + Member Lounge+Creche+Av Room + Surveillance Room +Toilets': 'B5',
#     'Block 6 (B6) Toilets': 'B6',
#     'Block 7(B7) Indoor Sports': 'B7',
#     'Block 9 (B9) Spa & Saloon': 'B9',
#     'Block 8 (B8) Squash Court': 'B8',
#     'Block 2 & 3 (B2 & B3) Cafe & Bar': 'B2 & B3',
#     'Block 4 (B4) Indoor Swimming Pool Changing Room & Toilets': 'B4',
#     'Block 11 (B11) Guest House': 'B11',
#     'Block 10 (B10) Gym': 'B10'
# }

# # KRA Column structure (0-indexed)
# KRA_COLUMNS = {
#     'BLOCK': 0,          # Column A: Block name
#     'MONTH1_ACTIVITY': 1,  # Column B: First month activity
#     'MONTH1_TARGET': 2,    # Column C: First month target %
#     'MONTH2_ACTIVITY': 3,  # Column D: Second month activity
#     'MONTH2_TARGET': 4,    # Column E: Second month target %
#     'MONTH3_ACTIVITY': 5,  # Column F: Third month activity
#     'MONTH3_TARGET': 6,    # Column G: Third month target %
# }

# # Tracker column positions (0-indexed)
# TRACKER_COLUMNS = {
#     'ACTIVITY_NAME': 6,    # Column G: Activity Name
#     'PERCENT_COMPLETE': 12  # Column M: % Complete
# }

# # -----------------------------------------------------------------------------
# # COS HELPERS
# # -----------------------------------------------------------------------------

# def init_cos():
#     return ibm_boto3.client(
#         's3',
#         ibm_api_key_id=COS_API_KEY,
#         ibm_service_instance_id=COS_CRN,
#         config=Config(signature_version='oauth'),
#         endpoint_url=COS_ENDPOINT,
#     )

# def download_file_bytes(cos, key):
#     if not key:
#         raise ValueError("File key cannot be None or empty")
#     obj = cos.get_object(Bucket=BUCKET, Key=key)
#     return obj['Body'].read()

# def list_files_in_folder(cos, folder_prefix):
#     """List all files in a specific folder (prefix) in the COS bucket"""
#     try:
#         response = cos.list_objects_v2(Bucket=BUCKET, Prefix=folder_prefix)
#         files = []
#         if 'Contents' in response:
#             for obj in response['Contents']:
#                 if not obj['Key'].endswith('/'):
#                     files.append(obj['Key'])
#         return files
#     except Exception as e:
#         logger.error(f"Error listing files in folder {folder_prefix}: {e}")
#         return []

# def extract_date_from_filename(filename):
#     """Extract date from filename in format (dd-mm-yyyy)"""
#     pattern = r'\((\d{2}-\d{2}-\d{4})\)'
#     match = re.search(pattern, filename)
#     if match:
#         date_str = match.group(1)
#         try:
#             return datetime.strptime(date_str, '%d-%m-%Y')
#         except ValueError:
#             logger.warning(f"Could not parse date {date_str} from filename {filename}")
#             return None
#     return None

# def get_month_name(month_num):
#     """Convert month number to month name"""
#     months = {
#         1: "January", 2: "February", 3: "March", 4: "April",
#         5: "May", 6: "June", 7: "July", 8: "August", 
#         9: "September", 10: "October", 11: "November", 12: "December"
#     }
#     return months.get(month_num, "Unknown")

# def get_month_number(month_name):
#     """Convert month name to month number"""
#     months = {
#         "January": 1, "February": 2, "March": 3, "April": 4,
#         "May": 5, "June": 6, "July": 7, "August": 8, 
#         "September": 9, "October": 10, "November": 11, "December": 12
#     }
#     return months.get(month_name, 1)

# def get_latest_kra_file(cos):
#     """Get the latest KRA milestone file from the Milestone/ folder"""
#     global WCC_KRA_KEY
    
#     logger.info("=== FINDING LATEST KRA MILESTONE FILE ===")
    
#     # List all files in Milestone folder
#     all_files = list_files_in_folder(cos, KRA_FOLDER)
#     logger.info(f"Found {len(all_files)} files in {KRA_FOLDER} folder")
    
#     # Pattern to match KRA milestone files
#     kra_pattern = r'KRA Milestones for.*\.xlsx$'
    
#     matching_files = []
    
#     for file_path in all_files:
#         filename = os.path.basename(file_path)
#         if re.search(kra_pattern, filename, re.IGNORECASE):
#             logger.info(f"Found KRA file: {filename}")
#             matching_files.append(file_path)
    
#     if matching_files:
#         # Sort by filename to get the latest
#         latest_file = sorted(matching_files)[-1]
#         WCC_KRA_KEY = latest_file
#         logger.info(f"✅ Latest KRA file: {WCC_KRA_KEY}")
#     else:
#         logger.error(f"❌ No KRA milestone files found in {KRA_FOLDER} folder!")
#         WCC_KRA_KEY = None
    
#     return WCC_KRA_KEY

# def extract_months_from_kra_filename(filename):
#     """Extract quarter months from KRA filename"""
#     months_pattern = r'for\s+((?:January|February|March|April|May|June|July|August|September|October|November|December)(?:\s+(?:January|February|March|April|May|June|July|August|September|October|November|December))*)\s+(\d{4})'
    
#     match = re.search(months_pattern, filename, re.IGNORECASE)
#     if match:
#         months_str = match.group(1)
#         year = int(match.group(2))
        
#         # Extract individual months
#         month_names = re.findall(r'January|February|March|April|May|June|July|August|September|October|November|December', 
#                                  months_str, re.IGNORECASE)
        
#         return [m.capitalize() for m in month_names], year
    
#     return None, None

# def get_tracker_for_month(cos, report_month, month_year):
#     """
#     Get tracker file for a specific report month using the month-shift logic.
    
#     Args:
#         cos: COS client
#         report_month: The month shown in the report (e.g., "June")
#         month_year: Year of the report month
    
#     Returns:
#         Path to tracker file or None
#     """
#     # Get the tracker month based on shift logic
#     tracker_month_name = MONTH_SHIFT.get(report_month)
    
#     if not tracker_month_name:
#         logger.warning(f"No month shift mapping found for {report_month}")
#         return None
    
#     # Calculate the tracker year (handle year rollover)
#     tracker_month_num = get_month_number(tracker_month_name)
#     report_month_num = get_month_number(report_month)
    
#     tracker_year = month_year
#     # If tracker month is earlier in the year than report month, it's next year
#     if tracker_month_num < report_month_num:
#         tracker_year += 1
    
#     logger.info(f"Looking for tracker for Report: {report_month} {month_year} → Tracker: {tracker_month_name} {tracker_year}")
    
#     # List all files in Wave City Club folder
#     all_files = list_files_in_folder(cos, TRACKER_FOLDER)
    
#     # Pattern to match Structure Work tracker files
#     tracker_pattern = r'Structure.*Work.*Tracker.*Wave.*City.*Club.*\.xlsx$'
    
#     matching_files = []
    
#     for file_path in all_files:
#         filename = os.path.basename(file_path)
#         if re.search(tracker_pattern, filename, re.IGNORECASE):
#             # Extract date from filename
#             file_date = extract_date_from_filename(filename)
#             if file_date and file_date.month == tracker_month_num and file_date.year == tracker_year:
#                 matching_files.append((file_path, file_date))
#                 logger.info(f"Found matching tracker: {filename} (dated {file_date.strftime('%d-%m-%Y')})")
    
#     if matching_files:
#         # Sort by date and get the latest
#         matching_files.sort(key=lambda x: x[1], reverse=True)
#         latest_tracker = matching_files[0][0]
#         logger.info(f"✅ Selected tracker: {os.path.basename(latest_tracker)}")
#         return latest_tracker
#     else:
#         logger.warning(f"⚠️ No tracker found for {report_month} {month_year} (looking for {tracker_month_name} {tracker_year} dated files)")
#         return None

# def setup_quarterly_months(kra_filename):
#     """Setup the quarterly months and years based on KRA filename"""
#     global MONTHS, MONTH_YEARS, TARGET_END_MONTH, TARGET_END_YEAR
    
#     months, year = extract_months_from_kra_filename(kra_filename)
    
#     if not months or not year:
#         logger.error(f"Could not extract months from KRA filename: {kra_filename}")
#         return False
    
#     MONTHS = months
#     TARGET_END_MONTH = MONTHS[-1]
#     TARGET_END_YEAR = year
    
#     # Assign years to each month (handle cross-year quarters)
#     for i, month in enumerate(MONTHS):
#         month_num = get_month_number(month)
        
#         # Handle December-January-February quarter
#         if i > 0:
#             prev_month_num = get_month_number(MONTHS[i-1])
#             # If current month number is less than previous, we've crossed into new year
#             if month_num < prev_month_num:
#                 year += 1
        
#         MONTH_YEARS[month] = year
    
#     return True

# def find_club_sheet(wb):
#     """Find the sheet with 'Club' in its name"""
#     for sheet_name in wb.sheetnames:
#         if 'club' in sheet_name.lower():
#             logger.info(f"Found Club sheet: {sheet_name}")
#             return sheet_name
    
#     logger.warning("No sheet with 'Club' in name found. Available sheets: " + ", ".join(wb.sheetnames))
#     return None

# # -----------------------------------------------------------------------------
# # DATA EXTRACTION
# # -----------------------------------------------------------------------------

# def get_wcc_targets_from_kra(cos):
#     """
#     Extract targets from KRA milestone file with updated logic:
#     - Look for sheet with "Club" in its name
#     - For each activity, fetch both the target activity name AND target %
#     - Store both values for comparison with tracker data
#     """
#     global WCC_KRA_KEY
    
#     if not WCC_KRA_KEY:
#         raise ValueError("WCC_KRA_KEY is not set. Call get_latest_kra_file() first.")
    
#     logger.info(f"Reading KRA file: {WCC_KRA_KEY}")
    
#     # Download and load KRA file
#     kra_bytes = download_file_bytes(cos, WCC_KRA_KEY)
#     kra_wb = load_workbook(BytesIO(kra_bytes), data_only=True)
    
#     # Find the Club sheet
#     club_sheet_name = find_club_sheet(kra_wb)
#     if not club_sheet_name:
#         # Fallback to active sheet
#         logger.warning("Using active sheet as fallback")
#         ws = kra_wb.active
#     else:
#         ws = kra_wb[club_sheet_name]
    
#     targets = {}
    
#     # Find header row (look for "Blocks" in column A)
#     header_row = None
#     for row_idx in range(1, min(10, ws.max_row + 1)):
#         cell_value = ws.cell(row=row_idx, column=1).values
#         if cell_value and 'block' in str(cell_value).lower():
#             header_row = row_idx
#             logger.info(f"Found header row at row {header_row}")
#             break
    
#     if not header_row:
#         logger.error("Could not find header row with 'Blocks'")
#         kra_wb.close()
#         return targets
    
#     # Start from row after header
#     data_start_row = header_row + 1
    
#     for row_idx in range(data_start_row, ws.max_row + 1):
#         row = list(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
        
#         # Extract block name from column A
#         block_name = row[KRA_COLUMNS['BLOCK']]
        
#         # Skip empty rows
#         if not block_name or pd.isna(block_name):
#             continue
        
#         # Stop if we hit another header row or non-block data
#         block_name_str = str(block_name).strip()
#         if any(month.lower() in block_name_str.lower() for month in ['january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december']):
#             logger.info(f"Stopping at row {row_idx}: Found header-like content '{block_name_str}'")
#             break
        
#         # Skip if block name is just "Activity" or other single-word non-block names
#         if block_name_str.lower() in ['activity', 'activities', 'milestone', 'milestones', 'target', 'targets', 'blocks']:
#             logger.info(f"Skipping non-block row: '{block_name_str}'")
#             continue
            
#         logger.info(f"Processing block: {block_name}")
        
#         targets[block_name] = {}
        
#         # Process each month
#         for i, month in enumerate(MONTHS):
#             month_year = MONTH_YEARS[month]
            
#             # Calculate column indices for this month
#             if i == 0:
#                 activity_col = KRA_COLUMNS['MONTH1_ACTIVITY']
#                 target_col = KRA_COLUMNS['MONTH1_TARGET']
#             elif i == 1:
#                 activity_col = KRA_COLUMNS['MONTH2_ACTIVITY']
#                 target_col = KRA_COLUMNS['MONTH2_TARGET']
#             elif i == 2:
#                 activity_col = KRA_COLUMNS['MONTH3_ACTIVITY']
#                 target_col = KRA_COLUMNS['MONTH3_TARGET']
#             else:
#                 logger.warning(f"More than 3 months found, skipping month {month}")
#                 continue
            
#             # Extract activity name and target %
#             activity_name = row[activity_col] if activity_col < len(row) else None
#             target_percent = row[target_col] if target_col < len(row) else None
            
#             # Validate target_percent is a number, not text
#             validated_target_percent = None
#             if target_percent and not pd.isna(target_percent):
#                 try:
#                     validated_target_percent = float(target_percent)
#                 except (ValueError, TypeError):
#                     logger.warning(f"Invalid target % value for {block_name} - {month}: '{target_percent}' (not a number)")
#                     validated_target_percent = None
            
#             # Store both activity and target % for this month
#             targets[block_name][month] = {
#                 'activity': activity_name if activity_name and not pd.isna(activity_name) else None,
#                 'target_percent': validated_target_percent
#             }
            
#             logger.info(f"  {month}: Activity='{activity_name}', Target %={validated_target_percent}")
    
#     kra_wb.close()
#     return targets

# def find_activity_completion_in_tracker(tracker_wb, sheet_name, target_activity):
#     """
#     Search for an activity in the tracker and return its % complete value.
    
#     Args:
#         tracker_wb: Loaded tracker workbook
#         sheet_name: Sheet name to search in
#         target_activity: Activity name to find
    
#     Returns:
#         Float value of % complete, or None if not found
#     """
#     if sheet_name not in tracker_wb.sheetnames:
#         logger.warning(f"Sheet '{sheet_name}' not found in tracker")
#         return None
    
#     ws = tracker_wb[sheet_name]
    
#     # Search through rows starting from row 3 (skip headers)
#     for row_idx in range(2, ws.max_row + 1):
#         row = list(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
        
#         # Get activity name from column G (index 6)
#         if len(row) > TRACKER_COLUMNS['ACTIVITY_NAME']:
#             activity_name = row[TRACKER_COLUMNS['ACTIVITY_NAME']]
            
#             if activity_name and str(activity_name).strip().lower() == str(target_activity).strip().lower():
#                 # Found the activity, get % complete from column M (index 12)
#                 if len(row) > TRACKER_COLUMNS['PERCENT_COMPLETE']:
#                     percent_complete = row[TRACKER_COLUMNS['PERCENT_COMPLETE']]
#                     try:
#                         return float(percent_complete) if percent_complete is not None else 0.0
#                     except (ValueError, TypeError):
#                         logger.warning(f"Could not convert % complete to float for activity '{activity_name}': {percent_complete}")
#                         return 0.0
    
#     logger.debug(f"Activity '{target_activity}' not found in sheet '{sheet_name}'")
#     return None

# def get_wcc_progress_from_tracker_all_months(cos, targets):
#     """
#     Extract progress data for all months with UPDATED LOGIC:
#     - Weightage = 100 for all blocks
#     - Weighted % = (% Complete / Weightage) * 100 for each month
#     - If tracker % > target %, display 100%
#     - Responsible and Delay columns shown once at the end
#     """
#     global TRACKER_PATHS, LOADED_TRACKERS
    
#     # Load all trackers into memory
#     for month in MONTHS:
#         tracker_path = TRACKER_PATHS.get(month)
#         if tracker_path and month not in LOADED_TRACKERS:
#             logger.info(f"Loading tracker for {month}: {tracker_path}")
#             tracker_bytes = download_file_bytes(cos, tracker_path)
#             LOADED_TRACKERS[month] = load_workbook(BytesIO(tracker_bytes), data_only=True)
    
#     # Build dataframe
#     data = []
#     milestone_counter = 1
    
#     for block_name, month_targets in targets.items():
#         # Get the tracker sheet name for this block
#         tracker_sheet = BLOCK_MAPPING.get(block_name)
        
#         if not tracker_sheet:
#             logger.warning(f"No tracker sheet mapping found for block: {block_name}")
#             continue
        
#         row_data = {
#             'Milestone': f'Milestone-{milestone_counter:02d}',
#             'Block': block_name
#         }
        
#         # Determine the target to complete by the end of quarter
#         last_month = MONTHS[-1]
#         last_month_target_info = month_targets.get(last_month, {})
#         target_to_complete = last_month_target_info.get('activity', 'N/A')
#         row_data[f'Target to be complete by {last_month}-{MONTH_YEARS[last_month]}'] = target_to_complete
        
#         # Process each month
#         for month in MONTHS:
#             month_year = MONTH_YEARS[month]
            
#             # Get target info for this month
#             target_info = month_targets.get(month, {})
#             target_activity = target_info.get('activity')
#             target_percent = target_info.get('target_percent')
            
#             # Column names for this month
#             target_col = f'Target - {month}-{month_year}'
#             status_col = f'% work done- {month} Status'
#             achieved_col = f'Achieved- {month} {month_year}'
#             weightage_col = f'Weightage- {month}'
#             weighted_pct_col = f'Weighted %- {month}'
            
#             row_data[target_col] = target_activity if target_activity else 'No target'
            
#             # Check if tracker is available for this month
#             if month not in LOADED_TRACKERS:
#                 # Leave blank if tracker not available
#                 row_data[status_col] = ''
#                 row_data[achieved_col] = ''
#                 row_data[weightage_col] = ''
#                 row_data[weighted_pct_col] = ''
#                 continue
            
#             # If no target activity, mark as N/A
#             if not target_activity:
#                 row_data[status_col] = 'N/A'
#                 row_data[achieved_col] = 'No target for this month'
#                 row_data[weightage_col] = ''
#                 row_data[weighted_pct_col] = ''
#                 continue
            
#             # Fetch completed % from tracker
#             tracker_wb = LOADED_TRACKERS[month]
#             completed_percent = find_activity_completion_in_tracker(tracker_wb, tracker_sheet, target_activity)
            
#             # Determine status based on comparison
#             if completed_percent is None:
#                 # Activity not found in tracker
#                 row_data[status_col] = '0%'
#                 row_data[achieved_col] = 'Activity not found in tracker'
#                 status_percent = 0.0
#             elif target_percent is not None and completed_percent > target_percent:
#                 # Tracker % > Target % -> display 100%
#                 row_data[status_col] = '100%'
#                 row_data[achieved_col] = f'Target exceeded ({completed_percent*100:.0f}% > {target_percent*100:.0f}%)'
#                 status_percent = 1.0
#             elif target_percent is not None and completed_percent == target_percent:
#                 # Target % matches completed %
#                 row_data[status_col] = '100%'
#                 row_data[achieved_col] = f'Target achieved ({completed_percent*100:.0f}% complete)'
#                 status_percent = 1.0
#             else:
#                 # Display completed % from tracker
#                 status_percent = completed_percent
#                 row_data[status_col] = f'{completed_percent*100:.0f}%'
                
#                 if target_percent is not None and completed_percent >= target_percent:
#                     row_data[achieved_col] = f'{completed_percent*100:.0f}% completed (Target: {target_percent*100:.0f}%)'
#                 elif target_percent is not None:
#                     row_data[achieved_col] = f'{completed_percent*100:.0f}% completed (Target: {target_percent*100:.0f}%)'
#                 else:
#                     row_data[achieved_col] = f'{completed_percent*100:.0f}% completed'
            
#             # Calculate Weightage and Weighted % for this month
#             weightage = 100  # Always 100
#             weighted_pct = (status_percent * 100 / weightage) * 100 if weightage > 0 else 0
#             row_data[weightage_col] = weightage
#             row_data[weighted_pct_col] = f'{weighted_pct:.0f}%'
        
#         # Add Responsible and Delay columns once at the end
#         row_data['Responsible'] = ''
#         row_data['Delay Reason'] = ''
        
#         data.append(row_data)
#         milestone_counter += 1
    
#     # Add summary row with average Weighted % for each month
#     summary_row = {'Milestone': 'AVERAGE WEIGHTED %', 'Block': ''}
#     summary_row[f'Target to be complete by {MONTHS[-1]}-{MONTH_YEARS[MONTHS[-1]]}'] = ''
    
#     for month in MONTHS:
#         month_year = MONTH_YEARS[month]
        
#         # Calculate average of Weighted % for this month
#         weighted_values = []
#         for row in data:
#             weighted_val = row.get(f'Weighted %- {month}', '')
#             if weighted_val and weighted_val != '':
#                 try:
#                     val = float(str(weighted_val).replace('%', ''))
#                     weighted_values.append(val)
#                 except (ValueError, TypeError):
#                     pass
        
#         # Calculate average
#         if weighted_values:
#             avg_weighted = sum(weighted_values) / len(weighted_values)
#             summary_row[f'Weighted %- {month}'] = f'{avg_weighted:.1f}%'
#         else:
#             summary_row[f'Weighted %- {month}'] = ''
        
#         # Leave other columns blank for summary row
#         summary_row[f'Target - {month}-{month_year}'] = ''
#         summary_row[f'% work done- {month} Status'] = ''
#         summary_row[f'Achieved- {month} {month_year}'] = ''
#         summary_row[f'Weightage- {month}'] = ''
    
#     summary_row['Responsible'] = ''
#     summary_row['Delay Reason'] = ''
    
#     data.append(summary_row)
    
#     # Create DataFrame
#     df = pd.DataFrame(data)
    
#     # Reorder columns to match format
#     column_order = ['Milestone', 'Block', f'Target to be complete by {MONTHS[-1]}-{MONTH_YEARS[MONTHS[-1]]}']
    
#     for month in MONTHS:
#         month_year = MONTH_YEARS[month]
#         column_order.extend([
#             f'Target - {month}-{month_year}',
#             f'% work done- {month} Status',
#             f'Achieved- {month} {month_year}',
#             f'Weightage- {month}',
#             f'Weighted %- {month}'
#         ])
    
#     column_order.extend(['Responsible', 'Delay Reason'])
    
#     df = df[column_order]
    
#     return df

# def apply_manual_overrides(df):
#     """
#     Apply any manual overrides if needed
#     This function can be customized based on business rules
#     """
#     logger.info("Applying manual overrides (if any)...")
#     return df

# # -----------------------------------------------------------------------------
# # REPORT GENERATION
# # -----------------------------------------------------------------------------

# def write_wcc_excel_report_consolidated(df, filename):
#     """Generate Excel report with proper formatting matching the format file"""
    
#     logger.info(f'Generating consolidated report: {filename}')
    
#     wb = Workbook()
#     ws = wb.active
#     ws.title = 'WCC Progress'
    
#     # Define styles
#     header_font = Font(name='Calibri', size=11, bold=True)
#     normal_font = Font(name='Calibri', size=10)
#     date_font = Font(name='Calibri', size=10, bold=True)
#     summary_font = Font(name='Calibri', size=11, bold=True)
    
#     center = Alignment(horizontal='center', vertical='center', wrap_text=True)
#     left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
#     border = Border(
#         left=Side(style='thin'),
#         right=Side(style='thin'),
#         top=Side(style='thin'),
#         bottom=Side(style='thin')
#     )
    
#     light_grey_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
#     light_blue_fill = PatternFill(start_color='B4C7E7', end_color='B4C7E7', fill_type='solid')
#     summary_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    
#     # Title row
#     ws.merge_cells('A1:T1')
#     ws['A1'] = 'Wave City Club- Progress Against Milestones'
#     ws['A1'].font = Font(name='Calibri', size=14, bold=True)
#     ws['A1'].alignment = center
    
#     # Date row
#     current_date = datetime.now().strftime('%d-%m-%Y')
#     ws.merge_cells('A2:T2')
#     ws['A2'] = f'Report Generated on: {current_date}'
#     ws['A2'].font = date_font
#     ws['A2'].alignment = center
    
#     # Month info row
#     month_info = ', '.join([f'{m} {MONTH_YEARS[m]}' for m in MONTHS])
#     ws.merge_cells('A3:T3')
#     ws['A3'] = f'Months Covered: {month_info}'
#     ws['A3'].font = date_font
#     ws['A3'].alignment = center
    
#     # Empty row
#     ws.merge_cells('A4:T4')
    
#     # Write dataframe starting at row 5
#     for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=5):
#         for c_idx, value in enumerate(row, start=1):
#             cell = ws.cell(row=r_idx, column=c_idx, value=value)
    
#     # Style header row (row 5)
#     header_row = 5
#     for cell in ws[header_row]:
#         cell.font = header_font
#         cell.alignment = center
#         cell.border = border
#         cell.fill = light_grey_fill
    
#     # Style data rows
#     data_start = 6
#     summary_row_idx = ws.max_row
#     data_end = summary_row_idx - 1
    
#     for row_num in range(data_start, data_end + 1):
#         for col_num in range(1, len(df.columns) + 1):
#             cell = ws.cell(row=row_num, column=col_num)
#             cell.font = normal_font
#             cell.border = border
            
#             # Alignment based on column type
#             if col_num in [1, 2, 3] or 'Target' in str(ws.cell(row=header_row, column=col_num).value or ''):
#                 cell.alignment = left
#             else:
#                 cell.alignment = center
    
#     # Style summary row (last row)
#     for col_num in range(1, len(df.columns) + 1):
#         cell = ws.cell(row=summary_row_idx, column=col_num)
#         cell.font = summary_font
#         cell.border = border
#         cell.fill = summary_fill
#         cell.alignment = center
    
#     # Dynamic column width adjustment
#     for col_num in range(1, len(df.columns) + 1):
#         col_letter = get_column_letter(col_num)
        
#         max_length = 0
#         for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=col_num, max_col=col_num):
#             for cell in row:
#                 if cell.value:
#                     max_length = max(max_length, len(str(cell.value)))
        
#         calculated_width = min(max(max_length + 2, 8), 15)
#         ws.column_dimensions[col_letter].width = calculated_width
    
#     # Set row heights
#     ws.row_dimensions[1].height = 25
#     ws.row_dimensions[2].height = 20
#     ws.row_dimensions[3].height = 20
#     for i in range(5, ws.max_row + 1):
#         ws.row_dimensions[i].height = 25
#     ws.row_dimensions[summary_row_idx].height = 30  # Make summary row taller
    
#     wb.save(filename)
#     logger.info(f'Report saved to {filename}')

# def get_unique_filename(base_name):
#     """If file exists, append (1), (2), etc."""
#     if not os.path.exists(base_name):
#         return base_name

#     name, ext = os.path.splitext(base_name)
#     counter = 1
#     new_name = f"{name}({counter}){ext}"
#     while os.path.exists(new_name):
#         counter += 1
#         new_name = f"{name}({counter}){ext}"
#     return new_name

# # -----------------------------------------------------------------------------
# # MAIN FUNCTION
# # -----------------------------------------------------------------------------

# def main():
#     """Main execution function for quarterly report generation"""
#     logger.info("=== STARTING WAVE CITY CLUB QUARTERLY REPORT GENERATION (UPDATED LOGIC) ===")
    
#     try:
#         # Initialize COS client
#         cos = init_cos()
        
#         # Step 1: Get latest KRA file from Milestone/ folder
#         logger.info("\n=== STEP 1: Finding Latest KRA File ===")
#         kra_file = get_latest_kra_file(cos)
        
#         if not kra_file:
#             logger.error("❌ Failed to find KRA file")
#             return
        
#         logger.info(f"✅ Using KRA file: {kra_file}")
        
#         # Step 2: Extract months from KRA filename and setup
#         logger.info("\n=== STEP 2: Setting Up Quarterly Months ===")
#         if not setup_quarterly_months(os.path.basename(kra_file)):
#             logger.error("❌ Failed to setup quarterly months")
#             return
        
#         logger.info(f"✅ Quarterly months: {MONTHS}")
#         logger.info(f"✅ Month-Year mapping: {MONTH_YEARS}")
        
#         # Step 3: Find appropriate trackers for each month from Wave City Club/ folder
#         logger.info("\n=== STEP 3: Finding Trackers for Each Month ===")
#         global TRACKER_PATHS
        
#         for month in MONTHS:
#             month_year = MONTH_YEARS[month]
#             tracker_path = get_tracker_for_month(cos, month, month_year)
            
#             if tracker_path:
#                 TRACKER_PATHS[month] = tracker_path
#                 logger.info(f"✅ {month} {month_year}: {tracker_path}")
#             else:
#                 logger.warning(f"⚠️ {month} {month_year}: No tracker found - column will be blank in report")
#                 TRACKER_PATHS[month] = None
        
#         # Continue even if no trackers found
#         logger.info(f"Found trackers for {sum(1 for v in TRACKER_PATHS.values() if v)} out of {len(MONTHS)} months")
        
#         # Step 4: Get targets from KRA file (Club sheet)
#         logger.info("\n=== STEP 4: Extracting Targets from KRA File ===")
#         targets = get_wcc_targets_from_kra(cos)
        
#         if not targets:
#             logger.error("❌ Failed to extract targets from KRA file")
#             return
        
#         logger.info(f"✅ Extracted targets for {len(targets)} blocks")
        
#         # Step 5: Extract progress data for all months with NEW LOGIC
#         logger.info("\n=== STEP 5: Extracting Progress Data from Trackers (NEW LOGIC) ===")
#         df = get_wcc_progress_from_tracker_all_months(cos, targets)
        
#         if df.empty:
#             logger.error("❌ Failed to generate progress data")
#             return
        
#         logger.info(f"✅ Generated progress data for {len(df)-1} milestones + 1 summary row")
        
#         # Step 5.5: Apply manual overrides
#         logger.info("\n=== STEP 5.5: Applying Manual Overrides ===")
#         df = apply_manual_overrides(df)
        
#         # Step 6: Generate Excel report
#         logger.info("\n=== STEP 6: Generating Excel Report ===")
#         current_date_for_filename = datetime.now().strftime('%d-%m-%Y')
        
#         # Create filename with quarter months
#         quarter_str = "_".join(MONTHS)
#         base_filename = f"Wave_City_Club_Milestone_Report_{quarter_str}_{current_date_for_filename}.xlsx"
#         filename = get_unique_filename(base_filename)
        
#         write_wcc_excel_report_consolidated(df, filename)
        
#         logger.info("\n=== WAVE CITY CLUB QUARTERLY REPORT GENERATION COMPLETE ===")
#         logger.info(f"✅ Report saved as: {filename}")
        
#         # Log summary
#         logger.info("\n=== REPORT SUMMARY ===")
#         logger.info(f"  Quarter Months: {', '.join([f'{m} {MONTH_YEARS[m]}' for m in MONTHS])}")
#         logger.info(f"  KRA File: {os.path.basename(kra_file)}")
#         logger.info(f"  Total Blocks: {len(targets)}")
#         logger.info(f"  Trackers Used:")
#         for month in MONTHS:
#             tracker = TRACKER_PATHS.get(month)
#             if tracker:
#                 tracker_date = extract_date_from_filename(os.path.basename(tracker))
#                 logger.info(f"    - {month} {MONTH_YEARS[month]}: {os.path.basename(tracker)} (dated {tracker_date.strftime('%d-%m-%Y') if tracker_date else 'Unknown'})")
#             else:
#                 logger.info(f"    - {month} {MONTH_YEARS[month]}: Not Available (column will be blank)")
#         logger.info(f"  Total Milestones: {len(df)-1} + 1 summary row")
        
#     except Exception as e:
#         logger.error(f"❌ Error in main execution: {e}")
#         import traceback
#         logger.error(traceback.format_exc())
#         raise

# if __name__ == "__main__":
#     main()






















# import os
# import re
# import logging
# from io import BytesIO
# from datetime import datetime
# from dateutil.relativedelta import relativedelta

# import pandas as pd
# from openpyxl import Workbook, load_workbook
# from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
# from openpyxl.utils import get_column_letter
# from openpyxl.utils.dataframe import dataframe_to_rows
# from dotenv import load_dotenv
# import ibm_boto3
# from ibm_botocore.client import Config

# # -----------------------------------------------------------------------------
# # CONFIG / CONSTANTS
# # -----------------------------------------------------------------------------
# load_dotenv()
# logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(levelname)s] %(message)s')
# logger = logging.getLogger(__name__)

# # Validate required environment variables
# required = {
#     'COS_API_KEY': os.getenv('COS_API_KEY'),
#     'COS_SERVICE_INSTANCE_CRN': os.getenv('COS_SERVICE_INSTANCE_CRN'),
#     'COS_ENDPOINT': os.getenv('COS_ENDPOINT'),
#     'COS_BUCKET_NAME': os.getenv('COS_BUCKET_NAME'),
# }
# missing = [k for k, v in required.items() if not v]
# if missing:
#     logger.error(f"Missing required environment variables: {', '.join(missing)}")
#     raise SystemExit(1)

# COS_API_KEY     = required['COS_API_KEY']
# COS_CRN         = required['COS_SERVICE_INSTANCE_CRN']
# COS_ENDPOINT    = required['COS_ENDPOINT']
# BUCKET          = required['COS_BUCKET_NAME']

# # Folder paths in COS
# KRA_FOLDER = "Milestone/"
# TRACKER_FOLDER = "Wave City Club/"

# # Dynamic KRA file path
# WCC_KRA_KEY = None

# # Dynamic tracker paths
# TRACKER_PATHS = {}  # Maps month names to tracker file paths
# LOADED_TRACKERS = {}  # Cache loaded workbooks

# # Dynamic months and years
# MONTHS = []
# MONTH_YEARS = {}  # Maps month name to year
# TARGET_END_MONTH = None
# TARGET_END_YEAR = None

# # Quarterly definitions (for reference)
# QUARTERLY_GROUPS = [
#     ['June', 'July', 'August'],           # Q1
#     ['September', 'October', 'November'], # Q2
#     ['December', 'January', 'February'],  # Q3
#     ['March', 'April', 'May']             # Q4
# ]

# # Month shift logic: Report Month -> Tracker Month
# MONTH_SHIFT = {
#     'June': 'July',
#     'July': 'August',
#     'August': 'September',
#     'September': 'October',
#     'October': 'November',
#     'November': 'December',
#     'December': 'January',
#     'January': 'February',
#     'February': 'March',
#     'March': 'April',
#     'April': 'May',
#     'May': 'June'
# }

# # Block mapping from KRA to tracker sheets
# BLOCK_MAPPING = {
#     'Block 1 (B1) Banquet Hall': 'B1 Banket Hall & Finedine ',
#     'Fine Dine': 'B1 Banket Hall & Finedine ',
#     'Block 5 (B5) Admin + Member Lounge+Creche+Av Room + Surveillance Room +Toilets': 'B5',
#     'Block 6 (B6) Toilets': 'B6',
#     'Block 7(B7) Indoor Sports': 'B7',
#     'Block 9 (B9) Spa & Saloon': 'B9',
#     'Block 8 (B8) Squash Court': 'B8',
#     'Block 2 & 3 (B2 & B3) Cafe & Bar': 'B2 & B3',
#     'Block 4 (B4) Indoor Swimming Pool Changing Room & Toilets': 'B4',
#     'Block 11 (B11) Guest House': 'B11',
#     'Block 10 (B10) Gym': 'B10'
# }

# # KRA Column structure (0-indexed)
# KRA_COLUMNS = {
#     'BLOCK': 0,          # Column A: Block name
#     'MONTH1_ACTIVITY': 1,  # Column B: First month activity
#     'MONTH1_TARGET': 2,    # Column C: First month target %
#     'MONTH2_ACTIVITY': 3,  # Column D: Second month activity
#     'MONTH2_TARGET': 4,    # Column E: Second month target %
#     'MONTH3_ACTIVITY': 5,  # Column F: Third month activity
#     'MONTH3_TARGET': 6,    # Column G: Third month target %
# }

# # Tracker column positions (0-indexed)
# TRACKER_COLUMNS = {
#     'ACTIVITY_NAME': 6,    # Column G: Activity Name
#     'PERCENT_COMPLETE': 12  # Column M: % Complete
# }

# # -----------------------------------------------------------------------------
# # COS HELPERS
# # -----------------------------------------------------------------------------

# def init_cos():
#     return ibm_boto3.client(
#         's3',
#         ibm_api_key_id=COS_API_KEY,
#         ibm_service_instance_id=COS_CRN,
#         config=Config(signature_version='oauth'),
#         endpoint_url=COS_ENDPOINT,
#     )

# def download_file_bytes(cos, key):
#     if not key:
#         raise ValueError("File key cannot be None or empty")
#     obj = cos.get_object(Bucket=BUCKET, Key=key)
#     return obj['Body'].read()

# def list_files_in_folder(cos, folder_prefix):
#     """List all files in a specific folder (prefix) in the COS bucket"""
#     try:
#         response = cos.list_objects_v2(Bucket=BUCKET, Prefix=folder_prefix)
#         files = []
#         if 'Contents' in response:
#             for obj in response['Contents']:
#                 if not obj['Key'].endswith('/'):
#                     files.append(obj['Key'])
#         return files
#     except Exception as e:
#         logger.error(f"Error listing files in folder {folder_prefix}: {e}")
#         return []

# def extract_date_from_filename(filename):
#     """Extract date from filename in format (dd-mm-yyyy)"""
#     pattern = r'\((\d{2}-\d{2}-\d{4})\)'
#     match = re.search(pattern, filename)
#     if match:
#         date_str = match.group(1)
#         try:
#             return datetime.strptime(date_str, '%d-%m-%Y')
#         except ValueError:
#             logger.warning(f"Could not parse date {date_str} from filename {filename}")
#             return None
#     return None

# def get_month_name(month_num):
#     """Convert month number to month name"""
#     months = {
#         1: "January", 2: "February", 3: "March", 4: "April",
#         5: "May", 6: "June", 7: "July", 8: "August", 
#         9: "September", 10: "October", 11: "November", 12: "December"
#     }
#     return months.get(month_num, "Unknown")

# def get_month_number(month_name):
#     """Convert month name to month number"""
#     months = {
#         "January": 1, "February": 2, "March": 3, "April": 4,
#         "May": 5, "June": 6, "July": 7, "August": 8, 
#         "September": 9, "October": 10, "November": 11, "December": 12
#     }
#     return months.get(month_name, 1)

# def get_latest_kra_file(cos):
#     """Get the latest KRA milestone file from the Milestone/ folder"""
#     global WCC_KRA_KEY
    
#     logger.info("=== FINDING LATEST KRA MILESTONE FILE ===")
    
#     # List all files in Milestone folder
#     all_files = list_files_in_folder(cos, KRA_FOLDER)
#     logger.info(f"Found {len(all_files)} files in {KRA_FOLDER} folder")
    
#     # Pattern to match KRA milestone files
#     kra_pattern = r'KRA Milestones for.*\.xlsx$'
    
#     matching_files = []
    
#     for file_path in all_files:
#         filename = os.path.basename(file_path)
#         if re.search(kra_pattern, filename, re.IGNORECASE):
#             logger.info(f"Found KRA file: {filename}")
            
#             # Extract date from filename
#             file_date = extract_date_from_filename(filename)
#             if file_date:
#                 matching_files.append((file_path, file_date))
#                 logger.info(f"  Extracted date: {file_date.strftime('%d-%m-%Y')}")
#             else:
#                 logger.warning(f"  Could not extract date from {filename}")
#                 matching_files.append((file_path, datetime.min))  # Fallback for files without dates
    
#     if matching_files:
#         # Sort by date (newest first)
#         matching_files.sort(key=lambda x: x[1], reverse=True)
#         latest_file = matching_files[0][0]
#         WCC_KRA_KEY = latest_file
#         logger.info(f"✅ Latest KRA file: {os.path.basename(WCC_KRA_KEY)} (dated {matching_files[0][1].strftime('%d-%m-%Y')})")
#     else:
#         logger.error(f"❌ No KRA milestone files found in {KRA_FOLDER} folder!")
#         WCC_KRA_KEY = None
    
#     return WCC_KRA_KEY

# def extract_months_from_kra_filename(filename):
#     """
#     Extract quarter months from KRA filename
#     Format: "KRA Milestones for [Month] [Month] [Month] [Year].xlsx"
#     Example: "KRA Milestones for December January February 2026.xlsx"
#     The year in the filename is the END year of the quarter
#     """
#     months_pattern = r'KRA\s+Milestones\s+for\s+((?:January|February|March|April|May|June|July|August|September|October|November|December)(?:\s+(?:January|February|March|April|May|June|July|August|September|October|November|December))*)\s+(\d{4})'
    
#     match = re.search(months_pattern, filename, re.IGNORECASE)
#     if match:
#         months_str = match.group(1)
#         end_year = int(match.group(2))
        
#         # Extract individual months in order
#         month_names = re.findall(r'January|February|March|April|May|June|July|August|September|October|November|December', 
#                                  months_str, re.IGNORECASE)
        
#         logger.info(f"Extracted from KRA filename '{filename}':")
#         logger.info(f"  Months: {month_names}, End Year: {end_year}")
        
#         return [m.capitalize() for m in month_names], end_year
    
#     logger.warning(f"Could not parse KRA filename: {filename}")
#     return None, None

# def get_tracker_for_month(cos, report_month, month_year):
#     """
#     Get tracker file for a specific report month using the month-shift logic.
#     Uses the LATEST tracker available that matches the required month/year.
    
#     Args:
#         cos: COS client
#         report_month: The month shown in the report (e.g., "June")
#         month_year: Year of the report month
    
#     Returns:
#         Path to tracker file or None
#     """
#     # Get the tracker month based on shift logic
#     tracker_month_name = MONTH_SHIFT.get(report_month)
    
#     if not tracker_month_name:
#         logger.warning(f"No month shift mapping found for {report_month}")
#         return None
    
#     # Calculate the tracker year (handle year rollover)
#     tracker_month_num = get_month_number(tracker_month_name)
#     report_month_num = get_month_number(report_month)
    
#     tracker_year = month_year
#     # If tracker month is earlier in the year than report month, it's next year
#     if tracker_month_num < report_month_num:
#         tracker_year += 1
    
#     logger.info(f"Looking for tracker for Report: {report_month} {month_year} → Tracker: {tracker_month_name} {tracker_year}")
    
#     # List all files in Wave City Club folder
#     all_files = list_files_in_folder(cos, TRACKER_FOLDER)
#     logger.info(f"Total files in {TRACKER_FOLDER}: {len(all_files)}")
    
#     # Pattern to match Structure Work tracker files
#     tracker_pattern = r'Structure.*Work.*Tracker.*Wave.*City.*Club.*\.xlsx$'
    
#     matching_files = []
    
#     for file_path in all_files:
#         filename = os.path.basename(file_path)
#         if re.search(tracker_pattern, filename, re.IGNORECASE):
#             # Extract date from filename
#             file_date = extract_date_from_filename(filename)
            
#             logger.info(f"  Found tracker: {filename}")
            
#             if file_date:
#                 logger.info(f"    Extracted date: {file_date.strftime('%d-%m-%Y')} (Month: {file_date.month}, Year: {file_date.year})")
#                 logger.info(f"    Looking for: Month={tracker_month_num}, Year={tracker_year}")
                
#                 # Check if this file matches the month and year we're looking for
#                 if file_date.month == tracker_month_num and file_date.year == tracker_year:
#                     matching_files.append((file_path, file_date))
#                     logger.info(f"    ✅ MATCH!")
#                 else:
#                     logger.info(f"    ❌ Date mismatch (need {tracker_month_name} {tracker_year})")
#             else:
#                 logger.warning(f"    Could not extract date from {filename}")
    
#     if matching_files:
#         # Sort by date and get the LATEST matching tracker
#         matching_files.sort(key=lambda x: x[1], reverse=True)
#         latest_tracker = matching_files[0][0]
#         logger.info(f"✅ Selected tracker: {os.path.basename(latest_tracker)} (dated {matching_files[0][1].strftime('%d-%m-%Y')})")
#         return latest_tracker
#     else:
#         logger.warning(f"⚠️ No tracker found for {report_month} {month_year} (looking for {tracker_month_name} {tracker_year} dated files)")
#         return None

# def setup_quarterly_months(kra_filename):
#     """
#     Setup the quarterly months and years based on KRA filename.
    
#     The year in the KRA filename represents the END year of the quarter.
#     For cross-year quarters (e.g., Dec-Jan-Feb 2026):
#     - December belongs to 2025
#     - January belongs to 2026
#     - February belongs to 2026
#     """
#     global MONTHS, MONTH_YEARS, TARGET_END_MONTH, TARGET_END_YEAR
    
#     months, end_year = extract_months_from_kra_filename(kra_filename)
    
#     if not months or not end_year:
#         logger.error(f"Could not extract months from KRA filename: {kra_filename}")
#         return False
    
#     MONTHS = months
#     TARGET_END_MONTH = MONTHS[-1]
#     TARGET_END_YEAR = end_year
    
#     # Determine starting year for the quarter
#     first_month_num = get_month_number(MONTHS[0])
#     last_month_num = get_month_number(MONTHS[-1])
    
#     # For cross-year quarters (first month number > last month number),
#     # the first month belongs to the previous year
#     if first_month_num > last_month_num:
#         # Cross-year quarter (e.g., Dec 2025, Jan 2026, Feb 2026)
#         start_year = end_year - 1
#         logger.info(f"Cross-year quarter detected: {MONTHS[0]} {start_year} - {MONTHS[-1]} {end_year}")
#     else:
#         # Same-year quarter (e.g., Jun 2025, Jul 2025, Aug 2025)
#         start_year = end_year
#         logger.info(f"Same-year quarter: all months in {start_year}")
    
#     # Assign years to each month
#     current_year = start_year
#     for i, month in enumerate(MONTHS):
#         month_num = get_month_number(month)
        
#         # Year increments when month number decreases (year rollover)
#         if i > 0:
#             prev_month_num = get_month_number(MONTHS[i-1])
#             if month_num < prev_month_num:
#                 current_year += 1
        
#         MONTH_YEARS[month] = current_year
    
#     logger.info(f"Month-Year mapping:")
#     for month in MONTHS:
#         logger.info(f"  {month}: {MONTH_YEARS[month]}")
    
#     return True


# def find_club_sheet(wb):
#     """Find the sheet with 'Club' in its name"""
#     for sheet_name in wb.sheetnames:
#         if 'club' in sheet_name.lower():
#             logger.info(f"Found Club sheet: {sheet_name}")
#             return sheet_name
    
#     logger.warning("No sheet with 'Club' in name found. Available sheets: " + ", ".join(wb.sheetnames))
#     return None

# # -----------------------------------------------------------------------------
# # DATA EXTRACTION
# # -----------------------------------------------------------------------------

# def get_wcc_targets_from_kra(cos):
#     """
#     Extract targets from KRA milestone file with updated logic:
#     - Look for sheet with "Club" in its name
#     - For each activity, fetch both the target activity name AND target %
#     - Store both values for comparison with tracker data
#     """
#     global WCC_KRA_KEY
    
#     if not WCC_KRA_KEY:
#         raise ValueError("WCC_KRA_KEY is not set. Call get_latest_kra_file() first.")
    
#     logger.info(f"Reading KRA file: {WCC_KRA_KEY}")
    
#     # Download and load KRA file
#     kra_bytes = download_file_bytes(cos, WCC_KRA_KEY)
#     kra_wb = load_workbook(BytesIO(kra_bytes), data_only=True)
    
#     # Find the Club sheet
#     club_sheet_name = find_club_sheet(kra_wb)
#     if not club_sheet_name:
#         # Fallback to active sheet
#         logger.warning("Using active sheet as fallback")
#         ws = kra_wb.active
#     else:
#         ws = kra_wb[club_sheet_name]
    
#     targets = {}
    
#     # Find header row (look for "Blocks" in column A)
#     header_row = None
#     for row_idx in range(1, min(10, ws.max_row + 1)):
#         cell_value = ws.cell(row=row_idx, column=1).value
#         if cell_value and 'block' in str(cell_value).lower():
#             header_row = row_idx
#             logger.info(f"Found header row at row {header_row}")
#             break
    
#     if not header_row:
#         logger.error("Could not find header row with 'Blocks'")
#         kra_wb.close()
#         return targets
    
#     # Start from row after header
#     data_start_row = header_row + 1
    
#     for row_idx in range(data_start_row, ws.max_row + 1):
#         row = list(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
        
#         # Extract block name from column A
#         block_name = row[KRA_COLUMNS['BLOCK']]
        
#         # Skip empty rows
#         if not block_name or pd.isna(block_name):
#             continue
        
#         # Stop if we hit another header row or non-block data
#         block_name_str = str(block_name).strip()
#         if any(month.lower() in block_name_str.lower() for month in ['january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december']):
#             logger.info(f"Stopping at row {row_idx}: Found header-like content '{block_name_str}'")
#             break
        
#         # Skip if block name is just "Activity" or other single-word non-block names
#         if block_name_str.lower() in ['activity', 'activities', 'milestone', 'milestones', 'target', 'targets', 'blocks']:
#             logger.info(f"Skipping non-block row: '{block_name_str}'")
#             continue
            
#         logger.info(f"Processing block: {block_name}")
        
#         targets[block_name] = {}
        
#         # Process each month
#         for i, month in enumerate(MONTHS):
#             month_year = MONTH_YEARS[month]
            
#             # Calculate column indices for this month
#             if i == 0:
#                 activity_col = KRA_COLUMNS['MONTH1_ACTIVITY']
#                 target_col = KRA_COLUMNS['MONTH1_TARGET']
#             elif i == 1:
#                 activity_col = KRA_COLUMNS['MONTH2_ACTIVITY']
#                 target_col = KRA_COLUMNS['MONTH2_TARGET']
#             elif i == 2:
#                 activity_col = KRA_COLUMNS['MONTH3_ACTIVITY']
#                 target_col = KRA_COLUMNS['MONTH3_TARGET']
#             else:
#                 logger.warning(f"More than 3 months found, skipping month {month}")
#                 continue
            
#             # Extract activity name and target %
#             activity_name = row[activity_col] if activity_col < len(row) else None
#             target_percent = row[target_col] if target_col < len(row) else None
            
#             # Validate target_percent is a number, not text
#             validated_target_percent = None
#             if target_percent and not pd.isna(target_percent):
#                 try:
#                     validated_target_percent = float(target_percent)
#                 except (ValueError, TypeError):
#                     logger.warning(f"Invalid target % value for {block_name} - {month}: '{target_percent}' (not a number)")
#                     validated_target_percent = None
            
#             # Store both activity and target % for this month
#             targets[block_name][month] = {
#                 'activity': activity_name if activity_name and not pd.isna(activity_name) else None,
#                 'target_percent': validated_target_percent
#             }
            
#             logger.info(f"  {month}: Activity='{activity_name}', Target %={validated_target_percent}")
    
#     kra_wb.close()
#     return targets

# def find_activity_completion_in_tracker(tracker_wb, sheet_name, target_activity):
#     """
#     Search for an activity in the tracker and return its % complete value.
    
#     Args:
#         tracker_wb: Loaded tracker workbook
#         sheet_name: Sheet name to search in
#         target_activity: Activity name to find
    
#     Returns:
#         Float value of % complete, or None if not found
#     """
#     if sheet_name not in tracker_wb.sheetnames:
#         logger.warning(f"Sheet '{sheet_name}' not found in tracker")
#         return None
    
#     ws = tracker_wb[sheet_name]
    
#     # Search through rows starting from row 3 (skip headers)
#     for row_idx in range(2, ws.max_row + 1):
#         row = list(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
        
#         # Get activity name from column G (index 6)
#         if len(row) > TRACKER_COLUMNS['ACTIVITY_NAME']:
#             activity_name = row[TRACKER_COLUMNS['ACTIVITY_NAME']]
            
#             if activity_name and str(activity_name).strip().lower() == str(target_activity).strip().lower():
#                 # Found the activity, get % complete from column M (index 12)
#                 if len(row) > TRACKER_COLUMNS['PERCENT_COMPLETE']:
#                     percent_complete = row[TRACKER_COLUMNS['PERCENT_COMPLETE']]
#                     try:
#                         return float(percent_complete) if percent_complete is not None else 0.0
#                     except (ValueError, TypeError):
#                         logger.warning(f"Could not convert % complete to float for activity '{activity_name}': {percent_complete}")
#                         return 0.0
    
#     logger.debug(f"Activity '{target_activity}' not found in sheet '{sheet_name}'")
#     return None

# def get_wcc_progress_from_tracker_all_months(cos, targets):
#     """
#     Extract progress data for all months with UPDATED LOGIC:
#     - Weightage = 100 for all blocks
#     - Weighted % = (% Complete / Weightage) * 100 for each month
#     - If tracker % > target %, display 100%
#     - Responsible and Delay columns shown once at the end
#     """
#     global TRACKER_PATHS, LOADED_TRACKERS
    
#     # Load all trackers into memory
#     for month in MONTHS:
#         tracker_path = TRACKER_PATHS.get(month)
#         if tracker_path and month not in LOADED_TRACKERS:
#             logger.info(f"Loading tracker for {month}: {tracker_path}")
#             tracker_bytes = download_file_bytes(cos, tracker_path)
#             LOADED_TRACKERS[month] = load_workbook(BytesIO(tracker_bytes), data_only=True)
    
#     # Build dataframe
#     data = []
#     milestone_counter = 1
    
#     for block_name, month_targets in targets.items():
#         # Get the tracker sheet name for this block
#         tracker_sheet = BLOCK_MAPPING.get(block_name)
        
#         if not tracker_sheet:
#             logger.warning(f"No tracker sheet mapping found for block: {block_name}")
#             continue
        
#         row_data = {
#             'Milestone': f'Milestone-{milestone_counter:02d}',
#             'Block': block_name
#         }
        
#         # Determine the target to complete by the end of quarter
#         last_month = MONTHS[-1]
#         last_month_target_info = month_targets.get(last_month, {})
#         target_to_complete = last_month_target_info.get('activity', 'N/A')
#         row_data[f'Target to be complete by {last_month}-{MONTH_YEARS[last_month]}'] = target_to_complete
        
#         # Process each month
#         for month in MONTHS:
#             month_year = MONTH_YEARS[month]
            
#             # Get target info for this month
#             target_info = month_targets.get(month, {})
#             target_activity = target_info.get('activity')
#             target_percent = target_info.get('target_percent')
            
#             # Column names for this month
#             target_col = f'Target - {month}-{month_year}'
#             status_col = f'% work done- {month} Status'
#             achieved_col = f'Achieved- {month} {month_year}'
#             weightage_col = f'Weightage- {month}'
#             weighted_pct_col = f'Weighted %- {month}'
            
#             row_data[target_col] = target_activity if target_activity else 'No target'
            
#             # Check if tracker is available for this month
#             if month not in LOADED_TRACKERS:
#                 # Leave blank if tracker not available
#                 row_data[status_col] = ''
#                 row_data[achieved_col] = ''
#                 row_data[weightage_col] = ''
#                 row_data[weighted_pct_col] = ''
#                 continue
            
#             # If no target activity, mark as N/A
#             if not target_activity:
#                 row_data[status_col] = 'N/A'
#                 row_data[achieved_col] = 'No target for this month'
#                 row_data[weightage_col] = ''
#                 row_data[weighted_pct_col] = ''
#                 continue
            
#             # Fetch completed % from tracker
#             tracker_wb = LOADED_TRACKERS[month]
#             completed_percent = find_activity_completion_in_tracker(tracker_wb, tracker_sheet, target_activity)
            
#             # Determine status based on comparison
#             if completed_percent is None:
#                 # Activity not found in tracker
#                 row_data[status_col] = '0%'
#                 row_data[achieved_col] = 'Activity not found in tracker'
#                 status_percent = 0.0
#             elif target_percent is not None and completed_percent > target_percent:
#                 # Tracker % > Target % -> display 100%
#                 row_data[status_col] = '100%'
#                 row_data[achieved_col] = f'Target exceeded ({completed_percent*100:.0f}% > {target_percent*100:.0f}%)'
#                 status_percent = 1.0
#             elif target_percent is not None and completed_percent == target_percent:
#                 # Target % matches completed %
#                 row_data[status_col] = '100%'
#                 row_data[achieved_col] = f'Target achieved ({completed_percent*100:.0f}% complete)'
#                 status_percent = 1.0
#             else:
#                 # Display completed % from tracker
#                 status_percent = completed_percent
#                 row_data[status_col] = f'{completed_percent*100:.0f}%'
                
#                 if target_percent is not None and completed_percent >= target_percent:
#                     row_data[achieved_col] = f'{completed_percent*100:.0f}% completed (Target: {target_percent*100:.0f}%)'
#                 elif target_percent is not None:
#                     row_data[achieved_col] = f'{completed_percent*100:.0f}% completed (Target: {target_percent*100:.0f}%)'
#                 else:
#                     row_data[achieved_col] = f'{completed_percent*100:.0f}% completed'
            
#             # Calculate Weightage and Weighted % for this month
#             weightage = 100  # Always 100
#             weighted_pct = (status_percent * 100 / weightage) * 100 if weightage > 0 else 0
#             row_data[weightage_col] = weightage
#             row_data[weighted_pct_col] = f'{weighted_pct:.0f}%'
        
#         # Add Responsible and Delay columns once at the end
#         row_data['Responsible'] = ''
#         row_data['Delay Reason'] = ''
        
#         data.append(row_data)
#         milestone_counter += 1
    
#     # Add summary row with average Weighted % for each month
#     summary_row = {'Milestone': 'AVERAGE WEIGHTED %', 'Block': ''}
#     summary_row[f'Target to be complete by {MONTHS[-1]}-{MONTH_YEARS[MONTHS[-1]]}'] = ''
    
#     for month in MONTHS:
#         month_year = MONTH_YEARS[month]
        
#         # Calculate average of Weighted % for this month
#         weighted_values = []
#         for row in data:
#             weighted_val = row.get(f'Weighted %- {month}', '')
#             if weighted_val and weighted_val != '':
#                 try:
#                     val = float(str(weighted_val).replace('%', ''))
#                     weighted_values.append(val)
#                 except (ValueError, TypeError):
#                     pass
        
#         # Calculate average
#         if weighted_values:
#             avg_weighted = sum(weighted_values) / len(weighted_values)
#             summary_row[f'Weighted %- {month}'] = f'{avg_weighted:.1f}%'
#         else:
#             summary_row[f'Weighted %- {month}'] = ''
        
#         # Leave other columns blank for summary row
#         summary_row[f'Target - {month}-{month_year}'] = ''
#         summary_row[f'% work done- {month} Status'] = ''
#         summary_row[f'Achieved- {month} {month_year}'] = ''
#         summary_row[f'Weightage- {month}'] = ''
    
#     summary_row['Responsible'] = ''
#     summary_row['Delay Reason'] = ''
    
#     data.append(summary_row)
    
#     # Create DataFrame
#     df = pd.DataFrame(data)
    
#     # Reorder columns to match format
#     column_order = ['Milestone', 'Block', f'Target to be complete by {MONTHS[-1]}-{MONTH_YEARS[MONTHS[-1]]}']
    
#     for month in MONTHS:
#         month_year = MONTH_YEARS[month]
#         column_order.extend([
#             f'Target - {month}-{month_year}',
#             f'% work done- {month} Status',
#             f'Achieved- {month} {month_year}',
#             f'Weightage- {month}',
#             f'Weighted %- {month}'
#         ])
    
#     column_order.extend(['Responsible', 'Delay Reason'])
    
#     df = df[column_order]
    
#     return df

# def apply_manual_overrides(df):
#     """
#     Apply any manual overrides if needed
#     This function can be customized based on business rules
#     """
#     logger.info("Applying manual overrides (if any)...")
#     return df

# # -----------------------------------------------------------------------------
# # REPORT GENERATION
# # -----------------------------------------------------------------------------

# def write_wcc_excel_report_consolidated(df, filename):
#     """Generate Excel report with proper formatting matching the format file"""
    
#     logger.info(f'Generating consolidated report: {filename}')
    
#     wb = Workbook()
#     ws = wb.active
#     ws.title = 'WCC Progress'
    
#     # Define styles
#     header_font = Font(name='Calibri', size=11, bold=True)
#     normal_font = Font(name='Calibri', size=10)
#     date_font = Font(name='Calibri', size=10, bold=True)
#     summary_font = Font(name='Calibri', size=11, bold=True)
    
#     center = Alignment(horizontal='center', vertical='center', wrap_text=True)
#     left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
#     border = Border(
#         left=Side(style='thin'),
#         right=Side(style='thin'),
#         top=Side(style='thin'),
#         bottom=Side(style='thin')
#     )
    
#     light_grey_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
#     light_blue_fill = PatternFill(start_color='B4C7E7', end_color='B4C7E7', fill_type='solid')
#     summary_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    
#     # Title row
#     ws.merge_cells('A1:T1')
#     ws['A1'] = 'Wave City Club- Progress Against Milestones'
#     ws['A1'].font = Font(name='Calibri', size=14, bold=True)
#     ws['A1'].alignment = center
    
#     # Date row
#     current_date = datetime.now().strftime('%d-%m-%Y')
#     ws.merge_cells('A2:T2')
#     ws['A2'] = f'Report Generated on: {current_date}'
#     ws['A2'].font = date_font
#     ws['A2'].alignment = center
    
#     # Month info row
#     month_info = ', '.join([f'{m} {MONTH_YEARS[m]}' for m in MONTHS])
#     ws.merge_cells('A3:T3')
#     ws['A3'] = f'Months Covered: {month_info}'
#     ws['A3'].font = date_font
#     ws['A3'].alignment = center
    
#     # Empty row
#     ws.merge_cells('A4:T4')
    
#     # Write dataframe starting at row 5
#     for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=5):
#         for c_idx, value in enumerate(row, start=1):
#             cell = ws.cell(row=r_idx, column=c_idx, value=value)
    
#     # Style header row (row 5)
#     header_row = 5
#     for cell in ws[header_row]:
#         cell.font = header_font
#         cell.alignment = center
#         cell.border = border
#         cell.fill = light_grey_fill
    
#     # Style data rows
#     data_start = 6
#     summary_row_idx = ws.max_row
#     data_end = summary_row_idx - 1
    
#     for row_num in range(data_start, data_end + 1):
#         for col_num in range(1, len(df.columns) + 1):
#             cell = ws.cell(row=row_num, column=col_num)
#             cell.font = normal_font
#             cell.border = border
            
#             # Alignment based on column type
#             if col_num in [1, 2, 3] or 'Target' in str(ws.cell(row=header_row, column=col_num).value or ''):
#                 cell.alignment = left
#             else:
#                 cell.alignment = center
    
#     # Style summary row (last row)
#     for col_num in range(1, len(df.columns) + 1):
#         cell = ws.cell(row=summary_row_idx, column=col_num)
#         cell.font = summary_font
#         cell.border = border
#         cell.fill = summary_fill
#         cell.alignment = center
    
#     # Dynamic column width adjustment
#     for col_num in range(1, len(df.columns) + 1):
#         col_letter = get_column_letter(col_num)
        
#         max_length = 0
#         for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=col_num, max_col=col_num):
#             for cell in row:
#                 if cell.value:
#                     max_length = max(max_length, len(str(cell.value)))
        
#         calculated_width = min(max(max_length + 2, 8), 15)
#         ws.column_dimensions[col_letter].width = calculated_width
    
#     # Set row heights
#     ws.row_dimensions[1].height = 25
#     ws.row_dimensions[2].height = 20
#     ws.row_dimensions[3].height = 20
#     for i in range(5, ws.max_row + 1):
#         ws.row_dimensions[i].height = 25
#     ws.row_dimensions[summary_row_idx].height = 30  # Make summary row taller
    
#     wb.save(filename)
#     logger.info(f'Report saved to {filename}')

# def get_unique_filename(base_name):
#     """If file exists, append (1), (2), etc."""
#     if not os.path.exists(base_name):
#         return base_name

#     name, ext = os.path.splitext(base_name)
#     counter = 1
#     new_name = f"{name}({counter}){ext}"
#     while os.path.exists(new_name):
#         counter += 1
#         new_name = f"{name}({counter}){ext}"
#     return new_name

# # -----------------------------------------------------------------------------
# # MAIN FUNCTION
# # -----------------------------------------------------------------------------

# def main():
#     """Main execution function for quarterly report generation"""
#     logger.info("=== STARTING WAVE CITY CLUB QUARTERLY REPORT GENERATION (UPDATED LOGIC) ===")
    
#     try:
#         # Initialize COS client
#         cos = init_cos()
        
#         # Step 1: Get latest KRA file from Milestone/ folder
#         logger.info("\n=== STEP 1: Finding Latest KRA File ===")
#         kra_file = get_latest_kra_file(cos)
        
#         if not kra_file:
#             logger.error("❌ Failed to find KRA file")
#             return
        
#         logger.info(f"✅ Using KRA file: {kra_file}")
        
#         # Step 2: Extract months from KRA filename and setup
#         logger.info("\n=== STEP 2: Setting Up Quarterly Months ===")
#         if not setup_quarterly_months(os.path.basename(kra_file)):
#             logger.error("❌ Failed to setup quarterly months")
#             return
        
#         logger.info(f"✅ Quarterly months: {MONTHS}")
#         logger.info(f"✅ Month-Year mapping: {MONTH_YEARS}")
        
#         # Step 3: Find appropriate trackers for each month from Wave City Club/ folder
#         logger.info("\n=== STEP 3: Finding Trackers for Each Month ===")
#         global TRACKER_PATHS
        
#         for month in MONTHS:
#             month_year = MONTH_YEARS[month]
#             tracker_path = get_tracker_for_month(cos, month, month_year)
            
#             if tracker_path:
#                 TRACKER_PATHS[month] = tracker_path
#                 logger.info(f"✅ {month} {month_year}: {tracker_path}")
#             else:
#                 logger.warning(f"⚠️ {month} {month_year}: No tracker found - column will be blank in report")
#                 TRACKER_PATHS[month] = None
        
#         # Continue even if no trackers found
#         logger.info(f"Found trackers for {sum(1 for v in TRACKER_PATHS.values() if v)} out of {len(MONTHS)} months")
        
#         # Step 4: Get targets from KRA file (Club sheet)
#         logger.info("\n=== STEP 4: Extracting Targets from KRA File ===")
#         targets = get_wcc_targets_from_kra(cos)
        
#         if not targets:
#             logger.error("❌ Failed to extract targets from KRA file")
#             return
        
#         logger.info(f"✅ Extracted targets for {len(targets)} blocks")
        
#         # Step 5: Extract progress data for all months with NEW LOGIC
#         logger.info("\n=== STEP 5: Extracting Progress Data from Trackers (NEW LOGIC) ===")
#         df = get_wcc_progress_from_tracker_all_months(cos, targets)
        
#         if df.empty:
#             logger.error("❌ Failed to generate progress data")
#             return
        
#         logger.info(f"✅ Generated progress data for {len(df)-1} milestones + 1 summary row")
        
#         # Step 5.5: Apply manual overrides
#         logger.info("\n=== STEP 5.5: Applying Manual Overrides ===")
#         df = apply_manual_overrides(df)
        
#         # Step 6: Generate Excel report
#         logger.info("\n=== STEP 6: Generating Excel Report ===")
#         current_date_for_filename = datetime.now().strftime('%d-%m-%Y')
        
#         # Create filename with quarter months
#         quarter_str = "_".join(MONTHS)
#         base_filename = f"Wave_City_Club_Milestone_Report_{quarter_str}_{current_date_for_filename}.xlsx"
#         filename = get_unique_filename(base_filename)
        
#         write_wcc_excel_report_consolidated(df, filename)
        
#         logger.info("\n=== WAVE CITY CLUB QUARTERLY REPORT GENERATION COMPLETE ===")
#         logger.info(f"✅ Report saved as: {filename}")
        
#         # Log summary
#         logger.info("\n=== REPORT SUMMARY ===")
#         logger.info(f"  Quarter Months: {', '.join([f'{m} {MONTH_YEARS[m]}' for m in MONTHS])}")
#         logger.info(f"  KRA File: {os.path.basename(kra_file)}")
#         logger.info(f"  Total Blocks: {len(targets)}")
#         logger.info(f"  Trackers Used:")
#         for month in MONTHS:
#             tracker = TRACKER_PATHS.get(month)
#             if tracker:
#                 tracker_date = extract_date_from_filename(os.path.basename(tracker))
#                 logger.info(f"    - {month} {MONTH_YEARS[month]}: {os.path.basename(tracker)} (dated {tracker_date.strftime('%d-%m-%Y') if tracker_date else 'Unknown'})")
#             else:
#                 logger.info(f"    - {month} {MONTH_YEARS[month]}: Not Available (column will be blank)")
#         logger.info(f"  Total Milestones: {len(df)-1} + 1 summary row")
        
#     except Exception as e:
#         logger.error(f"❌ Error in main execution: {e}")
#         import traceback
#         logger.error(traceback.format_exc())
#         raise

# if __name__ == "__main__":
#     main()



