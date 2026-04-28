import os
import re
import logging
from io import BytesIO
from datetime import datetime

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from dotenv import load_dotenv
import ibm_boto3
from ibm_botocore.client import Config

# ============================================================================
# SETUP & CONFIGURATION
# ============================================================================
load_dotenv()
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger(__name__)

COS_API_KEY = os.getenv("COS_API_KEY")
COS_CRN = os.getenv("COS_SERVICE_INSTANCE_CRN")
COS_ENDPOINT = os.getenv("COS_ENDPOINT")
BUCKET = os.getenv("COS_BUCKET_NAME")

GREEN_HEX = "FF92D050"

QUARTERS = {
    'Q1': ['June', 'July', 'August'],
    'Q2': ['September', 'October', 'November'],
    'Q3': ['December', 'January', 'February'],
    'Q4': ['March', 'April', 'May']
}

MONTH_SHIFT_MAP = {
    'June': 'July', 'July': 'August', 'August': 'September',
    'September': 'October', 'October': 'November', 'November': 'December',
    'December': 'January', 'January': 'February', 'February': 'March',
    'March': 'April', 'April': 'May', 'May': 'June'
}

MONTH_TO_NUM = {
    'January': 1, 'February': 2, 'March': 3, 'April': 4,
    'May': 5, 'June': 6, 'July': 7, 'August': 8,
    'September': 9, 'October': 10, 'November': 11, 'December': 12
}

# ============================================================================
# CLOUD STORAGE
# ============================================================================
def init_cos():
    """Initialize IBM COS client"""
    return ibm_boto3.client(
        "s3",
        ibm_api_key_id=COS_API_KEY,
        ibm_service_instance_id=COS_CRN,
        config=Config(signature_version="oauth"),
        endpoint_url=COS_ENDPOINT,
    )

def list_files(cos, prefix):
    """List files from COS bucket"""
    try:
        response = cos.list_objects_v2(Bucket=BUCKET, Prefix=prefix)
        return [obj['Key'] for obj in response.get('Contents', []) if not obj['Key'].endswith('/')]
    except Exception as e:
        logger.error(f"Error listing files: {e}")
        return []

def download_file(cos, key):
    """Download file from COS bucket"""
    obj = cos.get_object(Bucket=BUCKET, Key=key)
    return obj["Body"].read()

# ============================================================================
# UTILITY FUNCTIONS
# ============================================================================
def extract_date_from_filename(filename):
    """Extract date from filename (DD-MM-YYYY)"""
    pattern = r'\((\d{2}-\d{2}-\d{4})\)'
    match = re.search(pattern, filename)
    if match:
        try:
            return datetime.strptime(match.group(1), '%d-%m-%Y')
        except ValueError:
            return None
    return None

def determine_quarter_from_kra(kra_filename):
    """Determine quarter from KRA filename"""
    pattern = r'KRA Milestones for (\w+) (\w+) (\w+) (\d{4})'
    match = re.search(pattern, kra_filename, re.IGNORECASE)
    
    if match:
        m1, m2, m3, year = match.groups()
        months = [m1, m2, m3]
        
        for q, qmonths in QUARTERS.items():
            if set([m.lower() for m in months]) == set([m.lower() for m in qmonths]):
                return q, qmonths, int(year)
    
    return None, None, None

def extract_number(cell_value):
    """Extract numeric value from cell"""
    if not cell_value or cell_value == "-":
        return 0.0
    match = re.search(r"(\d+\.?\d*)", str(cell_value))
    return float(match.group(1)) if match else 0.0

def get_cell_hex_color(cell):
    """Extract hex color from cell"""
    if cell.fill and cell.fill.start_color:
        color_value = str(cell.fill.start_color.rgb)
        if color_value.startswith('FF') and len(color_value) > 8:
            return color_value[2:]
        return color_value
    return None

def is_cell_green(cell, green_hex=GREEN_HEX):
    """Check if cell has green background"""
    cell_color = get_cell_hex_color(cell)
    if not cell_color:
        return False
    
    cell_color_normalized = cell_color.upper()
    target_normalized = green_hex.upper()
    
    if target_normalized.startswith('FF') and len(target_normalized) == 8:
        target_normalized = target_normalized[2:]
    
    if cell_color_normalized.startswith('FF') and len(cell_color_normalized) == 8:
        cell_color_normalized = cell_color_normalized[2:]
    
    return cell_color_normalized == target_normalized

# ============================================================================
# KRA PARSER
# ============================================================================
class DynamicKRAParser:
    """Parse KRA sheet dynamically"""
    
    @staticmethod
    def find_section_headers(sheet):
        """Find section headers"""
        sections = {}
        nta_found = False
        
        for row_idx in range(1, sheet.max_row + 1):
            cell = sheet.cell(row=row_idx, column=1)
            if not cell.value:
                continue
            
            cell_value = str(cell.value).strip()
            
            if re.search(r'(Tower|Green|Structure|External Development|External Dev|NTA)', cell_value, re.IGNORECASE):
                # Skip duplicate NTA sections
                if 'NTA' in cell_value.upper() and nta_found:
                    logger.info(f"Skipping duplicate NTA section: '{cell_value}'")
                    continue
                
                if 'NTA' in cell_value.upper():
                    nta_found = True
                
                sections[cell_value] = row_idx
                logger.info(f"Found section: '{cell_value}'")
        
        return sections
    
    @staticmethod
    def find_month_columns(sheet, start_row, report_months):
        """Find month columns"""
        month_cols = {}
        
        for row_idx in range(start_row, min(start_row + 3, sheet.max_row + 1)):
            for col_idx in range(1, 10):
                cell = sheet.cell(row=row_idx, column=col_idx)
                if cell.value and str(cell.value).strip() in report_months:
                    month_cols[str(cell.value).strip()] = col_idx
            
            if len(month_cols) >= 2:
                return month_cols, row_idx
        
        return {}, None
    
    @staticmethod
    def extract_activities(sheet, header_row, month_cols, report_months, is_structure=False, is_external=False, is_nta=False):
        """Extract activities"""
        activities = []
        
        if is_structure:
            for row_idx in range(header_row + 1, min(header_row + 20, sheet.max_row + 1)):
                cell = sheet.cell(row=row_idx, column=1)
                if cell.value and str(cell.value).strip().lower() == 'total':
                    activities.append({'name': 'Slab Casting', 'row': row_idx, 'unit': 'Slab'})
                    break
        elif is_nta:
            for row_idx in range(header_row + 1, min(header_row + 50, sheet.max_row + 1)):
                cell = sheet.cell(row=row_idx, column=1)
                if not cell.value:
                    continue
                
                activity_name = str(cell.value).strip()
                
                if activity_name.lower().startswith('total'):
                    break
                
                if not activity_name:
                    continue
                
                has_data = False
                for col_idx in month_cols.values():
                    val = sheet.cell(row=row_idx, column=col_idx).value
                    if val and extract_number(val) > 0:
                        has_data = True
                        break
                
                if has_data:
                    activities.append({'name': activity_name, 'row': row_idx, 'unit': 'Flat'})
        elif is_external:
            for row_idx in range(header_row + 1, min(header_row + 50, sheet.max_row + 1)):
                cell = sheet.cell(row=row_idx, column=1)
                if not cell.value:
                    continue
                
                activity_name = str(cell.value).strip()
                
                if activity_name.lower().startswith('total'):
                    break
                
                if not activity_name:
                    continue
                
                has_data = False
                for col_idx in month_cols.values():
                    val = sheet.cell(row=row_idx, column=col_idx).value
                    if val and extract_number(val) > 0:
                        has_data = True
                        break
                
                if has_data:
                    activities.append({'name': activity_name, 'row': row_idx, 'unit': 'Flat'})
        else:
            for row_idx in range(header_row + 1, min(header_row + 50, sheet.max_row + 1)):
                cell = sheet.cell(row=row_idx, column=1)
                if not cell.value:
                    continue
                
                activity_name = str(cell.value).strip()
                
                if re.search(r'(Tower|Green|Finishing|Structure|External Development|NTA)\s*\d*', activity_name, re.IGNORECASE):
                    if activity_name.lower().startswith('tower') or 'external' in activity_name.lower():
                        break
                    continue
                
                if not activity_name or activity_name.lower().startswith('total'):
                    continue
                
                has_data = False
                for col_idx in month_cols.values():
                    val = sheet.cell(row=row_idx, column=col_idx).value
                    if val and extract_number(val) > 0:
                        has_data = True
                        break
                
                if has_data:
                    activities.append({'name': activity_name, 'row': row_idx, 'unit': 'Flat'})
        
        return activities
    
    @classmethod
    def parse_kra_dynamic(cls, sheet, report_months):
        """Parse KRA sheet with data type detection (count vs percentage)"""
        logger.info("=== PARSING KRA ===\n")
        
        sections = cls.find_section_headers(sheet)
        if not sections:
            logger.error("No sections found")
            return {}
        
        parsed = {}
        
        # Keywords that indicate percentage-based activities
        PERCENTAGE_KEYWORDS = [
            'handover', 'lift', 'external','lift handover'
        ]
        
        for section_name, start_row in sorted(sections.items(), key=lambda x: x[1]):
            logger.info(f"Parsing: {section_name}")
            
            month_cols, month_row = cls.find_month_columns(sheet, start_row, report_months)
            
            if not month_cols:
                logger.warning(f"No month columns found")
                continue
            
            is_structure = 'Structure' in section_name or 'Slab' in section_name
            is_external = 'External Development' in section_name
            is_nta = 'NTA' in section_name
            
            activities = cls.extract_activities(sheet, month_row, month_cols, report_months, is_structure, is_external, is_nta)
            
            if not activities:
                logger.warning(f"No activities found for {section_name}")
                # Create empty activity placeholder to show empty row
                activities = [{'name': '', 'row': month_row + 1, 'unit': 'Flat'}]
            
            logger.info(f"Found {len(activities)} activities")
            
            targets = {}
            activity_data_types = {}  # NEW: Track data type per activity
            
            for activity in activities:
                targets[activity['name']] = {}
                activity_data_types[activity['name']] = 'count'  # Default
                
                # Skip reading targets for empty activities
                if not activity['name']:
                    for month in report_months:
                        targets[activity['name']][month] = 0
                    continue
                
                row_idx = activity['row']
                cell_has_percentage_symbol = False
                
                # ===== COLLECT VALUES AND CHECK FOR PERCENTAGE SYMBOL =====
                for month, col_idx in month_cols.items():
                    cell_value = sheet.cell(row=row_idx, column=col_idx).value
                    cell_str = str(cell_value).strip() if cell_value else ""
                    
                    # Check if cell contains % symbol BEFORE extracting number
                    if '%' in cell_str:
                        cell_has_percentage_symbol = True
                    
                    val = extract_number(cell_value)
                    
                    if 0 <= val <= 1 and is_external:
                        val = val * 100
                    
                    targets[activity['name']][month] = val
                
                # ===== AUTO-DETECT DATA TYPE =====
                detected_type = 'count'  # Default
                detection_reason = ""
                
                # Rule 1: If cell has % symbol → PERCENTAGE
                if cell_has_percentage_symbol:
                    detected_type = 'percentage'
                    detection_reason = "% symbol in cell"
                
                # Rule 2: Check activity name for percentage keywords → PERCENTAGE
                elif any(kw in activity['name'].lower() for kw in PERCENTAGE_KEYWORDS):
                    detected_type = 'percentage'
                    detection_reason = "activity name keyword"
                
                # Rule 3: External or NTA section → PERCENTAGE
                elif is_external or is_nta:
                    detected_type = 'percentage'
                    detection_reason = "external/NTA section type"
                
                else:
                    detected_type = 'count'
                    detection_reason = "default count type"
                
                activity_data_types[activity['name']] = detected_type
                logger.info(f"  ✓ {activity['name']} → {detected_type.upper()} ({detection_reason})")
            
            parsed[section_name] = {
                'activities': activities,
                'month_cols': month_cols,
                'targets': targets,
                'data_types': activity_data_types,  # NEW: Add data types to output
                'is_structure': is_structure,
                'is_external': is_external,
                'is_nta': is_nta
            }
        
        return parsed

# ============================================================================
# TRACKER PARSERS
# ============================================================================
class TowerTrackerParser:
    """Parse tower tracker files - finishing work"""
    
    @staticmethod
    def find_headers(ws):
        """Find Activity and Actual Finish columns"""
        header_row = None
        activity_col = None
        finish_col = None
        
        for row_idx in range(1, min(10, ws.max_row + 1)):
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    cell_val = str(cell.value).strip().lower()
                    if 'activity' in cell_val or 'task' in cell_val:
                        activity_col = col_idx
                    if 'actual' in cell_val and 'finish' in cell_val:
                        finish_col = col_idx
            
            if activity_col and finish_col:
                header_row = row_idx
                break
        
        return header_row, activity_col, finish_col
    
    @staticmethod
    def parse_date(finish_cell):
        """Parse date from cell"""
        if isinstance(finish_cell, datetime):
            return finish_cell
        elif isinstance(finish_cell, str):
            for fmt in ["%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y", "%d/%m/%Y"]:
                try:
                    return datetime.strptime(finish_cell.strip(), fmt)
                except ValueError:
                    continue
        return None
    
    @staticmethod
    def extract_percentage(cell_value):
        """Extract percentage from cell if it exists"""
        if not cell_value:
            return None
        
        cell_str = str(cell_value).strip()
        
        # Check if cell contains % symbol
        if '%' not in cell_str:
            return None
        
        # Remove % and extract number
        cell_str = cell_str.replace("%", "").strip()
        match = re.search(r"(\d+(?:\.\d+)?)", cell_str)
        
        if match:
            try:
                pct = float(match.group(1))
                # If decimal format (0-1), convert to percentage
                if 0 <= pct < 1:
                    pct = pct * 100
                return max(0, min(pct, 100))
            except ValueError:
                return None
        
        return None
    
    @classmethod
    def parse(cls, tracker_bytes, activities, month_num):
        """Parse tower tracker - returns both counts and percentages"""
        logger.info(f"=== PARSING TOWER TRACKER ===")
        wb = load_workbook(filename=tracker_bytes, data_only=True)
        
        tracker_counts = {act['name']: 0 for act in activities}
        tracker_percentages = {act['name']: None for act in activities}
        found_counts = {act['name']: False for act in activities}
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            header_row, activity_col, finish_col = cls.find_headers(ws)
            if not header_row or not finish_col:
                continue
            
            for activity in activities:
                activity_name = activity['name']
                
                if found_counts[activity_name]:
                    continue
                
                sheet_count = 0
                sheet_percentage = None
                
                for row_idx in range(header_row + 1, ws.max_row + 1):
                    activity_cell = ws.cell(row=row_idx, column=activity_col).value
                    finish_cell = ws.cell(row=row_idx, column=finish_col).value
                    
                    if not activity_cell:
                        continue
                    
                    if activity_name.lower() not in str(activity_cell).strip().lower():
                        continue
                    
                    # Check if finish_cell contains a percentage
                    pct = cls.extract_percentage(finish_cell)
                    if pct is not None:
                        sheet_percentage = pct
                        logger.info(f"{activity_name}: {pct}% (from tracker cell)")
                        continue
                    
                    # Otherwise treat as date count
                    finish_date = cls.parse_date(finish_cell)
                    if finish_date and finish_date.month == month_num:
                        sheet_count += 1
                
                if sheet_percentage is not None:
                    tracker_percentages[activity_name] = sheet_percentage
                    tracker_counts[activity_name] = sheet_percentage
                    found_counts[activity_name] = True
                    logger.info(f"{activity_name}: {sheet_percentage}% (percentage-based)")
                elif sheet_count > 0:
                    tracker_counts[activity_name] = sheet_count
                    found_counts[activity_name] = True
                    logger.info(f"{activity_name}: {sheet_count} completed (count-based)")
        
        logger.info(f"Tower tracker parsed")
        return tracker_counts, 'count'


class ExternalDevelopmentParser:
    """Parse external development tracker files"""
    
    @staticmethod
    def find_headers(ws):
        """Find Activity and % Complete columns"""
        header_row = None
        activity_col = None
        status_col = None
        
        for row_idx in range(1, min(6, ws.max_row + 1)):
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if not cell.value:
                    continue
                    
                cell_val = str(cell.value).strip().lower()
                
                if 'activity' in cell_val:
                    activity_col = col_idx
                    
                if '% complete' in cell_val or cell_val == '% complete':
                    status_col = col_idx
            
            if activity_col and status_col:
                header_row = row_idx
                logger.info(f"Found headers at row {row_idx}: Activity col={activity_col}, Status col={status_col}")
                break
        
        return header_row, activity_col, status_col
    
    @staticmethod
    def parse_percentage(status_value):
        """Extract percentage completion from status"""
        if status_value is None:
            return 0.0
        
        try:
            val = float(status_value)
            if 0 < val <= 1:
                val = val * 100
            val = max(0, min(val, 100))
            return val
        except (ValueError, TypeError):
            pass
        
        status_str = str(status_value).strip()
        
        if not status_str or status_str == "-":
            return 0.0
        
        status_str = status_str.replace("%", "").strip()
        
        match = re.search(r"(\d+(?:\.\d+)?)", status_str)
        if match:
            try:
                pct = float(match.group(1))
                if 0 < pct <= 1:
                    pct = pct * 100
                pct = max(0, min(pct, 100))
                return pct
            except ValueError:
                return 0.0
        
        return 0.0
    
    @classmethod
    def parse_with_targets(cls, tracker_bytes, activities, kra_targets):
        """Parse tracker and extract % complete"""
        logger.info(f"=== PARSING EXTERNAL DEVELOPMENT TRACKER WITH TARGETS ===")
        
        wb = load_workbook(filename=tracker_bytes, data_only=True)
        tracker_data = {}
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            logger.info(f"Processing sheet: {sheet_name}")
            
            header_row, activity_col, status_col = cls.find_headers(ws)
            
            if not header_row or not activity_col or not status_col:
                logger.warning(f"Could not find headers in sheet: {sheet_name}")
                continue
            
            for row_idx in range(header_row + 1, ws.max_row + 1):
                activity_cell = ws.cell(row=row_idx, column=activity_col)
                status_cell = ws.cell(row=row_idx, column=status_col)
                
                if not activity_cell.value:
                    continue
                
                activity_header = str(activity_cell.value).strip()
                completion_pct = cls.parse_percentage(status_cell.value)
                
                logger.info(f"Row {row_idx}: {activity_header} | Completion: {completion_pct}%")
                
                matching_activity = None
                for activity in activities:
                    kra_activity_name = activity['name']
                    if (kra_activity_name.lower() in activity_header.lower() or 
                        activity_header.lower() in kra_activity_name.lower()):
                        matching_activity = kra_activity_name
                        break
                
                if not matching_activity:
                    logger.warning(f"No KRA match found for: {activity_header}")
                    continue
                
                if matching_activity not in tracker_data:
                    tracker_data[matching_activity] = completion_pct
                    logger.info(f"Activity: {activity_header}, Completion: {completion_pct}%")
        
        logger.info(f"Total extracted: {len(tracker_data)} activities")
        return tracker_data


class StructureWorkParser:
    """Parse Structure Work Tracker (Eligo format) to count green slab casting dates."""

    TOWER_COLUMNS = {
        'Tower H': ['AB', 'AF', 'AJ', 'AN','AR','AV','AZ'],
        'Tower G': ['N', 'R', 'V'],
        'Tower F': ['D', 'H']
    }

    FLOOR_LABELS = ['1F', '2F', '3F', '4F', '5F', '6F', '7F', '8F']

    @staticmethod
    def extract_tower_from_section(section_name):
        """Extract tower name (e.g., 'Tower H') from section name."""
        match = re.search(r'Tower\s+([A-Z])', str(section_name), re.IGNORECASE)
        return f"Tower {match.group(1).upper()}" if match else None

    @staticmethod
    def col_letter_to_index(col_letter):
        """Convert Excel column letter to numeric index."""
        result = 0
        for char in col_letter.upper():
            result = result * 26 + (ord(char) - ord('A') + 1)
        return result

    @staticmethod
    def find_target_sheet(wb):
        """Find the Eligo slab cycle sheet."""
        target_names = [
            "Revised Baselines- 25 days SC",
            "ELIGO SLAB CYCLE"
        ]
        
        for target in target_names:
            if target in wb.sheetnames:
                logger.info(f"Found sheet: {target}")
                return wb[target]
        
        for name in wb.sheetnames:
            name_lower = name.lower()
            if ('revised' in name_lower and 'baseline' in name_lower) or \
               ('eligo' in name_lower and 'slab' in name_lower) or \
               ('slab' in name_lower and 'cycle' in name_lower):
                logger.info(f"Found sheet (partial match): {name}")
                return wb[name]
        
        logger.warning(f"Sheet not found. Available: {wb.sheetnames}")
        return None

    @staticmethod
    def find_floor_rows(ws):
        """Find rows containing floor data (1F, 2F, etc.)."""
        floor_rows = {}
        
        for row_idx in range(1, min(50, ws.max_row + 1)):
            for col_idx in range(1, 5):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    cell_str = str(cell_value).strip().upper()
                    if cell_str in ['1F', '2F', '3F', '4F', '5F', '6F', '7F', '8F']:
                        floor_rows[cell_str] = row_idx
                        logger.debug(f"Found floor {cell_str} at row {row_idx}")
        
        return floor_rows

    @staticmethod
    def parse_date(value):
        """Parse date from cell value."""
        if not value:
            return None
        if isinstance(value, datetime):
            return value

        s = str(value).strip()
        if not s or s in ('-', '###', ''):
            return None

        for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y", "%d.%m.%Y", "%d-%b-%Y"):
            try:
                return datetime.strptime(s, fmt)
            except ValueError:
                continue
        return None

    @classmethod
    def parse_green_dates(cls, tracker_bytes, section_name, target_month_num=None):
        """Extract green anticipated dates (slab castings) for a given tower."""
        logger.info("=== PARSING STRUCTURE - SLAB CASTING (ELIGO - GREEN ANTICIPATED DATES) ===")

        target_tower = cls.extract_tower_from_section(section_name)
        if not target_tower:
            logger.warning(f"Could not extract tower from section name: {section_name}")
            return []

        tower_cols = cls.TOWER_COLUMNS.get(target_tower)
        if not tower_cols:
            logger.warning(f"No column mapping found for {target_tower}")
            return []

        logger.info(f"Tower: {target_tower} | Target month: {target_month_num}")
        logger.info(f"Scanning Anticipated columns: {tower_cols}")

        try:
            wb = load_workbook(filename=tracker_bytes, data_only=False)
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return []

        ws = cls.find_target_sheet(wb)
        if not ws:
            logger.warning(f"No Eligo Slab Cycle sheet found")
            return []

        logger.info(f"Using sheet: {ws.title}")

        floor_rows = cls.find_floor_rows(ws)
        if not floor_rows:
            logger.warning("Could not find floor rows, using default range")
            floor_rows = {f"{i}F": (i+3) for i in range(1, 9)}

        logger.info(f"Found {len(floor_rows)} floor rows: {list(floor_rows.keys())}")

        green_dates = []
        green_cells_found = 0
        total_cells_checked = 0

        col_indices = [cls.col_letter_to_index(col) for col in tower_cols]

        for floor_label in sorted(floor_rows.keys()):
            row_idx = floor_rows[floor_label]
            
            for col_letter, col_idx in zip(tower_cols, col_indices):
                total_cells_checked += 1
                cell = ws.cell(row=row_idx, column=col_idx)
                
                if not cell.value:
                    continue
                
                if is_cell_green(cell):
                    green_cells_found += 1
                    parsed_date = cls.parse_date(cell.value)
                    
                    if parsed_date:
                        if target_month_num is None or parsed_date.month == target_month_num:
                            green_dates.append(parsed_date)
                            logger.info(f"✓ Green date found: {parsed_date.strftime('%d-%b-%Y')} "
                                       f"(Floor {floor_label}, Col {col_letter})")
                    else:
                        logger.warning(f"Could not parse date from {col_letter}{row_idx}: '{cell.value}'")
                else:
                    cell_color = get_cell_hex_color(cell)
                    logger.debug(f"Cell {col_letter}{row_idx}: Not green (color: {cell_color}, value: {cell.value})")

        logger.info(f"Checked {total_cells_checked} cells | Found {green_cells_found} green cells")

        if not green_dates:
            logger.warning(f"No green anticipated dates found for {target_tower}")
            return []

        latest_year = max(d.year for d in green_dates)
        filtered_dates = [d for d in green_dates if d.year == latest_year]

        logger.info(f"Total green dates: {len(green_dates)} | "
                   f"Filtered to {latest_year}: {len(filtered_dates)} dates")

        return sorted(filtered_dates)

        
# ============================================================================
# REPORT GENERATOR
# ============================================================================
class EligoReportGenerator:
    
    
    def _forward_fill_percentages(self, activity_name, counts, months):
        """
        Forward fill percentage values across months.
        If an activity has a percentage value in a later month, copy it to all earlier months.
        This applies to completion-based activities like C-Gypsum, Lift Installation, etc.
        """
        months_list = list(months)
        filled_counts = dict(counts[activity_name])
        
        # Find the first non-zero percentage value (scanning from right/last month)
        target_value = None
        target_month_idx = None
        
        for idx in range(len(months_list) - 1, -1, -1):
            month = months_list[idx]
            month_val = filled_counts.get(month, 0)
            try:
                pct = float(str(month_val).strip('%')) if isinstance(month_val, str) else float(month_val)
            except (ValueError, AttributeError):
                pct = 0.0
            
            if pct > 0:
                target_value = pct
                target_month_idx = idx
                break
        
        # If found a non-zero value, copy it to all earlier months
        if target_value is not None and target_month_idx is not None:
            for idx in range(target_month_idx):
                filled_counts[months_list[idx]] = target_value
        
        return filled_counts

    def _is_completion_activity(self, activity_name):
        """
        Check if activity is a completion-based milestone (reaches 100% and stays there).
        Examples: Lift Installation, C-Gypsum and POP punning, Finishing, etc.
        """
        completion_keywords = [
            'lift installation',
            'c-gypsum',
            'c-gypsum and pop',
            'pop punning',
            'final handover',
            'completion',
            'ready for possession',
            'finishing'
        ]
        
        activity_lower = activity_name.lower()
        return any(keyword in activity_lower for keyword in completion_keywords)

    def _is_common_area_section(self, section_name):
        """Check if section is from Common Area sheet"""
        section_lower = section_name.lower()
        return 'common area' in section_lower or 'comman area' in section_lower or 'common' in section_lower
    
    
    
    """Generate quarterly report for ELIGO"""
    
    def __init__(self):
        self.cos = init_cos()
        self.kra_key = None
        self.current_quarter = None
        self.quarter_months = []
        self.tracker_months = []
        self.quarter_year = None
        self.kra_data = {}
        self.tracker_keys = {}
        self.structure_green_dates = {}
    
    def get_latest_kra(self):
        """Find latest KRA file using year and quarter comparison"""
        logger.info("=== FINDING LATEST KRA ===")
        
        all_files = list_files(self.cos, "")
        kra_files = [f for f in all_files if 'KRA' in os.path.basename(f).upper() and f.endswith('.xlsx')]
        
        if not kra_files:
            logger.error("No KRA files found")
            return False
        
        logger.info(f"Found {len(kra_files)} KRA files")
        
        # Step 1: Extract quarter info from each KRA filename
        kra_with_quarters = []
        
        for file_path in kra_files:
            filename = os.path.basename(file_path)
            quarter, months, year = determine_quarter_from_kra(filename)
            
            if quarter and months and year:
                # Convert quarter to numeric for sorting (Q1→1, Q2→2, Q3→3, Q4→4)
                quarter_num = int(quarter[1])
                
                kra_with_quarters.append({
                    'path': file_path,
                    'filename': filename,
                    'quarter': quarter,
                    'months': months,
                    'year': year,
                    'sort_key': (year, quarter_num)
                })
                
                logger.info(f"  ✓ {filename} → Q{quarter_num} {year}")
            else:
                logger.warning(f"  ✗ {filename} → Could not extract quarter info")
        
        # Step 2: Validate we have at least one valid KRA file
        if not kra_with_quarters:
            logger.error("No KRA files with valid quarter information found")
            logger.error("Expected format: 'KRA Milestones for Month1 Month2 Month3 YYYY.xlsx'")
            return False
        
        # Step 3: Sort by (Year, Quarter) and pick the latest
        kra_with_quarters.sort(key=lambda x: x['sort_key'])
        latest = kra_with_quarters[-1]
        
        # Step 4: Set instance variables from latest KRA
        self.kra_key = latest['path']
        self.current_quarter = latest['quarter']
        self.quarter_months = latest['months']
        self.quarter_year = latest['year']
        self.tracker_months = [MONTH_SHIFT_MAP.get(m, m) for m in self.quarter_months]
        
        logger.info(f"\n{'='*80}")
        logger.info(f"Latest KRA: {latest['filename']}")
        logger.info(f"Quarter: {self.current_quarter} {self.quarter_year}")
        logger.info(f"Report Months: {self.quarter_months}")
        logger.info(f"Tracker Months: {self.tracker_months}")
        logger.info(f"{'='*80}\n")
        
        return True
    
    def parse_kra(self):
        """Parse KRA file"""
        logger.info("\n=== PARSING KRA FILE ===")
        
        raw = download_file(self.cos, self.kra_key)
        wb = load_workbook(filename=BytesIO(raw), data_only=True)
        
        kra_sheet = None
        for sheet_name in wb.sheetnames:
            if 'eligo' in sheet_name.lower() and 'target' in sheet_name.lower():
                kra_sheet = wb[sheet_name]
                logger.info(f"Using sheet: {sheet_name}")
                break
        
        if not kra_sheet:
            logger.error("No ELIGO sheet found")
            return False
        
        parser = DynamicKRAParser()
        self.kra_data = parser.parse_kra_dynamic(kra_sheet, self.quarter_months)
        
        if not self.kra_data:
            logger.error("Failed to parse KRA")
            return False
        
        logger.info(f"Parsed {len(self.kra_data)} sections")
        return True
    
    def find_trackers(self):
        """Find tracker files - match tracker month to report month"""
        logger.info("\n=== FINDING TRACKER FILES ===")
        logger.info(f"Report Months: {self.quarter_months}")
        logger.info(f"Tracker Months: {self.tracker_months}")
        
        eligo_files = list_files(self.cos, "Eligo/")
        logger.info(f"Total files in Eligo/: {len(eligo_files)}\n")
        
        for section_name in self.kra_data.keys():
            self.tracker_keys[section_name] = {}
            
            # Determine tracker pattern and type
            is_external = False
            is_nta = False
            is_structure = False
            tracker_pattern = None
            
            if 'External Development' in section_name:
                tracker_pattern = 'External Development'
                is_external = True
            elif 'NTA' in section_name:
                tracker_pattern = 'NTA.*Finishing'
                is_nta = True
            elif 'Structure' in section_name:
                tracker_pattern = 'Structure Work Tracker'
                is_structure = True
            else:
                tower_match = re.search(r'Tower\s+([A-Z])', section_name, re.IGNORECASE)
                if not tower_match:
                    logger.warning(f"Could not extract tower from: {section_name}")
                    continue
                tracker_pattern = f'Tower {tower_match.group(1).upper()}.*Finishing'
            
            logger.info(f"{'='*80}")
            logger.info(f"Section: {section_name}")
            logger.info(f"Pattern: {tracker_pattern}")
            logger.info(f"Type: {'TOWER' if not (is_external or is_nta or is_structure) else 'EXTERNAL' if is_external else 'NTA' if is_nta else 'STRUCTURE'}")
            logger.info(f"{'='*80}")
            
            # ===== TOWER SECTIONS: Match by tracker month =====
            if not (is_external or is_nta or is_structure):
                for report_month, tracker_month in zip(self.quarter_months, self.tracker_months):
                    tracker_month_num = MONTH_TO_NUM.get(tracker_month)
                    
                    # Calculate the year for this tracker month
                    # For Q3 2026: all tracker months (Jan, Feb, Mar) are in 2026
                    tracker_year = self.quarter_year
                    if self.current_quarter == 'Q4' and tracker_month == 'June':
                        tracker_year = self.quarter_year + 1
                    
                    logger.info(f"\n  Report: {report_month} → Tracker: {tracker_month} ({tracker_month_num}/{tracker_year})")
                    
                    found = False
                    for file_path in eligo_files:
                        filename = os.path.basename(file_path)
                        
                        # Check pattern match
                        if not re.search(tracker_pattern, filename, re.IGNORECASE):
                            continue
                        
                        # Extract date
                        file_date = extract_date_from_filename(filename)
                        if not file_date:
                            logger.debug(f"    Could not extract date from: {filename}")
                            continue
                        
                        # Check if month and year match
                        if file_date.month == tracker_month_num and file_date.year == tracker_year:
                            self.tracker_keys[section_name][report_month] = file_path
                            logger.info(f"    ✅ Found: {filename}")
                            found = True
                            break
                    
                    if not found:
                        logger.warning(f"    ❌ Not found")
            
            # ===== EXTERNAL/NTA/STRUCTURE: Use OLD LOGIC (find any matching tracker) =====
            else:
                for report_month, tracker_month in zip(self.quarter_months, self.tracker_months):
                    tracker_month_num = MONTH_TO_NUM.get(tracker_month)
                    
                    # Calculate the year for this tracker month
                    # For Q3 2026: all tracker months (Jan, Feb, Mar) are in 2026
                    tracker_year = self.quarter_year
                    if self.current_quarter == 'Q4' and tracker_month == 'June':
                        tracker_year = self.quarter_year + 1
                    
                    logger.info(f"\n  Report: {report_month} → Tracker: {tracker_month} ({tracker_month_num}/{tracker_year})")
                    
                    found = False
                    for file_path in eligo_files:
                        filename = os.path.basename(file_path)
                        
                        if re.search(tracker_pattern, filename, re.IGNORECASE):
                            file_date = extract_date_from_filename(filename)
                            
                            if file_date and file_date.month == tracker_month_num and file_date.year == tracker_year:
                                self.tracker_keys[section_name][report_month] = file_path
                                logger.info(f"    ✅ Found: {filename}")
                                found = True
                                break
                    
                    if not found:
                        logger.warning(f"    ❌ Not found")
            
            logger.info(f"\n{'='*80}")
            logger.info("FINAL TRACKER ASSIGNMENT")
            logger.info(f"{'='*80}")
            for section_name, months_dict in self.tracker_keys.items():
                logger.info(f"\n{section_name}:")
                for month in self.quarter_months:
                    if month in months_dict:
                        logger.info(f"  {month}: ✅ {os.path.basename(months_dict[month])}")
                    else:
                        logger.info(f"  {month}: ❌ (blank)")
    
    def _parse_tracker(self, tracker_file, section_name, activities, is_structure, is_nta=False, report_month=None):
        """Parse tracker file and return (counts, data_type)"""
        try:
            raw = download_file(self.cos, tracker_file)
            
            if is_structure:
                logger.info(f"Parser: STRUCTURE WORK (SLAB CASTING)")
                
                report_month_num = MONTH_TO_NUM.get(report_month)
                logger.info(f"Extracting green dates for: {report_month} (month {report_month_num})")
                
                green_dates = StructureWorkParser.parse_green_dates(BytesIO(raw), section_name, report_month_num)
                
                if not green_dates:
                    logger.info(f"No dates found for month {report_month_num}, trying all green dates...")
                    green_dates = StructureWorkParser.parse_green_dates(BytesIO(raw), section_name, None)
                    if green_dates:
                        logger.info(f"Found {len(green_dates)} green dates (all months, latest year)")
                
                if green_dates:
                    self.structure_green_dates[report_month] = green_dates
                    logger.info(f"Found {len(green_dates)} green dates for {report_month}")
                else:
                    logger.warning(f"No green dates found")
                
                tracker_counts = {'Slab Casting': len(green_dates)}
                return tracker_counts, 'count'
            
            elif is_nta:
                logger.info(f"Parser: NTA FINISHING")
                month_num = MONTH_TO_NUM.get(report_month)
                if not month_num:
                    logger.warning(f"No valid month")
                    return {act['name']: 0 for act in activities}, 'count'
                
                tracker_counts, data_type = TowerTrackerParser.parse(BytesIO(raw), activities, month_num)
                return tracker_counts, data_type
                
            elif 'External Development' in section_name:
                logger.info(f"Parser: EXTERNAL DEVELOPMENT")
                
                kra_targets = self.kra_data.get(section_name, {}).get('targets', {})
                
                tracker_percentages = ExternalDevelopmentParser.parse_with_targets(
                    BytesIO(raw), 
                    activities, 
                    kra_targets
                )
                
                tracker_counts = {}
                for activity in activities:
                    act_name = activity['name']
                    if act_name in tracker_percentages:
                        tracker_counts[act_name] = tracker_percentages[act_name]
                    else:
                        tracker_counts[act_name] = 0
                
                logger.info(f"External Development data type: percentage")
                return tracker_counts, 'percentage'
            
            else:
                logger.info(f"Parser: TOWER FINISHING")
                month_num = MONTH_TO_NUM.get(report_month)
                if not month_num:
                    logger.warning(f"No valid month")
                    return {act['name']: 0 for act in activities}, 'count'
                
                tracker_counts, data_type = TowerTrackerParser.parse(BytesIO(raw), activities, month_num)
                return tracker_counts, data_type
        
        except Exception as e:
            logger.error(f"Error: {e}")
            import traceback
            logger.debug(traceback.format_exc())
            return {act['name']: 0 for act in activities}, 'count'
    
    def build_report_data(self):
        """Build report data from trackers"""
        logger.info("\n=== BUILDING REPORT DATA ===")
        
        report_dfs = {}
        
        for section_name, section_data in self.kra_data.items():
            logger.info(f"\nProcessing: {section_name}")
            
            activities = section_data['activities']
            targets = section_data['targets']
            kra_data_types = section_data.get('data_types', {})  # NEW: Get KRA data types
            is_structure = section_data.get('is_structure', False)
            is_external = section_data.get('is_external', False)
            is_nta = section_data.get('is_nta', False)
            
            counts = {act['name']: {month: 0 for month in self.quarter_months} for act in activities}
            data_types = {}
            activity_tracker_months = {act['name']: None for act in activities}
            
            # ===== INITIALIZE DATA TYPES FROM KRA =====
            for activity in activities:
                act_name = activity['name']
                data_types[act_name] = kra_data_types.get(act_name, 'count')
            
            # Track which months have trackers for each activity
            months_with_trackers = {act['name']: set() for act in activities}
            
            if section_name in self.tracker_keys:
                logger.info(f"Found trackers for {len(self.tracker_keys[section_name])} months")
                for report_month, tracker_file in self.tracker_keys[section_name].items():
                    logger.info(f"Parsing {report_month}: {os.path.basename(tracker_file)}")
                    tracker_counts, data_type = self._parse_tracker(
                        tracker_file, section_name, activities, is_structure, is_nta, report_month
                    )
                    
                    for activity_name, count in tracker_counts.items():
                        if activity_name in counts:
                            counts[activity_name][report_month] = count
                            # Keep KRA data type (don't override with tracker type)
                            months_with_trackers[activity_name].add(report_month)
                            logger.info(f"{activity_name}: {count} (KRA type: {data_types[activity_name]})")
                            
                            if count > 0 and activity_tracker_months[activity_name] is None:
                                activity_tracker_months[activity_name] = report_month
                                logger.info(f"  → Tracker month for {activity_name}: {report_month}")
            else:
                logger.warning(f"No trackers found for {section_name}")
            
            df = self._build_dataframe(
                section_name, activities, targets, counts, data_types, 
                activity_tracker_months, months_with_trackers
            )
            report_dfs[section_name] = df
        
        return report_dfs
    
    def _build_dataframe(self, section_name, activities, targets, counts, data_types, 
        activity_tracker_months=None, months_with_trackers=None, tracker_file_used=None):
        """Build milestone dataframe with debug logging - handles count and percentage data types"""
        data = []
        total_acts = len(activities)
        weightage = round(100 / total_acts, 2) if total_acts else 0
        
        is_external = 'External Development' in section_name
        is_nta = 'NTA' in section_name
        is_tower = not (is_external or is_nta or 'Structure' in section_name)
        
        logger.info(f"\n{'='*80}")
        logger.info(f"BUILDING DATAFRAME: {section_name}")
        logger.info(f"is_tower={is_tower}, is_nta={is_nta}, is_external={is_external}")
        logger.info(f"{'='*80}")
        
        if activity_tracker_months is None:
            activity_tracker_months = {}
        if months_with_trackers is None:
            months_with_trackers = {act['name']: set() for act in activities}
        if tracker_file_used is None:
            tracker_file_used = {act['name']: {month: None for month in self.quarter_months} for act in activities}
        
        for i, activity in enumerate(activities):
            name = activity['name']
            unit = activity.get('unit', 'Flat')
            unit_plural = f"{unit}s"
            
            # ===== GET DATA TYPE FOR THIS ACTIVITY =====
            activity_data_type = data_types.get(name, 'count')
            
            logger.info(f"\n{'─'*80}")
            logger.info(f"Activity {i+1}: {name}")
            logger.info(f"Unit: {unit}")
            logger.info(f"Targets: {targets.get(name, {})}")
            logger.info(f"Counts: {counts.get(name, {})}")
            logger.info(f"Data Type: {activity_data_type}")
            logger.info(f"Months with trackers: {months_with_trackers.get(name, set())}")
            logger.info(f"{'─'*80}")
            
            # Skip empty activities
            if not name:
                row = {
                    "Milestone": "",
                    "Activity": "",
                    "Target": "",
                    "Weightage": "",
                    "Weighted Delay against Targets": "",
                    f"Delay Reasons_{self.quarter_months[-1]} {self.quarter_year}": "",
                }
                for m in self.quarter_months:
                    row[f"% Work Done against Target-Till {m}"] = ""
                    row[f"Target achieved in {m}"] = ""
                row["Total achieved"] = ""
                data.append(row)
                continue
            
            is_completion = self._is_completion_activity(name)
            if is_external and is_completion:
                counts[name] = self._forward_fill_percentages(name, counts, self.quarter_months)
            
            row = {
                "Milestone": f"{i+1:02d}",
                "Activity": name,
                "Weightage": weightage,
                "Weighted Delay against Targets": "",
                f"Delay Reasons_{self.quarter_months[-1]} {self.quarter_year}": "",
            }
            
            total_target = 0
            target_parts = []
            is_common_area = self._is_common_area_section(section_name)
            
            # ===== SET TARGET DISPLAY =====
            if is_external or is_common_area:
                row["Target"] = "External/Common Area"
            elif activity_data_type == 'percentage':
                # For percentage activities, show ONLY non-zero values with month names
                percentage_parts = []
                for month in self.quarter_months:
                    target_val = targets[name].get(month, 0)
                    # Convert decimal to percentage if needed
                    try:
                        target_val = float(target_val)
                        # Convert decimal (0-1) to percentage (0-100)
                        if 0 <= target_val < 1:
                            target_val = target_val * 100
                        elif target_val == 1.0:
                            target_val = 100.0
                        
                        # Format as integer if whole number, else keep 2 decimals
                        if target_val == int(target_val):
                            target_val = int(target_val)
                        else:
                            target_val = round(target_val, 2)
                    except (ValueError, TypeError):
                        target_val = 0
                    
                    # MODIFIED: Only add to display if value is > 0
                    if target_val > 0:
                        percentage_parts.append(f"{target_val}% ({month})")
                
                row["Target"] = ", ".join(percentage_parts) if percentage_parts else "0%"
            else:
                for month in self.quarter_months:
                    target = int(targets[name].get(month, 0))
                    total_target += target
                    if target > 0:
                        target_parts.append(f"{target}-{month}")
                row["Target"] = f"{total_target} {unit_plural} ({', '.join(target_parts)})" if target_parts else f"0 {unit_plural}"
            
            cum_done = 0
            cum_target = 0
            total_achieved = 0
            last_pct = 0.0
            last_tracker_file = None
            
            logger.info(f"\nMonth-by-month calculation:")
            
            for month_idx, month in enumerate(self.quarter_months):
                logger.info(f"\n  {month}:")
                
                # ===== PERCENTAGE-BASED ACTIVITIES =====
                if activity_data_type == 'percentage':
                   logger.info(f"    [PERCENTAGE-BASED]")
                   
                   # Get value from tracker (counts) - this contains the % Complete value
                   month_value = counts[name].get(month, 0)
                   has_tracker = month in months_with_trackers[name]
                   
                   # Find which month has the KRA target (e.g., February for "33% (February)")
                   target_month_index = None
                   target_pct = None
                   for idx, m in enumerate(self.quarter_months):
                       kra_target = targets[name].get(m, 0)
                       if kra_target != 0:
                           target_month_index = idx
                           try:
                               target_pct = float(kra_target)
                               if 0 <= target_pct < 1:
                                   target_pct = target_pct * 100
                               elif target_pct == 1.0:
                                   target_pct = 100.0
                           except (ValueError, TypeError):
                               target_pct = 0.0
                           break
                   
                   logger.info(f"    Target month index: {target_month_index}, Target %: {target_pct}")
                   logger.info(f"    Current month index: {month_idx}")
                   logger.info(f"    Tracker % Complete value: {month_value}, Has tracker: {has_tracker}")
                   
                   # Check if month_value is actually meaningful (not just 0 from initialization)
                   has_meaningful_tracker_value = has_tracker and month_value != 0
                   
                   logger.info(f"    Has meaningful tracker value: {has_meaningful_tracker_value}")
                   
                   # NEW LOGIC: Prioritize KRA target forward-fill over tracker data
                   if target_month_index is not None and target_pct is not None:
                       # Has KRA target → Use forward-fill logic
                       
                       if month_idx < target_month_index:
                            # Before target month
                            if has_tracker:
                                # Has tracker → Show 100%
                                logger.info(f"    Before target month WITH tracker → Show 100%")
                                row[f"% Work Done against Target-Till {month}"] = "100%"
                                row[f"Target achieved in {month}"] = "100%"
                                last_pct = 100.0
                            else:
                                # No tracker → Show 0%
                                logger.info(f"    Before target month WITHOUT tracker → Show 0%")
                                row[f"% Work Done against Target-Till {month}"] = ""
                                row[f"Target achieved in {month}"] = ""
                                last_pct = 0.0
                           
                       elif month_idx == target_month_index:
                           # Target month itself → Check if tracker has data
                           if has_meaningful_tracker_value:
                               # Use tracker data for target month
                               try:
                                   if isinstance(month_value, str):
                                       pct = float(month_value.strip('%'))
                                   else:
                                       pct = float(month_value)
                               except (ValueError, TypeError, AttributeError):
                                   pct = target_pct
                               
                               if 0 <= pct < 1:
                                   pct = pct * 100
                               elif pct == 1.0:
                                   pct = 100.0
                               
                               pct = min(max(pct, 0.0), 100.0)
                               
                               if pct == int(pct):
                                   pct_display = f"{int(pct)}%"
                               else:
                                   pct_display = f"{pct:.1f}%"
                               
                               logger.info(f"    Target month with tracker → {pct_display}")
                               row[f"% Work Done against Target-Till {month}"] = ""
                               row[f"Target achieved in {month}"] = f"{pct_display} Complete"
                               last_pct = pct
                           else:
                               # No tracker, use KRA target
                               if target_pct == int(target_pct):
                                   pct_display = f"{int(target_pct)}%"
                               else:
                                   pct_display = f"{target_pct:.1f}%"
                               
                               logger.info(f"    Target month without tracker → {pct_display}")
                               row[f"% Work Done against Target-Till {month}"] = ""
                               row[f"Target achieved in {month}"] = pct_display
                               last_pct = target_pct
                               
                       else:
                           # After target month → Blank
                           logger.info(f"    After target month → Blank in both columns")
                           row[f"% Work Done against Target-Till {month}"] = ""
                           row[f"Target achieved in {month}"] = ""
                   
                   elif has_meaningful_tracker_value:
                       # Has tracker but no KRA target → Show tracker value
                       try:
                           if isinstance(month_value, str):
                               pct = float(month_value.strip('%'))
                           else:
                               pct = float(month_value)
                       except (ValueError, TypeError, AttributeError):
                           pct = 0.0
                       
                       logger.info(f"    Has meaningful tracker (no KRA target) → {month_value}")
                       
                       if 0 <= pct < 1:
                           pct = pct * 100
                       elif pct == 1.0:
                           pct = 100.0
                       
                       pct = min(max(pct, 0.0), 100.0)
                       
                       if pct == int(pct):
                           pct_display = f"{int(pct)}%"
                       else:
                           pct_display = f"{pct:.1f}%"
                       
                       logger.info(f"    Final percentage: {pct_display}")
                       
                       row[f"% Work Done against Target-Till {month}"] = pct_display
                       row[f"Target achieved in {month}"] = f"{pct_display} Complete"
                       last_pct = pct
                   
                   else:
                       # No KRA target and no meaningful tracker → KEEP BLANK
                       logger.info("    No KRA target and no meaningful tracker → leave blank")
                       row[f"% Work Done against Target-Till {month}"] = ""
                       row[f"Target achieved in {month}"] = ""
    
                
                # ===== EXTERNAL/COMMON AREA ACTIVITIES =====
                elif is_external or is_common_area:
                    logger.info(f"    [EXTERNAL/COMMON AREA]")
                    row[f"% Work Done against Target-Till {month}"] = ""
                    row[f"Target achieved in {month}"] = ""
                
                # ===== COUNT-BASED ACTIVITIES =====
                else:
                    logger.info(f"    [COUNT-BASED]")
                    month_done = counts[name].get(month, 0)
                    month_target = int(targets[name].get(month, 0))
                    has_tracker = month in months_with_trackers[name]
                    
                    logger.info(f"    Done: {month_done}, Target: {month_target}, Has Tracker: {has_tracker}")
                    
                    if has_tracker:
                        cum_done += month_done
                        cum_target += month_target
                        total_achieved += month_done
                        
                        logger.info(f"    Cumulative: {cum_done}/{cum_target}")
                        
                        # Calculate percentage
                        if cum_done > 0 and cum_target == 0:
                            pct = 100.0
                            logger.info(f"    Logic: Done({cum_done}) > 0 and Target({cum_target}) == 0 → 100%")
                        elif cum_target == 0:
                            pct = 0.0
                            logger.info(f"    Logic: Done and Target both 0 → 0%")
                        else:
                            pct = (cum_done / cum_target) * 100
                            logger.info(f"    Logic: ({cum_done}/{cum_target}) * 100 = {pct}%")
                        
                        pct = min(pct, 100.0)
                        pct = round(pct, 2)
                        
                        logger.info(f"    Percentage: {pct}%")
                        
                        row[f"% Work Done against Target-Till {month}"] = f"{pct}%"
                        row[f"Target achieved in {month}"] = f"{int(month_done)} out of {int(month_target)} {unit_plural}"
                        last_pct = pct
                    
                    else:
                        logger.info(f"    NO TRACKER - leaving blank")
                        row[f"% Work Done against Target-Till {month}"] = ""
                        row[f"Target achieved in {month}"] = f"0 out of {int(month_target)} {unit_plural}"

            # ===== HARD-CODED OVERRIDES FOR FEBRUARY (USER REQUEST) =====
            if 'February' in self.quarter_months:
                if re.search(r'\bTower\s*F\b', section_name, re.IGNORECASE) and name.strip().lower() == 'external paint 1st coat':
                    row["% Work Done against Target-Till February"] = "100%"
                if re.search(r'\bTower\s*G\b', section_name, re.IGNORECASE) and name.strip().lower() == 'lift handover':
                    row["% Work Done against Target-Till February"] = "0%"
            
            # ===== SET TOTAL ACHIEVED =====
            if is_external or is_common_area:
                row["Total achieved"] = "N/A"
                row["Weighted Delay against Targets"] = ""
            elif activity_data_type == 'percentage':
                row["Total achieved"] = ""
                # Calculate weighted delay for percentage activities
                weighted_delay = last_pct * weightage / 100
                row["Weighted Delay against Targets"] = f"{int(weighted_delay)}%"
            else:
                row["Total achieved"] = f"{int(total_achieved)} {unit_plural}"
                row["Weighted Delay against Targets"] = f"{int(last_pct * weightage / 100)}%"
            
            logger.info(f"\nFinal row:")
            logger.info(f"  Total achieved: {row['Total achieved']}")
            logger.info(f"  Last pct: {last_pct}%")
            logger.info(f"  Weighted delay: {row['Weighted Delay against Targets']}")
            
            data.append(row)
        
        cols = ["Milestone", "Activity", "Target"]
        for m in self.quarter_months:
            cols.append(f"% Work Done against Target-Till {m}")
        cols.extend(["Weightage", "Weighted Delay against Targets"])
        for m in self.quarter_months:
            cols.append(f"Target achieved in {m}")
        cols.extend(["Total achieved", f"Delay Reasons_{self.quarter_months[-1]} {self.quarter_year}"])
        
        return pd.DataFrame(data, columns=cols)
        
    def _sort_sections(self, sections):
        """Sort sections: Structure, Towers, NTA, External"""
        structure_sections = {}
        tower_sections = {}
        nta_sections = {}
        external_sections = {}
        
        for section_name in sections:
            name_lower = section_name.lower()
            
            if 'structure' in name_lower or 'slab' in name_lower:
                structure_sections[section_name] = sections[section_name]
            elif 'nta' in name_lower:
                nta_sections[section_name] = sections[section_name]
            elif 'external development' in name_lower:
                external_sections[section_name] = sections[section_name]
            else:
                tower_sections[section_name] = sections[section_name]
        
        sorted_towers = sorted(
            tower_sections.keys(),
            key=lambda x: int(re.search(r'\d+', x).group()) if re.search(r'\d+', x) else x.lower()
        )
        
        sorted_nta = sorted(nta_sections.keys())
        
        sorted_external = sorted(
            external_sections.keys(),
            key=lambda x: int(re.search(r'\d+', x).group()) if re.search(r'\d+', x) else x.lower()
        )
        
        result = {}
        for section in structure_sections:
            result[section] = structure_sections[section]
        for section in sorted_towers:
            result[section] = tower_sections[section]
        for section in sorted_nta:
            result[section] = nta_sections[section]
        for section in sorted_external:
            result[section] = external_sections[section]
        
        return result
    
    def write_report(self, report_dfs):
        """Write Excel report"""
        filename = f"Eligo_Milestone_Report_{self.current_quarter}_{self.quarter_year}_{datetime.now():%Y%m%d}.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Time Delivery Milestones"
        
        ws.append([f"Eligo Time Delivery Milestones Report - {self.current_quarter} ({', '.join(self.quarter_months)}) {self.quarter_year}"])
        ws.append([f"Report Generated on: {datetime.now().strftime('%d-%m-%Y')}"])
        ws.append([])
        
        yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        grey = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        bold = Font(bold=True)
        thin = Side(style="thin")
        border = Border(top=thin, bottom=thin, left=thin, right=thin)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        sorted_report_dfs = self._sort_sections(report_dfs)
        
        for section_name in sorted_report_dfs.keys():
            df = sorted_report_dfs[section_name]
            
            ws.append([f"{section_name} Progress Against Milestones"])
            title_row = ws.max_row
            
            if not df.empty:
                num_cols = len(df.columns)
            else:
                num_cols = 11
            
            ws.merge_cells(f'A{title_row}:{get_column_letter(num_cols)}{title_row}')
            ws[f'A{title_row}'].font = bold
            ws[f'A{title_row}'].fill = grey
            ws[f'A{title_row}'].alignment = center
            
            if df.empty:
                header_cols = ["Milestone", "Activity", "Target"]
                for m in self.quarter_months:
                    header_cols.append(f"% Work Done against Target-Till {m}")
                header_cols.extend(["Weightage", "Weighted Delay against Targets"])
                for m in self.quarter_months:
                    header_cols.append(f"Target achieved in {m}")
                header_cols.extend(["Total achieved", f"Delay Reasons_{self.quarter_months[-1]} {self.quarter_year}"])
                
                ws.append(header_cols)
                header_row = ws.max_row
                
                for col_idx in range(1, len(header_cols) + 1):
                    ws.cell(header_row, col_idx).font = bold
                    ws.cell(header_row, col_idx).fill = grey
                    ws.cell(header_row, col_idx).border = border
                    ws.cell(header_row, col_idx).alignment = center
                
                logger.warning(f"No activities found for {section_name}")
            else:
                for r in dataframe_to_rows(df, index=False, header=True):
                    ws.append(r)
                
                header_row = title_row + 1
                for col_idx in range(1, len(df.columns) + 1):
                    ws.cell(header_row, col_idx).font = bold
                    ws.cell(header_row, col_idx).fill = grey
                    ws.cell(header_row, col_idx).border = border
                
                for row in ws.iter_rows(title_row + 2, ws.max_row):
                    for col_idx, cell in enumerate(row, 1):
                        cell.border = border
                        cell.alignment = center if cell.column > 2 else left
                
                # Calculate and display total delay
                try:
                    total_delay = sum(
                        float(str(v).strip().rstrip('%')) 
                        for v in df["Weighted Delay against Targets"] 
                        if v and str(v).strip() != ""
                    )
                except (ValueError, AttributeError):
                    total_delay = 0
                
                ws.append(["Total"])
                total_row = ws.max_row
                
                # Find the column index for "Weighted Delay against Targets"
                weighted_delay_col_idx = None
                for col_idx, col_name in enumerate(df.columns, 1):
                    if col_name == "Weighted Delay against Targets":
                        weighted_delay_col_idx = col_idx
                        break
                
                # Write total delay value in the correct column
                if weighted_delay_col_idx:
                    ws.cell(total_row, weighted_delay_col_idx).value = f"{round(total_delay, 2)}%"
                
                # Format total row
                for cell in ws[total_row]:
                    cell.font = bold
                    cell.fill = yellow
                    cell.border = border
            
            ws.append([])
        
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 50)
        
        wb.save(filename)
        logger.info(f"\nReport saved: {filename}")
        
        logger.info(f"\n=== GREEN DATES SUMMARY ===")
        if self.structure_green_dates:
            for month, dates in self.structure_green_dates.items():
                dates_str = ", ".join([d.strftime('%d-%m-%Y') for d in dates])
                logger.info(f"{month}: {dates_str}")
        else:
            logger.info("No green dates extracted")
    
    def generate(self):
        """Generate complete report"""
        logger.info("="*80)
        logger.info("=== ELIGO QUARTERLY REPORT GENERATOR ===")
        logger.info("="*80)
        
        if not self.get_latest_kra():
            return False
        
        if not self.parse_kra():
            return False
        
        self.find_trackers()
        report_dfs = self.build_report_data()
        
        if not report_dfs:
            logger.error("No data generated")
            return False
        
        self.write_report(report_dfs)
        
        logger.info("\n" + "="*80)
        logger.info("REPORT GENERATION COMPLETE")
        logger.info(f"Quarter: {self.current_quarter} {self.quarter_year}")
        logger.info(f"Sections: {len(report_dfs)}")
        logger.info(f"Green Dates Extracted: {len(self.structure_green_dates)} months")
        logger.info("="*80)
        
        return True

# ============================================================================
# MAIN
# ============================================================================
def main():
    generator = EligoReportGenerator()
    success = generator.generate()
    return 0 if success else 1

if __name__ == "__main__":
    exit(main())
    
