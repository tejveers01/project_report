import os
import logging
from io import BytesIO
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
import ibm_boto3
from ibm_botocore.client import Config
import re
from typing import Optional, Tuple, List, Dict, Any

# ======================= CONFIGURATION =======================
load_dotenv()
logging.basicConfig(level=logging.DEBUG, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger(__name__)

# Cloud Storage Configuration
COS_API_KEY = os.getenv("COS_API_KEY")
COS_CRN = os.getenv("COS_SERVICE_INSTANCE_CRN")
COS_ENDPOINT = os.getenv("COS_ENDPOINT")
BUCKET = os.getenv("COS_BUCKET_NAME")
KRA_FOLDER = os.getenv("KRA_FOLDER", "")
EDEN_TRACKER_FOLDER = os.getenv("EDEN_TRACKER_FOLDER", "Eden/")

# ======================= HARDCODED COLUMN MAPPINGS =======================
# KRA Sheet Columns (1-indexed for openpyxl)
KRA_TOWER_COL = 1  # Column A: Tower name

KRA_SEP_ACTIVITY_COL = 2   # Column B: September activity
KRA_SEP_TARGET_COL = 3     # Column C: September % target

KRA_OCT_ACTIVITY_COL = 4   # Column D: October activity
KRA_OCT_TARGET_COL = 5     # Column E: October % target

KRA_NOV_ACTIVITY_COL = 6   # Column F: November activity
KRA_NOV_TARGET_COL = 7     # Column G: November % target

# Tracker Sheet Columns (1-indexed for openpyxl)
TRACKER_TOWER_COL = 1           # Column A: Tower number
TRACKER_ACTIVITY_NO_COL = 2     # Column B: Activity number
TRACKER_LOOKAHEAD_COL = 3       # Column C: Monthly lookahead ID
TRACKER_TASK_NAME_COL = 4       # Column D: Task name
TRACKER_ACTUAL_START_COL = 5    # Column E: Actual start
TRACKER_ACTUAL_FINISH_COL = 6   # Column F: Actual finish
TRACKER_PCT_COMPLETE_COL = 7    # Column G: % Complete ← THE KEY COLUMN
TRACKER_DURATION_COL = 8        # Column H: Duration

# Quarterly structure
QUARTERS = {
    "Q1": ["June", "July", "August"],
    "Q2": ["September", "October", "November"],
    "Q3": ["December", "January", "February"],
    "Q4": ["March", "April", "May"]
}

# Month to tracker month mapping
MONTH_TO_TRACKER_MAPPING = {
    "June": 7, "July": 8, "August": 9,
    "September": 10, "October": 11, "November": 12,
    "December": 1, "January": 2, "February": 3,
    "March": 4, "April": 5, "May": 6
}

# Validate environment variables
required_vars = {
    'COS_API_KEY': COS_API_KEY,
    'COS_SERVICE_INSTANCE_CRN': COS_CRN,
    'COS_ENDPOINT': COS_ENDPOINT,
    'COS_BUCKET_NAME': BUCKET
}

missing_vars = [var_name for var_name, var_value in required_vars.items() if not var_value]
if missing_vars:
    error_msg = f"Missing required environment variables: {', '.join(missing_vars)}"
    logger.error(error_msg)
    raise ValueError(error_msg)

# ======================= CLOUD STORAGE HELPERS =======================

def init_cos():
    """Initialize IBM Cloud Object Storage client."""
    return ibm_boto3.client(
        "s3",
        ibm_api_key_id=COS_API_KEY,
        ibm_service_instance_id=COS_CRN,
        config=Config(signature_version="oauth"),
        endpoint_url=COS_ENDPOINT
    )

def download_file_bytes(cos, key: str) -> bytes:
    """Download file from cloud storage as bytes."""
    return cos.get_object(Bucket=BUCKET, Key=key)["Body"].read()

# ======================= FILE DISCOVERY =======================

def find_latest_kra_file(cos_client, bucket_name: str, folder_prefix: str = "") -> Optional[Tuple[str, List[str], int]]:
    """Find the latest KRA Milestones file."""
    logger.info(f"\n{'='*70}")
    logger.info(f"SEARCHING FOR LATEST KRA FILE")
    logger.info(f"{'='*70}")
    
    try:
        response = cos_client.list_objects_v2(Bucket=bucket_name, Prefix=folder_prefix)
        
        if 'Contents' not in response:
            logger.error(f"No files found in folder '{folder_prefix}'")
            return None
        
        kra_files = []
        
        for obj in response['Contents']:
            file_key = obj['Key']
            filename = os.path.basename(file_key)
            filename_lower = filename.lower()
            
            if file_key.endswith('/'):
                continue
            
            is_kra = 'kra' in filename_lower and 'milestone' in filename_lower
            is_excel = filename_lower.endswith(('.xlsx', '.xls'))
            
            if is_kra and is_excel:
                months_pattern = r'(January|February|March|April|May|June|July|August|September|October|November|December)'
                found_months = re.findall(months_pattern, filename, re.IGNORECASE)
                found_months = [m.capitalize() for m in found_months]
                
                year_match = re.search(r'(\d{4})', filename)
                year = int(year_match.group(1)) if year_match else datetime.now().year
                
                kra_files.append({
                    'key': file_key,
                    'filename': filename,
                    'months': found_months,
                    'year': year,
                    'last_modified': obj['LastModified']
                })
                
                logger.info(f"Found: {filename}")
        
        if not kra_files:
            logger.error("No KRA Milestone files found")
            return None
        
        kra_files.sort(key=lambda f: f['last_modified'], reverse=True)
        latest = kra_files[0]
        
        logger.info(f"\nSelected: {latest['filename']}")
        logger.info(f"Months: {', '.join(latest['months'])}")
        logger.info(f"Year: {latest['year']}")
        
        return latest['key'], latest['months'], latest['year']
        
    except Exception as e:
        logger.error(f"Error searching for KRA file: {str(e)}")
        raise


def calculate_tracker_year(report_month: str, kra_year: int) -> int:
    """
    Calculate correct year for tracker file based on month shift logic.
    
    Month Shift: Report Month → Tracker Month
    September → October, October → November, November → December,
    December → January, January → February, February → March
    """
    tracker_month_num = MONTH_TO_TRACKER_MAPPING.get(report_month)
    report_month_num = MONTH_TO_TRACKER_MAPPING.get(report_month)
    
    if not tracker_month_num:
        return kra_year
    
    # If tracker month is earlier in the year than report month, it's next year
    if tracker_month_num < report_month_num:
        return kra_year + 1
    
    return kra_year

def find_tracker_for_month(cos_client, bucket_name: str, target_month: int, target_year: int,
                          folder_prefix: str = "Eden/") -> Optional[str]:
    """
    Find tracker file for SPECIFIC month and year.
    If multiple exist for that month, use the latest one.
    If none exist, return None (so report shows blank for that month).
    """
    logger.info(f"  Searching for tracker: Month {target_month:02d}/{target_year}")
    
    try:
        response = cos_client.list_objects_v2(Bucket=bucket_name, Prefix=folder_prefix)
        
        if 'Contents' not in response:
            logger.info(f"    ✗ No files in folder")
            return None
        
        matching_trackers = []
        
        for obj in response['Contents']:
            file_key = obj['Key']
            filename = os.path.basename(file_key)
            filename_lower = filename.lower()
            
            if file_key.endswith('/'):
                continue
            
            is_tracker = any(pattern in filename_lower for pattern in 
                           ['structure work tracker', 'tracker', 'structure tracker'])
            is_excel = filename_lower.endswith(('.xlsx', '.xls'))
            
            if is_tracker and is_excel:
                # Extract date from filename: (dd-mm-yyyy)
                date_pattern = r'\((\d{1,2})-(\d{1,2})-(\d{2,4})\)'
                date_match = re.search(date_pattern, filename)
                
                if date_match:
                    day, month, year = date_match.groups()
                    file_month = int(month)
                    file_year = int(year)
                    
                    # Handle 2-digit years
                    if file_year < 100:
                        file_year += 2000
                    
                    logger.debug(f"    Found: {filename} → {file_month:02d}/{file_year}")
                    
                    # Check if this file matches the month and year we're looking for
                    if file_month == target_month and file_year == target_year:
                        matching_trackers.append({
                            'key': file_key,
                            'filename': filename,
                            'day': int(day),
                            'date': obj['LastModified']
                        })
                        logger.debug(f"      ✓ MATCH!")
        
        if not matching_trackers:
            logger.info(f"    ✗ No tracker found for month {target_month:02d}/{target_year}")
            return None
        
        # If multiple trackers for this month, use the latest (highest day number)
        matching_trackers.sort(key=lambda t: t['day'], reverse=True)
        latest_tracker = matching_trackers[0]['key']
        
        logger.info(f"    ✓ Found: {os.path.basename(latest_tracker)}")
        return latest_tracker
        
    except Exception as e:
        logger.error(f"Error searching for tracker: {str(e)}")
        return None


# ======================= KRA DATA EXTRACTION =======================

def find_project_sheet(workbook, project_name: str):
    """Find sheet containing project name."""
    for sheet_name in workbook.sheetnames:
        if project_name.upper() in sheet_name.upper():
            return workbook[sheet_name]
    return None

class ActivityTarget:
    """Represents a target activity from KRA."""
    
    def __init__(self, tower: str, activity_text: str, target_pct: float, month: str):
        self.tower = tower
        self.activity_text = activity_text  # Full text as it appears
        self.target_pct = target_pct
        self.month = month
        self.actual_pct = 0.0
        self.status = ""
    
    def __repr__(self):
        return f"{self.tower} | {self.month} | {self.activity_text} ({self.target_pct}%)"

def get_kra_column_mapping(quarter_months: List[str]) -> Dict[str, Tuple[int, int]]:
    """Create dynamic column mapping based on months in KRA file."""
    mapping = {}
    col_pairs = [(2, 3), (4, 5), (6, 7)]
    
    for idx, month in enumerate(quarter_months[:3]):
        activity_col, target_col = col_pairs[idx]
        mapping[month] = (activity_col, target_col)
        logger.info(f"Column mapping: {month} → Activity Col {activity_col}, Target Col {target_col}")
    
    return mapping

def parse_kra_targets_dynamic(worksheet, quarter_months: List[str]) -> Dict[str, List[ActivityTarget]]:
    """
    Parse KRA targets with HIERARCHICAL activity structure.
    
    KRA shows multi-line hierarchy:
    - Main row: Tower name + first activity level
    - Sub-rows: Child activities (indented/under same tower)
    
    Returns activities as newline-separated hierarchy for tracker matching.
    Example: "Upper Basement\nColumn/Shear Wall\nChecking & Casting Work"
    """
    logger.info("\n" + "="*70)
    logger.info("PARSING KRA TARGETS (Hierarchical Activities)")
    logger.info("="*70)
    
    # Get dynamic column mapping
    col_mapping = get_kra_column_mapping(quarter_months)
    
    tower_targets = {}
    
    # Find header row
    header_row = None
    for row_idx in range(1, min(10, worksheet.max_row + 1)):
        cell_value = worksheet.cell(row_idx, 1).value
        if cell_value and 'tower' in str(cell_value).lower():
            header_row = row_idx
            logger.info(f"Found header row at row {header_row}")
            break
    
    if not header_row:
        logger.error("Could not find header row")
        return tower_targets
    
    # Parse data rows - build hierarchy
    data_start = header_row + 1
    current_tower = None
    activity_hierarchy = {}  # Track multi-line activities per tower per month
    
    for row_idx in range(data_start, worksheet.max_row + 1):
        tower_cell = worksheet.cell(row_idx, 1).value
        
        # NEW TOWER
        if tower_cell and not pd.isna(tower_cell):
            tower_name = str(tower_cell).strip()
            
            # Skip non-tower rows
            if any(x in tower_name.lower() for x in ['activity', 'target', 'milestone header']):
                current_tower = None
                activity_hierarchy = {}
                continue
            
            current_tower = tower_name
            activity_hierarchy = {}  # Reset hierarchy for new tower
            
            if current_tower not in tower_targets:
                tower_targets[current_tower] = []
            
            logger.info(f"\nProcessing: {current_tower}")
        
        # EXTRACT HIERARCHICAL ACTIVITIES FOR CURRENT TOWER
        if current_tower:
            for month in quarter_months:
                if month not in col_mapping:
                    continue
                
                activity_col, target_col = col_mapping[month]
                
                activity = worksheet.cell(row_idx, activity_col).value
                target = worksheet.cell(row_idx, target_col).value
                
                # Build hierarchy: add this row's activity to the chain
                if activity and not pd.isna(activity):
                    activity_text = str(activity).strip()
                    
                    # Initialize month hierarchy if not exists
                    if month not in activity_hierarchy:
                        activity_hierarchy[month] = []
                    
                    # Add to hierarchy
                    activity_hierarchy[month].append(activity_text)
                
                # When we hit a TARGET, save the complete hierarchy up to this point
                if target and not pd.isna(target):
                    try:
                        target_pct = float(target) * 100 if isinstance(target, (int, float)) else 0
                        
                        if target_pct > 0 and month in activity_hierarchy and activity_hierarchy[month]:
                            # Build full hierarchy path
                            full_activity = "\n".join(activity_hierarchy[month])
                            
                            target_obj = ActivityTarget(current_tower, full_activity, target_pct, month)
                            tower_targets[current_tower].append(target_obj)
                            
                            logger.info(f"  {month}: {full_activity[:60].replace(chr(10), ' → ')} → {target_pct:.0f}%")
                            
                            # Reset hierarchy for next target in this month
                            activity_hierarchy[month] = []
                    except (ValueError, TypeError):
                        pass
    
    logger.info(f"\n✓ Extracted targets for {len(tower_targets)} towers")
    for tower in sorted(tower_targets.keys()):
        targets = tower_targets[tower]
        if targets:
            logger.info(f"  {tower}: {len(targets)} target(s)")
    
    return tower_targets


def _parse_month_targets(worksheet, row_idx: int, tower_name: str, tower_targets: Dict, 
                         month: str, activity_col: int, target_col: int):
    """Helper function to parse targets for a specific month."""
    activity = worksheet.cell(row_idx, activity_col).value
    target = worksheet.cell(row_idx, target_col).value
    
    if target and isinstance(target, (int, float)) and target > 0:
        activity_text = str(activity).strip() if activity else "Activity"
        target_obj = ActivityTarget(tower_name, activity_text, float(target) * 100, month)
        tower_targets[tower_name].append(target_obj)
        logger.info(f"  {month}: {activity_text} → {target*100}%")


def _parse_sub_activity(worksheet, row_idx: int, tower_name: str, tower_targets: Dict,
                       month: str, activity_col: int, target_col: int):
    """Helper function to parse sub-activities (multi-line activities)."""
    activity = worksheet.cell(row_idx, activity_col).value
    target = worksheet.cell(row_idx, target_col).value
    
    if target and isinstance(target, (int, float)) and target > 0:
        # Build hierarchical activity text
        activity_parts = []
        current_activity = str(activity).strip() if activity else ""
        
        for back_row in range(max(row_idx - 3, 5), row_idx + 1):
            cell_val = worksheet.cell(back_row, activity_col).value
            if cell_val and str(cell_val).strip():
                cell_str = str(cell_val).strip()
                # Don't include section headers in activity text
                if cell_str not in [tower_name, "NTA Finishing Work Milestone", "External Development Work Milestone"]:
                    # Skip if this is a duplicate of the current row's activity (avoid double-counting)
                    if back_row < row_idx and cell_str == current_activity:
                        continue
                    activity_parts.append(cell_str)
        
        activity_text = "\n".join(activity_parts) if activity_parts else current_activity
        target_obj = ActivityTarget(tower_name, activity_text, float(target) * 100, month)
        tower_targets[tower_name].append(target_obj)
        logger.info(f"  {month} (sub): {activity_text[:50]}... → {target*100}%")

# ======================= TRACKER DATA EXTRACTION =======================

def normalize_text(text: str) -> str:
    """Normalize text for matching."""
    if not text:
        return ""
    # Convert to lowercase, remove extra spaces, remove special chars
    text = re.sub(r'\s+', ' ', str(text).lower().strip())
    text = re.sub(r'[^\w\s]', ' ', text)
    return ' '.join(text.split())

def find_activity_in_tracker(tracker_wb, tower_name: str, activity_text: str, month: str = None) -> Optional[float]:
    """
    Find matching activity in tracker and return % complete.
    Enhanced to handle all milestone types:
    1. Regular towers/NTAs
    2. Tower Finishing Work
    3. NTA Finishing Work Milestone (header) - no tracker lookup
    4. Individual NTA Finishing Work (NTA 01, NTA 02, etc.)
    5. External Development Work - no tracker lookup
    
    """
    
    # Special handling for sections without tracker sheets
    if tower_name in ["NTA Finishing Work Milestone", "External Development Work"]:
        logger.debug(f"    {tower_name} - section header, no individual tracker sheet")
        return None
    
    # Extract base tower name for sheet matching
    base_tower = tower_name
    
    # Handle different milestone types
    if "Finishing Work" in tower_name:
        # "Tower 7 Finishing Work Milestone" -> "Tower 7"
        # "NTA 01 Finishing Work" -> "NTA 01"
        base_tower = tower_name.replace("Finishing Work Milestone", "").replace("Finishing Work", "").strip()
    
    # Find tower sheet
    tower_sheet = None
    sheet_search_terms = []
    
    if "Tower" in base_tower:
        # Extract number: "Tower 7" -> "7"
        tower_num = base_tower.replace("Tower", "").strip()
        # Be specific: only match sheets that have "Tower" in them
        for sheet_name in tracker_wb.sheetnames:
            if "tower" in sheet_name.lower() and tower_num in sheet_name:
                tower_sheet = tracker_wb[sheet_name]
                break
    elif "NTA" in base_tower:
        # NTAs should look in "Non Tower Area" sheet
        for sheet_name in tracker_wb.sheetnames:
            if "non tower" in sheet_name.lower() or "nta" in sheet_name.lower():
                tower_sheet = tracker_wb[sheet_name]
                break
    else:
        logger.debug(f"    Cannot extract tower identifier from: {tower_name}")
        return None
    
    if not tower_sheet:
        logger.debug(f"    Sheet not found for {tower_name}")
        return None
    
    # Define row ranges for each NTA tower to constrain search
    NTA_ROW_RANGES = {
        "NTA 01": (6, 33),
        "NTA 02": (35, 62),
        "NTA 03": (64, 91),
        "NTA 04": (122, 149),
        "NTA 05": (93, 120),
        "NTA 06": (151, 178),
        "NTA 07": (180, 207),
        "NTA 08": (209, 236),
        "NTA 09": (238, 266),
        "NTA 10": (268, 296)
    }
    
    # Determine row range based on tower
    row_start = 3  # Default start row
    row_end = tower_sheet.max_row + 1  # Default end row
    
    # Check if this is an NTA tower and apply row constraints
    if "NTA" in base_tower:
        # Normalize the NTA identifier (handle "NTA 01", "NTA 1", "NTA01", etc.)
        nta_num = base_tower.replace("NTA", "").strip()
        # Pad single digit with zero
        if len(nta_num) == 1:
            nta_num = "0" + nta_num
        nta_key = f"NTA {nta_num}"
        
        if nta_key in NTA_ROW_RANGES:
            row_start, row_end = NTA_ROW_RANGES[nta_key]
            row_end += 1  # Make it inclusive
            logger.debug(f"    Using NTA row range: {row_start}-{row_end-1}")
    
    # Split activity text into hierarchy levels - handle both newline and comma separation
    activity_lines = []
    if '\n' in activity_text:
        # Newline-separated hierarchy (sub-activities)
        activity_lines = [line.strip() for line in activity_text.split('\n') if line.strip()]
    elif ',' in activity_text:
        # Comma-separated hierarchy (single-line activities)
        activity_lines = [line.strip() for line in activity_text.split(',') if line.strip()]
    else:
        # Single term
        activity_lines = [activity_text.strip()] if activity_text.strip() else []
    
    if not activity_lines:
        return None
    
    logger.debug(f"    Searching in {tower_sheet.title}")
    logger.debug(f"    Hierarchy: {' → '.join(activity_lines)}")
    
    # Strategy: Find the PARENT level first, then find the CHILD within the next 10 rows
    parent_term = normalize_text(activity_lines[0])
    
    # Define conflicting terms for the parent
    conflicting_terms = []
    if 'upper' in parent_term:
        conflicting_terms.append('lower')
    elif 'lower' in parent_term:
        conflicting_terms.append('upper')
    elif 'ground' in parent_term:
        conflicting_terms.extend(['1st', '2nd', '3rd', '4th'])
    elif '1st' in parent_term:
        conflicting_terms.extend(['ground', '2nd', '3rd', '4th'])
    elif '2nd' in parent_term:
        conflicting_terms.extend(['ground', '1st', '3rd', '4th'])
    elif '3rd' in parent_term:
        conflicting_terms.extend(['ground', '1st', '2nd', '4th'])
    
    # Special handling for finishing work keywords
    if any(term in parent_term for term in ['finishing', 'paint', 'plastering', 'false ceiling', 'flooring', 'tiles', 'fixtures']):
        conflicting_terms.extend(['structure', 'rcc', 'concrete', 'casting', 'shuttering', 'reinforcement'])
    
    logger.debug(f"    Parent term: '{parent_term}'")
    logger.debug(f"    Conflicting terms: {conflicting_terms}")
    
    # STEP 1: Find the parent row (within row range constraints)
    parent_row = None
    for row_idx in range(row_start, row_end):
        task_name = tower_sheet.cell(row_idx, TRACKER_TASK_NAME_COL).value
        
        if not task_name:
            continue
        
        task_normalized = normalize_text(task_name)
        
        # Check if this row matches the parent
        if parent_term in task_normalized:
            # Make sure it doesn't contain conflicting terms
            has_conflict = any(conflict in task_normalized for conflict in conflicting_terms)
            
            if not has_conflict:
                parent_row = row_idx
                logger.debug(f"    Found parent at row {row_idx}: {task_name.strip()}")
                break
    
    if not parent_row:
        logger.debug(f"    Parent '{parent_term}' not found")
        return None
    
    # STEP 2: If we have child levels, search within next 10 rows
    if len(activity_lines) > 1:
        child_terms = [normalize_text(line) for line in activity_lines[1:]]
        logger.debug(f"    Child terms: {child_terms}")
        
        best_match = None
        best_match_score = 0
        best_match_row = None
        
        # Search within next 25 rows after parent (but not beyond row_end)
        for row_idx in range(parent_row + 1, min(parent_row + 26, row_end)):
            task_name = tower_sheet.cell(row_idx, TRACKER_TASK_NAME_COL).value
            
            if not task_name:
                continue
            
            task_normalized = normalize_text(task_name)
            
            # Calculate match score - prioritize matching deeper (later) terms
            # Strongly prefer exact matches over partial matches
            match_count = 0
            for idx, term in enumerate(child_terms):
                if term in task_normalized:
                    # Base score for matching this term (later terms get higher scores)
                    term_score = (idx + 1)
                    
                    # Strong bonus for exact match (term equals the whole task name)
                    if term == task_normalized:
                        term_score += 100  # Very high bonus for exact match
                    # Medium bonus if term is a major portion (>70%) of task name
                    elif len(term) >= len(task_normalized) * 0.7:
                        term_score += 20
                    # Small penalty if task name has many extra words beyond our term
                    elif len(task_normalized) > len(term) * 1.5:
                        term_score -= 5  # Penalize if tracker has significantly more words
                    
                    match_count += term_score
            
            # Only update if we have a BETTER match
            if match_count > best_match_score:
                pct_value = tower_sheet.cell(row_idx, TRACKER_PCT_COMPLETE_COL).value
                
                if pct_value is not None:
                    try:
                        if isinstance(pct_value, (int, float)):
                            pct_complete = float(pct_value) * 100
                        else:
                            pct_complete = float(str(pct_value).replace('%', ''))
                        
                        if 0 <= pct_complete <= 100:
                            best_match_score = match_count
                            best_match = pct_complete
                            best_match_row = row_idx
                            logger.debug(f"    Child match at row {row_idx}: {task_name.strip()[:50]} = {pct_complete:.1f}% (score: {match_count})")
                    except (ValueError, TypeError):
                        pass
        
        if best_match is not None:
            logger.debug(f"    ✓ SELECTED: row {best_match_row}, {best_match:.1f}%")
            return best_match
        else:
            logger.debug(f"    Child terms not found within 10 rows of parent")
            return None
    
    else:
        # Single level activity - return parent row's % Complete
        pct_value = tower_sheet.cell(parent_row, TRACKER_PCT_COMPLETE_COL).value
        
        if pct_value is not None:
            try:
                if isinstance(pct_value, (int, float)):
                    pct_complete = float(pct_value) * 100
                else:
                    pct_complete = float(str(pct_value).replace('%', ''))
                
                if 0 <= pct_complete <= 100:
                    logger.debug(f"    ✓ SELECTED parent row: {pct_complete:.1f}%")
                    return pct_complete
            except (ValueError, TypeError):
                pass
        
        return None
# ======================= REPORT GENERATION =======================

def sort_towers(tower_name: str) -> tuple:
    """Custom sort key ensuring proper order of all milestone types"""
    tower_lower = tower_name.lower()
    
    # Priority 0: Regular Towers (Structure Work)
    if tower_name.startswith('Tower') and 'finishing' not in tower_lower:
        match = re.search(r'Tower\s*(\d+)', tower_name)
        if match:
            return (0, int(match.group(1)), tower_name)
        return (0, 999, tower_name)
    
    # Priority 1: Regular NTAs (Structure Work) - Must NOT be in finishing section
    elif tower_name.startswith('NTA') and 'finishing' not in tower_lower and 'work' not in tower_lower:
        # This catches only "NTA 01", "NTA 02" that are structure work
        match = re.search(r'NTA\s*(\d+)', tower_name)
        if match:
            return (1, int(match.group(1)), tower_name)
        return (1, 999, tower_name)
    
    # Priority 2: Tower Finishing Work
    elif 'tower' in tower_lower and 'finishing' in tower_lower:
        match = re.search(r'Tower\s*(\d+)', tower_name, re.IGNORECASE)
        if match:
            return (2, int(match.group(1)), tower_name)
        return (2, 999, tower_name)
    
    # Priority 3: NTA Finishing Work Milestone Section
    # Sub-priority 0: The header "NTA Finishing Work Milestone:"
    # Sub-priority 1-99: Individual NTAs "NTA 01 Finishing Work", "NTA 02 Finishing Work"...
    elif 'nta' in tower_lower and 'finishing' in tower_lower:
        # The header comes first
        if tower_name == "NTA Finishing Work Milestone":
            return (3, 0, tower_name)
        # Individual NTA Finishing Work entries
        # Match pattern: "NTA 01 Finishing Work", "NTA 02 Finishing Work", etc.
        match = re.search(r'NTA\s*(\d+)', tower_name, re.IGNORECASE)
        if match:
            nta_num = int(match.group(1))
            return (3, nta_num, tower_name)
        return (3, 999, tower_name)
    
    # Priority 4: External Development Work
    elif 'external' in tower_lower or 'development' in tower_lower:
        return (4, 0, tower_name)
    
    # Priority 5: Others
    else:
        return (5, 999, tower_name)

def generate_report(tower_targets: Dict[str, List[ActivityTarget]], 
                   tracker_workbooks: Dict[str, Any], months: List[str], year: int) -> pd.DataFrame:
    """
    Generate milestone report DataFrame.
    SPECIAL: "NTA Finishing Work Milestone" appears as a section header row with no data
    """
    logger.info("\n" + "="*70)
    logger.info("GENERATING REPORT")
    logger.info("="*70)
    
    report_rows = []
    
    # Sort towers
    sorted_tower_names = sorted(tower_targets.keys(), key=sort_towers)
    
    logger.info(f"\nSorted tower order:")
    for idx, tower in enumerate(sorted_tower_names, 1):
        logger.info(f"  {idx}. {tower} (Priority: {sort_towers(tower)})")
    
    for tower_name in sorted_tower_names:
        # Skip only the invalid "NTA" entry
        if tower_name.strip().upper() == "NTA":
            logger.info(f"\nSkipping: {tower_name} (invalid)")
            continue
        
        # SPECIAL CASE: "NTA Finishing Work Milestone" is a section header only
        if tower_name == "NTA Finishing Work Milestone":
            logger.info(f"\nAdding section header: {tower_name}")
            
            # Create a header row with colon appended to tower name
            header_row = {'Tower': f"{tower_name}:"}  # Add colon here
            
            for month in months:
                header_row[f"Activity- {month} {year}"] = ""
                header_row[f"% Complete- {month}"] = ""
                header_row[f"Status- {month}"] = ""
                header_row[f"Weightage- {month}"] = ""
                header_row[f"Weighted %- {month}"] = ""
            
            header_row[f"Target till {months[-1]}"] = ""
            header_row['Responsible'] = ""
            header_row['Delay Reason'] = ""
            
            report_rows.append(header_row)
            continue
        
        logger.info(f"\nProcessing: {tower_name}")
        
        row_data = {'Tower': tower_name}
        
        # Process each month
        for month in months:
            month_targets = [t for t in tower_targets[tower_name] if t.month == month]
            tracker_wb = tracker_workbooks.get(month)
            
            if not month_targets:
                # No targets for this month
                row_data[f"Activity- {month} {year}"] = ""
                row_data[f"% Complete- {month}"] = ""
                row_data[f"Status- {month}"] = ""
                row_data[f"Weightage- {month}"] = ""
                row_data[f"Weighted %- {month}"] = ""
                continue
            
            # We have targets
            activities_text = "\n".join([t.activity_text for t in month_targets])
            row_data[f"Activity- {month} {year}"] = activities_text
            
            if tracker_wb:
                total_actual = 0
                matched = 0
                
                for target in month_targets:
                    actual_pct = find_activity_in_tracker(tracker_wb, tower_name, target.activity_text, month)
                    
                    if actual_pct is not None:
                        # If actual meets or exceeds target, show 100%
                        if actual_pct >= target.target_pct:
                            target.actual_pct = 100.0
                            target.status = "Achieved"
                            matched += 1
                        else:
                            # Below target - show actual percentage
                            target.actual_pct = actual_pct
                            target.status = "Not Matched"
                        
                        logger.info(f"  {month}: {target.activity_text[:40]} = {target.actual_pct:.0f}%")
                        total_actual += target.actual_pct
                    else:
                        target.status = "Not Found"
                
                avg_actual = total_actual / len(month_targets) if month_targets else 0
                
                if matched == len(month_targets) and matched > 0:
                    status = "Achieved"
                elif matched > 0:
                    status = "Partial"
                else:
                    status = "Not Achieved"
                
                row_data[f"% Complete- {month}"] = f"{avg_actual:.0f}%"
                row_data[f"Status- {month}"] = status
                
                # Weightage is 100 for each month
                weightage = 100
                weighted_pct = (avg_actual / 100) * weightage
                row_data[f"Weightage- {month}"] = weightage
                row_data[f"Weighted %- {month}"] = f"{weighted_pct:.1f}%"
            else:
                row_data[f"% Complete- {month}"] = ""
                row_data[f"Status- {month}"] = ""
                row_data[f"Weightage- {month}"] = ""
                row_data[f"Weighted %- {month}"] = ""
        
        # Summary columns
        last_month = months[-1]
        last_targets = [t for t in tower_targets[tower_name] if t.month == last_month]
        row_data[f"Target till {last_month}"] = "\n".join([t.activity_text for t in last_targets])
        
        row_data['Responsible'] = ""
        row_data['Delay Reason'] = ""
        
        report_rows.append(row_data)
    
    # Add summary row
    summary_row = {'Tower': 'AVERAGE WEIGHTED %'}
    
    for month in months:
        weighted_values = []
        for row in report_rows:
            # Skip the NTA Finishing Work Milestone header row in calculations
            if row['Tower'] == "NTA Finishing Work Milestone:" or row['Tower'] == "NTA Finishing Work Milestone":
                continue
                
            weighted_val = row.get(f"Weighted %- {month}", "")
            if weighted_val and weighted_val != "":
                try:
                    val = float(str(weighted_val).replace('%', ''))
                    weighted_values.append(val)
                except (ValueError, TypeError):
                    pass
        
        if weighted_values:
            avg_weighted = sum(weighted_values) / len(weighted_values)
            summary_row[f"Weighted %- {month}"] = f"{avg_weighted:.1f}%"
        else:
            summary_row[f"Weighted %- {month}"] = ""
        
        summary_row[f"Activity- {month} {year}"] = ""
        summary_row[f"% Complete- {month}"] = ""
        summary_row[f"Status- {month}"] = ""
        summary_row[f"Weightage- {month}"] = ""
    
    summary_row[f"Target till {months[-1]}"] = ""
    summary_row['Responsible'] = ""
    summary_row['Delay Reason'] = ""
    
    report_rows.append(summary_row)
    
    return pd.DataFrame(report_rows)

def format_report(worksheet, dataframe):
    """Apply formatting to report."""
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    summary_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    summary_font = Font(bold=True, size=11)
    
    # Format title
    worksheet.cell(1, 1).font = Font(bold=True, size=14)
    worksheet.cell(2, 1).font = Font(italic=True, size=10)
    
    # Format headers
    for col in range(1, worksheet.max_column + 1):
        cell = worksheet.cell(4, col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Format data
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Last row is the summary row
    summary_row_idx = worksheet.max_row
    
    for row in range(5, worksheet.max_row + 1):
        is_summary_row = (row == summary_row_idx)
        
        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row, col)
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            cell.border = thin_border
            
            # Special formatting for summary row
            if is_summary_row:
                cell.fill = summary_fill
                cell.font = summary_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Column widths
    for col_idx, column in enumerate(dataframe.columns, start=1):
        col_letter = get_column_letter(col_idx)
        if 'Activity' in column or 'Target' in column:
            worksheet.column_dimensions[col_letter].width = 40
        else:
            worksheet.column_dimensions[col_letter].width = 15
    
    worksheet.row_dimensions[4].height = 50
    worksheet.row_dimensions[summary_row_idx].height = 30  # Make summary row slightly taller

# ======================= MAIN =======================

def main():
    """Main execution."""
    try:
        logger.info("\n" + "="*70)
        logger.info("MILESTONE REPORT GENERATOR v2.0")
        logger.info("="*70)
        
        # Step 1: Find KRA
        logger.info("\nSTEP 1: Finding latest KRA file")
        cos = init_cos()
        
        kra_result = find_latest_kra_file(cos, BUCKET, KRA_FOLDER)
        if not kra_result:
            logger.error("Could not find KRA file")
            return
        
        kra_key, quarter_months, kra_year = kra_result
        
        # Step 2: Load KRA
        logger.info("\nSTEP 2: Loading KRA and parsing targets")
        kra_bytes = download_file_bytes(cos, kra_key)
        kra_wb = load_workbook(filename=BytesIO(kra_bytes), data_only=True)
        
        kra_ws = find_project_sheet(kra_wb, "EDEN")
        if not kra_ws:
            logger.error("EDEN sheet not found")
            return
        
        # Use DYNAMIC parser that reads the actual quarter months
        tower_targets = parse_kra_targets_dynamic(kra_ws, quarter_months)
        
        if not tower_targets:
            logger.error("No targets found in KRA")
            return
        
        # Step 3: Load trackers ONLY for months that exist
        logger.info("\nSTEP 3: Loading tracker files (ONLY if they exist)")
        tracker_workbooks = {}
        
        for month in quarter_months:
            tracker_month_num = MONTH_TO_TRACKER_MAPPING.get(month)
            if not tracker_month_num:
                logger.warning(f"  {month}: No mapping found")
                continue
            
            tracker_year = calculate_tracker_year(month, kra_year)
            logger.info(f"\n  {month} {kra_year} requires tracker: {tracker_month_num:02d}/{tracker_year}")
            
            # CRITICAL: Look for SPECIFIC month/year tracker
            tracker_key = find_tracker_for_month(cos, BUCKET, tracker_month_num, tracker_year, EDEN_TRACKER_FOLDER)
            
            if tracker_key:
                logger.info(f"    Loading tracker...")
                tracker_bytes = download_file_bytes(cos, tracker_key)
                tracker_wb = load_workbook(filename=BytesIO(tracker_bytes), data_only=True)
                tracker_workbooks[month] = tracker_wb
                logger.info(f"    ✓ Loaded successfully")
            else:
                logger.warning(f"    ✗ Tracker NOT FOUND - {month} column will be BLANK")
        
        logger.info(f"\n  Summary: {len(tracker_workbooks)}/{len(quarter_months)} trackers loaded")
        
        # Step 4: Generate report
        logger.info("\nSTEP 4: Generating report")
        report_df = generate_report(tower_targets, tracker_workbooks, quarter_months, kra_year)
        
        # Step 5: Save
        logger.info("\nSTEP 5: Saving report")
        output_file = f"Eden_Milestone_Report_{'_'.join(quarter_months)}_{kra_year}.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Progress Report"
        
        ws.append(["Eden- Progress Against Milestones"])
        ws.append([f"Report Generated: {datetime.now().strftime('%B %d, %Y')}"])
        ws.append([])
        
        for r in dataframe_to_rows(report_df, index=False, header=True):
            ws.append(r)
        
        format_report(ws, report_df)
        wb.save(output_file)
        
        logger.info(f"\n{'='*70}")
        logger.info("REPORT COMPLETE")
        logger.info(f"{'='*70}")
        logger.info(f"File: {output_file}")
        logger.info(f"Towers: {len(report_df)}")
        logger.info(f"Months with tracker data: {list(tracker_workbooks.keys())}")
        logger.info(f"Months with BLANK columns: {[m for m in quarter_months if m not in tracker_workbooks]}")
        logger.info(f"{'='*70}\n")
        
    except Exception as e:
        logger.error(f"Error: {str(e)}", exc_info=True)
        raise

if __name__ == "__main__":
    main()
























