import os
import sys
import re
import logging
from io import BytesIO
from datetime import datetime

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import ibm_boto3
from ibm_botocore.client import Config

CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR = os.path.dirname(CURRENT_DIR)
if ROOT_DIR not in sys.path:
    sys.path.insert(0, ROOT_DIR)

from env_loader import load_root_env

# ============================================================================
# SETUP & CONFIGURATION
# ============================================================================
load_root_env()
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger(__name__)

COS_API_KEY = os.getenv("COS_API_KEY")
COS_CRN = os.getenv("COS_SERVICE_INSTANCE_CRN")
COS_ENDPOINT = os.getenv("COS_ENDPOINT")
BUCKET = os.getenv("COS_BUCKET_NAME")

GREEN_HEX = "FF92D050"

QUARTERS = {
    'Q1': ['June', 'July', 'August'],
    'Q2': ['September', 'October', 'November'],
    'Q3': ['December', 'January', 'February'],
    'Q4': ['March', 'April', 'May']
}

MONTH_SHIFT_MAP = {
    'June': 'July', 'July': 'August', 'August': 'September',
    'September': 'October', 'October': 'November', 'November': 'December',
    'December': 'January', 'January': 'February', 'February': 'March',
    'March': 'April', 'April': 'May', 'May': 'June'
}

MONTH_TO_NUM = {
    'January': 1, 'February': 2, 'March': 3, 'April': 4,
    'May': 5, 'June': 6, 'July': 7, 'August': 8,
    'September': 9, 'October': 10, 'November': 11, 'December': 12
}

# ============================================================================
# CLOUD STORAGE
# ============================================================================
def init_cos():
    """Initialize IBM COS client"""
    return ibm_boto3.client(
        "s3",
        ibm_api_key_id=COS_API_KEY,
        ibm_service_instance_id=COS_CRN,
        config=Config(signature_version="oauth"),
        endpoint_url=COS_ENDPOINT,
    )

def list_files(cos, prefix):
    """List files from COS bucket"""
    try:
        response = cos.list_objects_v2(Bucket=BUCKET, Prefix=prefix)
        return [obj['Key'] for obj in response.get('Contents', []) if not obj['Key'].endswith('/')]
    except Exception as e:
        logger.error(f"Error listing files: {e}")
        return []

def download_file(cos, key):
    """Download file from COS bucket"""
    obj = cos.get_object(Bucket=BUCKET, Key=key)
    return obj["Body"].read()

# ============================================================================
# UTILITY FUNCTIONS
# ============================================================================
def extract_unit_from_cell(cell_value):
    """Extract unit from cell value (e.g., 'Flats', 'activity', 'Slab')"""
    if not cell_value:
        return 'Flat'
    
    cell_str = str(cell_value).strip()
    
    # Check for common unit patterns
    if 'activity' in cell_str.lower():
        return 'activity'
    elif 'slab' in cell_str.lower():
        return 'Slab'
    elif 'flat' in cell_str.lower():
        return 'Flat'
    
    # Default to Flat
    return 'Flat'


def extract_date_from_filename(filename):
    """Extract date from filename (DD-MM-YYYY)"""
    pattern = r'\((\d{2}-\d{2}-\d{4})\)'
    match = re.search(pattern, filename)
    if match:
        try:
            return datetime.strptime(match.group(1), '%d-%m-%Y')
        except ValueError:
            return None
    return None

def determine_quarter_from_kra(kra_filename):
    """Determine quarter from KRA filename"""
    pattern = r'KRA Milestones for (\w+) (\w+) (\w+) (\d{4})'
    match = re.search(pattern, kra_filename, re.IGNORECASE)
    
    if match:
        m1, m2, m3, year = match.groups()
        months = [m1, m2, m3]
        
        for q, qmonths in QUARTERS.items():
            if set([m.lower() for m in months]) == set([m.lower() for m in qmonths]):
                return q, qmonths, int(year)
    
    return None, None, None

def extract_number(cell_value):
    """Extract numeric value from cell"""
    if not cell_value or cell_value == "-":
        return 0.0
    match = re.search(r"(\d+\.?\d*)", str(cell_value))
    return float(match.group(1)) if match else 0.0

def get_cell_hex_color(cell):
    """Extract hex color from cell"""
    if cell.fill and cell.fill.start_color:
        color_value = str(cell.fill.start_color.rgb)
        if color_value.startswith('FF') and len(color_value) > 8:
            return color_value[2:]
        return color_value
    return None

def is_cell_green(cell, green_hex=GREEN_HEX):
    """Check if cell has green background"""
    cell_color = get_cell_hex_color(cell)
    if not cell_color:
        return False
    
    cell_color_normalized = cell_color.upper()
    target_normalized = green_hex.upper()
    
    if target_normalized.startswith('FF') and len(target_normalized) == 8:
        target_normalized = target_normalized[2:]
    
    if cell_color_normalized.startswith('FF') and len(cell_color_normalized) == 8:
        cell_color_normalized = cell_color_normalized[2:]
    
    return cell_color_normalized == target_normalized

# ============================================================================
# KRA PARSER
# ============================================================================
class DynamicKRAParser:
    """Parse KRA sheet dynamically with unit detection"""
    
    @staticmethod
    def find_section_headers(sheet):
        """Find section headers"""
        sections = {}
        
        for row_idx in range(1, sheet.max_row + 1):
            cell = sheet.cell(row=row_idx, column=1)
            if not cell.value:
                continue
            
            cell_value = str(cell.value).strip()
            
            if re.search(r'(Tower|Green|Structure|External Development|External Dev)', cell_value, re.IGNORECASE):
                sections[cell_value] = row_idx
                logger.info(f"Found section: '{cell_value}'")
        
        return sections
    
    @staticmethod
    def find_month_columns(sheet, start_row, report_months):
        """Find month columns"""
        month_cols = {}
        
        for row_idx in range(start_row, min(start_row + 3, sheet.max_row + 1)):
            for col_idx in range(1, 10):
                cell = sheet.cell(row=row_idx, column=col_idx)
                if cell.value and str(cell.value).strip() in report_months:
                    month_cols[str(cell.value).strip()] = col_idx
            
            if len(month_cols) >= 2:
                return month_cols, row_idx
        
        return {}, None
    
    @staticmethod
    def extract_activities(sheet, header_row, month_cols, report_months, is_structure=False, is_external=False):
        """Extract activities with their units"""
        activities = []
        
        if is_structure:
            for row_idx in range(header_row + 1, min(header_row + 20, sheet.max_row + 1)):
                cell = sheet.cell(row=row_idx, column=1)
                if cell.value and str(cell.value).strip().lower() == 'total':
                    activities.append({'name': 'Slab Casting', 'row': row_idx, 'unit': 'Slab'})
                    break
        elif is_external:
            for row_idx in range(header_row + 1, min(header_row + 50, sheet.max_row + 1)):
                cell = sheet.cell(row=row_idx, column=1)
                if not cell.value:
                    continue
                
                activity_name = str(cell.value).strip()
                
                if activity_name.lower().startswith('total'):
                    break
                
                if not activity_name:
                    continue
                
                has_data = False
                for col_idx in month_cols.values():
                    val = sheet.cell(row=row_idx, column=col_idx).value
                    if val and extract_number(val) > 0:
                        has_data = True
                        break
                
                if has_data:
                    unit = 'Flat'  # default
                    
                    # Check cells in this row for unit information
                    for col_idx in month_cols.values():
                        cell_val = sheet.cell(row=row_idx, column=col_idx).value
                        detected_unit = extract_unit_from_cell(cell_val)
                        if detected_unit != 'Flat':  # If we find a non-default unit
                            unit = detected_unit
                            break
                    
                    activities.append({'name': activity_name, 'row': row_idx, 'unit': unit})
        else:
            for row_idx in range(header_row + 1, min(header_row + 50, sheet.max_row + 1)):
                cell = sheet.cell(row=row_idx, column=1)
                if not cell.value:
                    continue
                
                activity_name = str(cell.value).strip()
                
                if re.search(r'(Tower|Green|Finishing|Structure|External Development)\s*\d*', activity_name, re.IGNORECASE):
                    if activity_name.lower().startswith('tower') or 'external' in activity_name.lower():
                        break
                    continue
                
                if not activity_name or activity_name.lower().startswith('total'):
                    continue
                
                has_data = False
                for col_idx in month_cols.values():
                    val = sheet.cell(row=row_idx, column=col_idx).value
                    if val and extract_number(val) > 0:
                        has_data = True
                        break
                
                if has_data:
                    unit = 'Flat'  # default
                    
                    # Check cells in this row for unit information
                    for col_idx in month_cols.values():
                        cell_val = sheet.cell(row=row_idx, column=col_idx).value
                        detected_unit = extract_unit_from_cell(cell_val)
                        if detected_unit != 'Flat':  # If we find a non-default unit
                            unit = detected_unit
                            break
                    
                    activities.append({'name': activity_name, 'row': row_idx, 'unit': unit})
        
        return activities
    
    @classmethod
    def parse_kra_dynamic(cls, sheet, report_months):
        """Parse KRA sheet"""
        logger.info("=== PARSING KRA ===\n")
        
        sections = cls.find_section_headers(sheet)
        if not sections:
            logger.error("No sections found")
            return {}
        
        parsed = {}
        
        for section_name, start_row in sorted(sections.items(), key=lambda x: x[1]):
            logger.info(f"Parsing: {section_name}")
            
            month_cols, month_row = cls.find_month_columns(sheet, start_row, report_months)
            
            if not month_cols:
                logger.warning(f"No month columns found")
                continue
            
            is_structure = 'Structure' in section_name or 'Slab' in section_name
            is_external = 'External Development' in section_name
            
            activities = cls.extract_activities(sheet, month_row, month_cols, report_months, is_structure, is_external)
            
            if not activities:
                logger.warning(f"No activities found for {section_name}")
                activities = []
            
            logger.info(f"Found {len(activities)} activities")
            
            for activity in activities:
                logger.info(f"  - {activity['name']}: unit={activity['unit']}")
            
            targets = {}
            for activity in activities:
                targets[activity['name']] = {}
                row_idx = activity['row']
                
                for month, col_idx in month_cols.items():
                    cell_value = sheet.cell(row=row_idx, column=col_idx).value
                    val = extract_number(cell_value)
                    
                    if 0 <= val <= 1 and is_external:
                        val = val * 100
                    
                    targets[activity['name']][month] = val
            
            parsed[section_name] = {
                'activities': activities,
                'month_cols': month_cols,
                'targets': targets,
                'is_structure': is_structure,
                'is_external': is_external
            }
        
        return parsed

# ============================================================================
# TRACKER PARSERS
# ============================================================================
class TowerTrackerParser:
    """Parse tower tracker files"""

    @staticmethod
    def normalize_activity(text):
        """Normalize activity text for tighter matching."""
        if text is None:
            return ""
        s = str(text).strip().lower()
        # Keep alphanumerics and spaces only
        s = re.sub(r"[^a-z0-9 ]+", " ", s)
        # Collapse whitespace
        s = re.sub(r"\s+", " ", s).strip()
        return s
    
    @staticmethod
    def find_headers(ws):
        """Find Activity and Actual Finish columns"""
        header_row = None
        activity_col = None
        finish_col = None
        
        for row_idx in range(1, min(10, ws.max_row + 1)):
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    cell_val = str(cell.value).strip().lower()
                    if 'activity' in cell_val or 'task' in cell_val:
                        activity_col = col_idx
                    if 'actual' in cell_val and 'finish' in cell_val:
                        finish_col = col_idx
            
            if activity_col and finish_col:
                header_row = row_idx
                break
        
        return header_row, activity_col, finish_col
    
    @staticmethod
    def parse_date(finish_cell):
        """Parse date from cell"""
        if isinstance(finish_cell, datetime):
            return finish_cell
        elif isinstance(finish_cell, str):
            for fmt in ["%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y", "%d/%m/%Y"]:
                try:
                    return datetime.strptime(finish_cell.strip(), fmt)
                except ValueError:
                    continue
        return None
    
    @classmethod
    def parse(cls, tracker_bytes, activities, month_num):
        """Parse tower tracker"""
        logger.info(f"=== PARSING TOWER TRACKER ===")
        wb = load_workbook(filename=tracker_bytes, data_only=True)
        
        tracker_counts = {act['name']: 0 for act in activities}
        found_counts = {act['name']: False for act in activities}
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            header_row, activity_col, finish_col = cls.find_headers(ws)
            if not header_row or not finish_col:
                continue
            
            for activity in activities:
                activity_name = activity['name']
                
                if found_counts[activity_name]:
                    continue
                
                sheet_count = 0
                
                for row_idx in range(header_row + 1, ws.max_row + 1):
                    activity_cell = ws.cell(row=row_idx, column=activity_col).value
                    finish_cell = ws.cell(row=row_idx, column=finish_col).value
                    
                    if not activity_cell:
                        continue
                    
                    # Tight match: normalized activity name must equal normalized cell text
                    if cls.normalize_activity(activity_name) != cls.normalize_activity(activity_cell):
                        continue
                    
                    finish_date = cls.parse_date(finish_cell)
                    if finish_date and finish_date.month == month_num:
                        sheet_count += 1
                
                if sheet_count > 0:
                    tracker_counts[activity_name] = sheet_count
                    found_counts[activity_name] = True
                    logger.info(f"{activity_name}: {sheet_count} completed")
        
        logger.info(f"Tower tracker parsed")
        return tracker_counts, 'count'


class ExternalDevelopmentParser:
    """Parse external development tracker files"""
    
    @staticmethod
    def find_headers(ws):
        """Find Activity and % Complete columns"""
        header_row = None
        activity_col = None
        status_col = None
        
        for row_idx in range(1, min(6, ws.max_row + 1)):
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if not cell.value:
                    continue
                    
                cell_val = str(cell.value).strip().lower()
                
                if 'activity' in cell_val:
                    activity_col = col_idx
                    
                if '% complete' in cell_val or cell_val == '% complete':
                    status_col = col_idx
            
            if activity_col and status_col:
                header_row = row_idx
                logger.info(f"Found headers at row {row_idx}: Activity col={activity_col}, Status col={status_col}")
                break
        
        return header_row, activity_col, status_col
    
    @staticmethod
    def parse_percentage(status_value):
        """Extract percentage completion from status"""
        if status_value is None:
            return 0.0
        
        try:
            val = float(status_value)
            if 0 < val <= 1:
                val = val * 100
            val = max(0, min(val, 100))
            return val
        except (ValueError, TypeError):
            pass
        
        status_str = str(status_value).strip()
        
        if not status_str or status_str == "-":
            return 0.0
        
        status_str = status_str.replace("%", "").strip()
        
        match = re.search(r"(\d+(?:\.\d+)?)", status_str)
        if match:
            try:
                pct = float(match.group(1))
                if 0 < pct <= 1:
                    pct = pct * 100
                pct = max(0, min(pct, 100))
                return pct
            except ValueError:
                return 0.0
        
        return 0.0
    
    @classmethod
    def parse_with_targets(cls, tracker_bytes, activities, kra_targets):
        """Parse tracker and extract % complete"""
        logger.info(f"=== PARSING EXTERNAL DEVELOPMENT TRACKER WITH TARGETS ===")
        
        wb = load_workbook(filename=tracker_bytes, data_only=True)
        tracker_data = {}
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            logger.info(f"Processing sheet: {sheet_name}")
            
            header_row, activity_col, status_col = cls.find_headers(ws)
            
            if not header_row or not activity_col or not status_col:
                logger.warning(f"Could not find headers in sheet: {sheet_name}")
                continue
            
            for row_idx in range(header_row + 1, ws.max_row + 1):
                activity_cell = ws.cell(row=row_idx, column=activity_col)
                status_cell = ws.cell(row=row_idx, column=status_col)
                
                if not activity_cell.value:
                    continue
                
                activity_header = str(activity_cell.value).strip()
                completion_pct = cls.parse_percentage(status_cell.value)
                
                logger.info(f"Row {row_idx}: {activity_header} | Completion: {completion_pct}%")
                
                matching_activity = None
                for activity in activities:
                    kra_activity_name = activity['name']
                    if (kra_activity_name.lower() in activity_header.lower() or 
                        activity_header.lower() in kra_activity_name.lower()):
                        matching_activity = kra_activity_name
                        break
                
                if not matching_activity:
                    logger.warning(f"No KRA match found for: {activity_header}")
                    continue
                
                if matching_activity not in tracker_data:
                    tracker_data[matching_activity] = completion_pct
                    logger.info(f"Activity: {activity_header}, Completion: {completion_pct}%")
        
        logger.info(f"Total extracted: {len(tracker_data)} activities")
        return tracker_data


class StructureWorkParser:
    """Parse structure work tracker for green dates"""
    
    TARGET_SHEET = "Revised baseline with 60d NGT"
    
    TOWER_COLUMNS = {
        'Tower 6': ['FM', 'FQ', 'FU', 'FY', 'GC', 'GG', 'GK'],
        'Tower 7': ['E1', 'EM', 'EQ', 'EU', 'EY', 'FC', 'FG'],
        'T5': ['DE', 'DI', 'DM', 'DQ', 'DU', 'DY', 'EC']
    }
    
    @staticmethod
    def extract_tower_from_section(section_name):
        """Extract tower from section name"""
        match = re.search(r'Tower\s+(\d+)', section_name, re.IGNORECASE)
        if match:
            return f"Tower {match.group(1)}"
        if 'T5' in section_name.upper():
            return 'T5'
        return None
    
    @staticmethod
    def col_letter_to_index(col_letter):
        """Convert column letter to index"""
        result = 0
        for char in col_letter.upper():
            result = result * 26 + (ord(char) - ord('A') + 1)
        return result
    
    @staticmethod
    def find_target_sheet(wb, target_sheet_name):
        """Find target sheet"""
        if target_sheet_name in wb.sheetnames:
            return wb[target_sheet_name]
        
        for sheet_name in wb.sheetnames:
            if sheet_name.lower() == target_sheet_name.lower():
                return wb[sheet_name]
        
        keywords = target_sheet_name.lower().split()
        for sheet_name in wb.sheetnames:
            if all(kw in sheet_name.lower() for kw in keywords):
                return wb[sheet_name]
        
        return None
    
    @staticmethod
    def is_valid_floor(floor_str):
        """Check if valid floor"""
        if not floor_str:
            return False
        floor_str = str(floor_str).strip()
        return floor_str.endswith('F') or floor_str.endswith('f')
    
    @staticmethod
    def parse_date(cell_value):
        """Parse date"""
        if not cell_value:
            return None
        if isinstance(cell_value, datetime):
            return cell_value
        
        cell_str = str(cell_value).strip()
        if not cell_str or '###' in cell_str or cell_str == '-':
            return None
        
        date_formats = ["%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y", "%d.%m.%Y", "%Y/%m/%d"]
        
        for fmt in date_formats:
            try:
                return datetime.strptime(cell_str, fmt)
            except (ValueError, TypeError):
                continue
        
        return None
    
    @classmethod
    def parse_green_dates(cls, tracker_bytes, section_name, target_month_num=None, tracker_year=None):
        """Extract green dates for tower - prioritizing current tracker year"""
        logger.info(f"=== PARSING STRUCTURE - GREEN DATES ===")
        
        target_tower = cls.extract_tower_from_section(section_name)
        if not target_tower:
            logger.warning(f"Could not extract tower")
            return []
        
        logger.info(f"Tower: {target_tower}")
        
        tower_cols = cls.TOWER_COLUMNS.get(target_tower)
        if not tower_cols:
            logger.warning(f"No column mapping found for {target_tower}")
            return []
        
        wb = load_workbook(filename=tracker_bytes, data_only=False)
        green_dates_list = []
        
        ws = cls.find_target_sheet(wb, cls.TARGET_SHEET)
        if not ws:
            logger.warning(f"Sheet '{cls.TARGET_SHEET}' not found")
            return []
        
        logger.info(f"Using sheet: {ws.title}")
        
        col_indices = [cls.col_letter_to_index(col) for col in tower_cols]
        floor_sections = [
            {'name': 'South', 'rows': list(range(4, 11))},
            {'name': 'North', 'rows': list(range(13, 21))}
        ]
        
        green_cells_found = 0
        
        for section in floor_sections:
            logger.info(f"Processing section: {section['name']}")
            
            for row_idx in section['rows']:
                first_cell = ws.cell(row=row_idx, column=1).value
                
                if not first_cell:
                    continue
                
                floor_str = str(first_cell).strip()
                
                if not cls.is_valid_floor(floor_str):
                    continue
                
                for i, col_idx in enumerate(col_indices):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    
                    if is_cell_green(cell) and cell.value:
                        green_cells_found += 1
                        
                        parsed_date = cls.parse_date(cell.value)
                        
                        if parsed_date:
                            date_str = parsed_date.strftime('%d-%m-%Y')
                            month = parsed_date.month
                            year = parsed_date.year
                            logger.info(f"{floor_str}: {date_str} (month {month})")
                            
                            if target_month_num is None or month == target_month_num:
                                if tracker_year is None or year == tracker_year:
                                    green_dates_list.append(parsed_date)
        
        logger.info(f"Found {green_cells_found} green cells | Extracted {len(green_dates_list)} dates")
        return green_dates_list

# ============================================================================
# REPORT GENERATOR
# ============================================================================
class VerdiaReportGenerator:
    """Generate quarterly report"""
    def __init__(self):
        self.cos = init_cos()
        self.kra_key = None
        self.current_quarter = None
        self.quarter_months = []
        self.tracker_months = []
        self.quarter_year = None
        self.kra_data = {}
        self.tracker_keys = {}
        self.structure_green_dates = {}
    
    def get_latest_kra(self):
        """Find latest KRA file using year and quarter comparison"""
        logger.info("=== FINDING LATEST KRA ===")
        
        all_files = list_files(self.cos, "")
        kra_files = [f for f in all_files if 'KRA' in os.path.basename(f).upper() and f.endswith('.xlsx')]
        
        if not kra_files:
            logger.error("No KRA files found")
            return False
        
        logger.info(f"Found {len(kra_files)} KRA files")
        
        # Step 1: Extract quarter info from each KRA filename
        kra_with_quarters = []
        
        for file_path in kra_files:
            filename = os.path.basename(file_path)
            quarter, months, year = determine_quarter_from_kra(filename)
            
            if quarter and months and year:
                # Convert quarter to numeric for sorting (Q1→1, Q2→2, Q3→3, Q4→4)
                quarter_num = int(quarter[1])
                
                kra_with_quarters.append({
                    'path': file_path,
                    'filename': filename,
                    'quarter': quarter,
                    'months': months,
                    'year': year,
                    'sort_key': (year, quarter_num)
                })
                
                logger.info(f"  ✓ {filename} → Q{quarter_num} {year}")
            else:
                logger.warning(f"  ✗ {filename} → Could not extract quarter info")
        
        # Step 2: Validate we have at least one valid KRA file
        if not kra_with_quarters:
            logger.error("No KRA files with valid quarter information found")
            logger.error("Expected format: 'KRA Milestones for Month1 Month2 Month3 YYYY.xlsx'")
            return False
        
        # Step 3: Sort by (Year, Quarter) and pick the latest
        kra_with_quarters.sort(key=lambda x: x['sort_key'])
        latest = kra_with_quarters[-1]
        
        # Step 4: Set instance variables from latest KRA
        self.kra_key = latest['path']
        self.current_quarter = latest['quarter']
        self.quarter_months = latest['months']
        self.quarter_year = latest['year']
        self.tracker_months = [MONTH_SHIFT_MAP.get(m, m) for m in self.quarter_months]
        
        logger.info(f"\n{'='*80}")
        logger.info(f"Latest KRA: {latest['filename']}")
        logger.info(f"Quarter: {self.current_quarter} {self.quarter_year}")
        logger.info(f"Report Months: {self.quarter_months}")
        logger.info(f"Tracker Months: {self.tracker_months}")
        logger.info(f"{'='*80}\n")
        
        return True
    
    def parse_kra(self):
        """Parse KRA file"""
        logger.info("\n=== PARSING KRA FILE ===")
        
        raw = download_file(self.cos, self.kra_key)
        wb = load_workbook(filename=BytesIO(raw), data_only=True)
        
        kra_sheet = None
        for sheet_name in wb.sheetnames:
            if 'veridia' in sheet_name.lower() and 'target' in sheet_name.lower():
                kra_sheet = wb[sheet_name]
                logger.info(f"Using sheet: {sheet_name}")
                break
        
        if not kra_sheet:
            logger.error("No Veridia sheet found")
            return False
        
        parser = DynamicKRAParser()
        self.kra_data = parser.parse_kra_dynamic(kra_sheet, self.quarter_months)
        
        if not self.kra_data:
            logger.error("Failed to parse KRA")
            return False
        
        logger.info(f"Parsed {len(self.kra_data)} sections")
        return True
    
    def find_trackers(self):
        """Find tracker files with month shifting logic"""
        logger.info("\n=== FINDING TRACKER FILES ===")
        
        veridia_files = list_files(self.cos, "Veridia/")
        
        for section_name in self.kra_data.keys():
            self.tracker_keys[section_name] = {}
            
            if 'External Development' in section_name:
                tracker_pattern = 'External Development'
            elif 'Structure' in section_name or 'Slab' in section_name:
                tracker_pattern = 'Structure Work Tracker'
            else:
                tower_match = re.search(r'(Tower \d+|Green \d+)', section_name, re.IGNORECASE)
                if not tower_match:
                    continue
                tracker_pattern = f'{tower_match.group(1)}.*Tracker'
            
            logger.info(f"\nLooking for: {tracker_pattern}")
            
            for idx, report_month in enumerate(self.quarter_months):
                tracker_month = self.tracker_months[idx]
                tracker_month_num = MONTH_TO_NUM.get(tracker_month)
                
                # Determine tracker year based on quarter
                tracker_year = self.quarter_year
                
                # Q3 2026: report months are Dec 2025, Jan 2026, Feb 2026
                # tracker months (shifted) are Jan 2026, Feb 2026, Mar 2026
                if self.current_quarter == 'Q3':
                    # For Q3, tracker_year is always quarter_year (2026)
                    tracker_year = self.quarter_year
                elif self.current_quarter == 'Q4' and tracker_month == 'June':
                    # Only June tracker goes to next year for Q4
                    tracker_year = self.quarter_year + 1
                
                found = False
                for file_path in veridia_files:
                    filename = os.path.basename(file_path)
                    
                    if re.search(tracker_pattern, filename, re.IGNORECASE):
                        file_date = extract_date_from_filename(filename)
                        
                        if file_date and file_date.month == tracker_month_num and file_date.year == tracker_year:
                            self.tracker_keys[section_name][report_month] = file_path
                            logger.info(f"{report_month} (tracker: {tracker_month}): {filename}")
                            found = True
                            break
                
                if not found:
                    logger.warning(f"{report_month} (tracker: {tracker_month}): Not found")
    
    def _parse_tracker(self, tracker_file, section_name, activities, is_structure, report_month=None):
        """Parse tracker file and return (counts, data_type)"""
        try:
            raw = download_file(self.cos, tracker_file)
            
            if is_structure:
                logger.info(f"Parser: STRUCTURE WORK")
                
                report_month_num = MONTH_TO_NUM.get(report_month)
                logger.info(f"Extracting green dates for: {report_month} (month {report_month_num})")
                
                file_date = extract_date_from_filename(tracker_file)
                tracker_file_year = file_date.year if file_date else None
                logger.info(f"Tracker file year: {tracker_file_year}")
                
                green_dates = StructureWorkParser.parse_green_dates(
                    BytesIO(raw), 
                    section_name, 
                    report_month_num,
                    tracker_year=tracker_file_year
                )
                
                if not green_dates:
                    logger.warning(f"No green dates found for {report_month} {tracker_file_year}")
                else:
                    self.structure_green_dates[report_month] = green_dates
                    logger.info(f"Found {len(green_dates)} green dates for {report_month}")
                
                tracker_counts = {'Slab Casting': len(green_dates)}
                return tracker_counts, 'count'
                
            elif 'External Development' in section_name:
                logger.info(f"Parser: EXTERNAL DEVELOPMENT")
                
                kra_targets = self.kra_data.get(section_name, {}).get('targets', {})
                
                tracker_percentages = ExternalDevelopmentParser.parse_with_targets(
                    BytesIO(raw), 
                    activities, 
                    kra_targets
                )
                
                tracker_counts = {}
                for activity in activities:
                    act_name = activity['name']
                    if act_name in tracker_percentages:
                        tracker_counts[act_name] = tracker_percentages[act_name]
                    else:
                        tracker_counts[act_name] = 0
                
                logger.info(f"External Development data type: percentage")
                return tracker_counts, 'percentage'
            
            else:
                logger.info(f"Parser: TOWER")
                month_num = MONTH_TO_NUM.get(report_month)
                if not month_num:
                    logger.warning(f"No valid month")
                    return {act['name']: 0 for act in activities}, 'count'
                
                tracker_counts, data_type = TowerTrackerParser.parse(BytesIO(raw), activities, month_num)
                return tracker_counts, data_type
        
        except Exception as e:
            logger.error(f"Error: {e}")
            import traceback
            logger.debug(traceback.format_exc())
            return {act['name']: 0 for act in activities}, 'count'
    
    def build_report_data(self):
        """Build report data from trackers"""
        logger.info("\n=== BUILDING REPORT DATA ===")
        
        report_dfs = {}
        
        for section_name, section_data in self.kra_data.items():
            logger.info(f"\nProcessing: {section_name}")
            
            activities = section_data['activities']
            targets = section_data['targets']
            is_structure = section_data.get('is_structure', False)
            is_external = section_data.get('is_external', False)
            
            counts = {act['name']: {month: None for month in self.quarter_months} for act in activities}
            data_types = {act['name']: 'count' for act in activities}
            activity_tracker_months = {act['name']: None for act in activities}
            months_with_tracker = {act['name']: [] for act in activities}
            
            if section_name in self.tracker_keys:
                logger.info(f"Found trackers for {len(self.tracker_keys[section_name])} months")
                for report_month, tracker_file in self.tracker_keys[section_name].items():
                    logger.info(f"Parsing {report_month}: {os.path.basename(tracker_file)}")
                    tracker_counts, data_type = self._parse_tracker(
                        tracker_file, section_name, activities, is_structure, report_month
                    )
                    
                    for activity_name, count in tracker_counts.items():
                        if activity_name in counts:
                            counts[activity_name][report_month] = count
                            data_types[activity_name] = data_type
                            months_with_tracker[activity_name].append(report_month)
                            logger.info(f"{activity_name}: {count} (type: {data_type})")
                            
                            if count > 0 and activity_tracker_months[activity_name] is None:
                                activity_tracker_months[activity_name] = report_month
                                logger.info(f"  → Tracker month for {activity_name}: {report_month}")
            else:
                logger.warning(f"No trackers found")
            
            df = self._build_dataframe(section_name, activities, targets, counts, data_types, 
                                       activity_tracker_months, months_with_tracker)
            report_dfs[section_name] = df
        
        return report_dfs
    
    
    def _build_dataframe(self, section_name, activities, targets, counts, data_types, 
                    activity_tracker_months=None, months_with_tracker=None):
        """Build milestone dataframe - show all KRA targets, only cumulate with trackers"""
        data = []
        total_acts = len(activities)
        weightage = round(100 / total_acts, 2) if total_acts else 0
    
        is_external = 'External Development' in section_name
    
        if activity_tracker_months is None:
            activity_tracker_months = {}
        if months_with_tracker is None:
            months_with_tracker = {act['name']: [] for act in activities}
    
        delay_reasons_col_name = f"Delay Reasons_{self.quarter_months[-1]} {self.quarter_year}"
    
        for i, activity in enumerate(activities):
            name = activity['name']
            unit = activity.get('unit', 'Flat')
            unit_plural = f"{unit}s"
            has_any_kra_target = any((targets[name].get(m, 0) or 0) > 0 for m in self.quarter_months)
    
            # Initialize row with all columns set to empty strings
            row = {
                "Milestone": f"{i+1:02d}",
                "Activity": name,
                "Weightage": weightage,
                "Weighted Delay against Targets": "",
                delay_reasons_col_name: "",  # ALWAYS BLANK
            }
    
            total_target = 0
            target_parts = []
            data_is_percentage = data_types.get(name, 'count') == 'percentage'
            months_available = months_with_tracker.get(name, [])
    
            if is_external:
                if not has_any_kra_target:
                    row["Target"] = ""
                else:
                    tracker_month = activity_tracker_months.get(name)

                    if tracker_month:
                        display_month = tracker_month
                        display_target = int(targets[name].get(tracker_month, 0))
                        if display_target == 0:
                            display_target = 100
                    else:
                        months_with_data = []
                        for month in self.quarter_months:
                            month_target = targets[name].get(month, 0)
                            month_done = counts[name].get(month)

                            if month_target > 0 or (month_done is not None and month_done > 0):
                                months_with_data.append(month)

                        if months_with_data:
                            display_month = months_with_data[0]
                            display_target = int(targets[name].get(display_month, 0))
                            if display_target == 0:
                                display_target = 100
                        else:
                            display_month = self.quarter_months[0]
                            display_target = 100

                    row["Target"] = f"{display_target}% by {display_month}"

            else:
                for month in self.quarter_months:
                    target = int(targets[name].get(month, 0))
                    total_target += target
                    if target > 0:
                        target_parts.append(f"{target}-{month}")

                row["Target"] = f"{total_target} {unit_plural} ({', '.join(target_parts)})" if target_parts else ""
    
            cum_done = 0
            cum_target = 0
            total_achieved = 0
            last_pct = None
            any_tracker_seen = False
    
            # For non-external: Check if work exists in ANY month to show cumulative progress
            has_any_tracker_data = False
            if not is_external:
                for month in self.quarter_months:
                    if counts[name].get(month) is not None:
                        has_any_tracker_data = True
                        break
    
            for month in self.quarter_months:
                month_done = counts[name].get(month)
                month_target = targets[name].get(month, 0)

                if not has_any_kra_target:
                    row[f"% Work Done against Target-Till {month}"] = ""
                    row[f"Target achieved in {month}"] = ""
                    continue

                if is_external:
                    tracker_exists = month in months_available
                
                    if not tracker_exists:
                        row[f"% Work Done against Target-Till {month}"] = ""
                        row[f"Target achieved in {month}"] = ""
                        continue
                
                    any_tracker_seen = True
                
                    if data_is_percentage:
                        pct = float(month_done or 0.0)
                        pct = min(max(pct, 0.0), 100.0)
                        target_pct = float(month_target) if month_target else 100.0
                    else:
                        if month_target and month_target > 0:
                            pct = float(month_done) if month_done is not None else 0.0
                        else:
                            pct = float(month_done) if month_done is not None else 0.0
                        target_pct = int(month_target) if month_target else 100
                
                    pct = round(pct, 2)
                    target_pct = round(target_pct, 2)
                
                    achievement_against_target = min((pct / target_pct * 100), 100) if target_pct > 0 else 0
                    achievement_against_target = round(achievement_against_target, 2)
                
                    row[f"% Work Done against Target-Till {month}"] = f"{pct}%"
                    row[f"Target achieved in {month}"] = f"{pct}% out of {target_pct}% target"
                    total_achieved = pct
                    last_pct = achievement_against_target
    
                else:
                    # TOWER/STRUCTURE LOGIC
                    # If there's tracker data anywhere, show cumulative progress
                    if has_any_tracker_data:
                        if month_done is not None:
                            month_done = int(month_done)
                            month_target = int(month_target)
    
                            cum_done += month_done
                            cum_target += month_target
                            total_achieved += month_done
    
                            # If cumulative target is zero, keep percentage blank.
                            # This avoids showing 100% when no target exists.
                            if cum_target > 0:
                                pct = (cum_done / cum_target) * 100
                                pct = min(pct, 100.0)
                                pct = round(pct, 2)
                                row[f"% Work Done against Target-Till {month}"] = f"{pct}%"
                            else:
                                row[f"% Work Done against Target-Till {month}"] = ""
                            row[f"Target achieved in {month}"] = f"{int(month_done)} out of {int(month_target)} {unit_plural}"
                            if cum_target > 0:
                                last_pct = pct
                        else:
                            # No data for this month yet
                            month_target = int(month_target)
                            row[f"% Work Done against Target-Till {month}"] = ""
                            row[f"Target achieved in {month}"] = f"0 out of {month_target} {unit_plural}"
                    else:
                        # No tracker data at all - leave blank
                        row[f"% Work Done against Target-Till {month}"] = ""
                        row[f"Target achieved in {month}"] = ""
    
            # Calculate Weighted Delay (only if tracker data exists)
            if not has_any_kra_target:
                row["Total achieved"] = ""
                row["Weighted Delay against Targets"] = ""
            elif is_external:
                if any_tracker_seen:
                    row["Total achieved"] = f"{int(total_achieved) if total_achieved == int(total_achieved) else total_achieved}%"
                    if last_pct is not None and last_pct != 0:
                        row["Weighted Delay against Targets"] = f"{round((last_pct * weightage) / 100, 2)}%"
                    else:
                        row["Weighted Delay against Targets"] = ""
                else:
                    row["Total achieved"] = ""
                    row["Weighted Delay against Targets"] = ""
            else:
                if has_any_tracker_data:
                    row["Total achieved"] = f"{int(total_achieved)} {unit_plural}"
                    if last_pct is not None:
                        row["Weighted Delay against Targets"] = f"{round((last_pct * weightage) / 100, 2)}%"
                    else:
                        row["Weighted Delay against Targets"] = ""
                else:
                    row["Total achieved"] = ""
                    row["Weighted Delay against Targets"] = ""
            
            # CRITICAL: ALWAYS keep Delay Reasons blank
            row[delay_reasons_col_name] = ""
    
            data.append(row)
    
        # Build column order carefully
        cols = ["Milestone", "Activity", "Target"]
        
        # Add % Work Done columns for each month
        for m in self.quarter_months:
            cols.append(f"% Work Done against Target-Till {m}")
        
        # Add summary columns
        cols.extend(["Weightage", "Weighted Delay against Targets", "Total achieved"])
        
        # Add Target achieved columns for each month
        for m in self.quarter_months:
            cols.append(f"Target achieved in {m}")
        
        # Add Delay Reasons column at the END
        cols.append(delay_reasons_col_name)
    
        return pd.DataFrame(data, columns=cols)
        
    def _sort_sections(self, sections):
        """Sort sections: Structure, Towers, External"""
        structure_sections = {}
        tower_sections = {}
        external_sections = {}
        
        for section_name in sections:
            name_lower = section_name.lower()
            
            if 'structure' in name_lower or 'slab' in name_lower:
                structure_sections[section_name] = sections[section_name]
            elif 'external development' in name_lower:
                external_sections[section_name] = sections[section_name]
            else:
                tower_sections[section_name] = sections[section_name]
        
        sorted_towers = sorted(
            tower_sections.keys(),
            key=lambda x: int(re.search(r'\d+', x).group()) if re.search(r'\d+', x) else x.lower()
        )
        
        sorted_external = sorted(
            external_sections.keys(),
            key=lambda x: int(re.search(r'\d+', x).group()) if re.search(r'\d+', x) else x.lower()
        )
        
        result = {}
        for section in structure_sections:
            result[section] = structure_sections[section]
        for section in sorted_towers:
            result[section] = tower_sections[section]
        for section in sorted_external:
            result[section] = external_sections[section]
        
        return result
    
    def write_report(self, report_dfs):
        """Write Excel report"""
        filename = f"Veridia_Milestone_Report_{self.current_quarter}_{self.quarter_year}_{datetime.now():%Y%m%d}.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Time Delivery Milestones"
        
        ws.append([f"Veridia Time Delivery Milestones Report - {self.current_quarter} ({', '.join(self.quarter_months)}) {self.quarter_year}"])
        ws.append([f"Report Generated on: {datetime.now().strftime('%d-%m-%Y')}"])
        ws.append([])
        
        yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        grey = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        bold = Font(bold=True)
        thin = Side(style="thin")
        border = Border(top=thin, bottom=thin, left=thin, right=thin)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        sorted_report_dfs = self._sort_sections(report_dfs)
        
        for section_name in sorted_report_dfs.keys():
            df = sorted_report_dfs[section_name]
            
            ws.append([f"{section_name} Progress Against Milestones"])
            title_row = ws.max_row
            
            if not df.empty:
                num_cols = len(df.columns)
            else:
                num_cols = 11
            
            ws.merge_cells(f'A{title_row}:{get_column_letter(num_cols)}{title_row}')
            ws[f'A{title_row}'].font = bold
            ws[f'A{title_row}'].fill = grey
            ws[f'A{title_row}'].alignment = center
            
            if df.empty:
                header_cols = ["Milestone", "Activity", "Target"]
                for m in self.quarter_months:
                    header_cols.append(f"% Work Done against Target-Till {m}")
                header_cols.extend(["Weightage", "Weighted Delay against Targets"])
                for m in self.quarter_months:
                    header_cols.append(f"Target achieved in {m}")
                header_cols.extend(["Total achieved", f"Delay Reasons_{self.quarter_months[-1]} {self.quarter_year}"])
                
                ws.append(header_cols)
                header_row = ws.max_row
                
                for col_idx in range(1, len(header_cols) + 1):
                    ws.cell(header_row, col_idx).font = bold
                    ws.cell(header_row, col_idx).fill = grey
                    ws.cell(header_row, col_idx).border = border
                    ws.cell(header_row, col_idx).alignment = center
                
                logger.warning(f"No activities found for {section_name}")
            else:
                for r in dataframe_to_rows(df, index=False, header=True):
                    ws.append(r)
                
                header_row = title_row + 1
                for col_idx in range(1, len(df.columns) + 1):
                    ws.cell(header_row, col_idx).font = bold
                    ws.cell(header_row, col_idx).fill = grey
                    ws.cell(header_row, col_idx).border = border
                
                for row in ws.iter_rows(title_row + 2, ws.max_row):
                    for col_idx, cell in enumerate(row, 1):
                        cell.border = border
                        cell.alignment = center if cell.column > 2 else left
                
                # Calculate and display total delay
                try:
                    total_delay = sum(
                        float(str(v).strip().rstrip('%')) 
                        for v in df["Weighted Delay against Targets"] 
                        if v and str(v).strip() != ""
                    )
                except (ValueError, AttributeError):
                    total_delay = 0
                
                ws.append(["Total"])
                total_row = ws.max_row
                
                # Find the column index for "Weighted Delay against Targets"
                weighted_delay_col_idx = None
                for col_idx, col_name in enumerate(df.columns, 1):
                    if col_name == "Weighted Delay against Targets":
                        weighted_delay_col_idx = col_idx
                        break
                
                # Write total delay value in the correct column
                if weighted_delay_col_idx:
                    ws.cell(total_row, weighted_delay_col_idx).value = f"{round(total_delay, 2)}%"
                
                # Format total row
                for cell in ws[total_row]:
                    cell.font = bold
                    cell.fill = yellow
                    cell.border = border
            
            ws.append([])
        
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 50)
        
        wb.save(filename)
        logger.info(f"\nReport saved: {filename}")
        
        logger.info(f"\n=== GREEN DATES SUMMARY ===")
        if self.structure_green_dates:
            for month, dates in self.structure_green_dates.items():
                dates_str = ", ".join([d.strftime('%d-%m-%Y') for d in dates])
                logger.info(f"{month}: {dates_str}")
        else:
            logger.info("No green dates extracted")
    
    def generate(self):
        """Generate complete report"""
        logger.info("="*80)
        logger.info("=== VERIDIA QUARTERLY REPORT GENERATOR ===")
        logger.info("="*80)
        
        if not self.get_latest_kra():
            return False
        
        if not self.parse_kra():
            return False
        
        self.find_trackers()
        report_dfs = self.build_report_data()
        
        if not report_dfs:
            logger.error("No data generated")
            return False
        
        self.write_report(report_dfs)
        
        logger.info("\n" + "="*80)
        logger.info("REPORT GENERATION COMPLETE")
        logger.info(f"Quarter: {self.current_quarter} {self.quarter_year}")
        logger.info(f"Sections: {len(report_dfs)}")
        logger.info(f"Green Dates Extracted: {len(self.structure_green_dates)} months")
        logger.info("="*80)
        
        return True

# ============================================================================
# MAIN
# ============================================================================
def main():
    generator = VerdiaReportGenerator()
    success = generator.generate()
    return 0 if success else 1

if __name__ == "__main__":
    exit(main())
