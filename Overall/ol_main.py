import streamlit as st
import os
import sys
import importlib.util
from pathlib import Path

CURRENT_DIR = Path(__file__).resolve().parent
ROOT_DIR = CURRENT_DIR.parent

if str(CURRENT_DIR) not in sys.path:
    sys.path.insert(0, str(CURRENT_DIR))
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

try:
    from dotenv import load_dotenv
    load_dotenv(ROOT_DIR / ".env", override=True)
except Exception:
    pass

def load_local_module(module_name):
    module_path = CURRENT_DIR / f"{module_name}.py"
    spec = importlib.util.spec_from_file_location(f"overall_{module_name}", module_path)
    if spec is None or spec.loader is None:
        raise ModuleNotFoundError(module_name)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module

def get_config_value(key, default=None):
    """Read config from Streamlit secrets first, then root .env / environment."""
    try:
        if key in st.secrets:
            return st.secrets[key]
    except Exception:
        pass
    return os.getenv(key, default)

import pandas as pd
import requests
import json
import openpyxl
import time
import math
from io import BytesIO
from datetime import datetime
import io
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import io
from datetime import date
import concurrent.futures
from dateutil.relativedelta import relativedelta
import re
from collections import defaultdict
from dateutil.relativedelta import relativedelta

try:
    import ibm_boto3
    from ibm_botocore.client import Config
    EWS_LIG = load_local_module("EWS_LIG")
    Tower_G_and_H = load_local_module("Tower_G_and_H")
    Veridia = load_local_module("Veridia")
    Wavecity = load_local_module("Wavecity")
    Finishing = load_local_module("Finishing")
    Eden = load_local_module("Eden")
    from Fileformat import *
except ModuleNotFoundError as exc:
    st.error(
        "Missing Python dependency: "
        f"`{exc.name}`. Install the project requirements in the active virtualenv and reload Streamlit."
    )
    st.code(r"venv\Scripts\python.exe -m pip install -r requirements.txt")
    st.stop()


if 'tower2_finishing' not in st.session_state:
    st.session_state.tower2_finishing = "0%"
if 'tower3_finishing' not in st.session_state:
    st.session_state.tower3_finishing = "0%"
if 'tower4_finishing' not in st.session_state:
    st.session_state.tower4_finishing = "0%"
if 'tower5_finishing' not in st.session_state:
    st.session_state.tower5_finishing = "0%"
if 'tower6_finishing' not in st.session_state:
    st.session_state.tower6_finishing = "0%"
if 'tower7_finishing' not in st.session_state:
    st.session_state.tower7_finishing = "0%"


if 'towerf_finishing' not in st.session_state:
    st.session_state.towerf_finishing = "0%"
if 'towerg_finishing' not in st.session_state:
    st.session_state.towerg_finishing = "0%"
if 'towerh_finishing' not in st.session_state:
    st.session_state.towerh_finishing = "0%"

if 'wavecity_finishing' not in st.session_state:
    st.session_state.wavecity_finishing = "0%"

# st.session_state.overall
if 'overalldf' not in st.session_state:
    st.session_state.overalldf = pd.DataFrame()

COS_API_KEY = get_config_value("COS_API_KEY")
COS_SERVICE_INSTANCE_ID = get_config_value("COS_SERVICE_INSTANCE_ID")
COS_ENDPOINT = get_config_value("COS_ENDPOINT", "https://s3.us-south.cloud-object-storage.appdomain.cloud")
COS_BUCKET = get_config_value("COS_BUCKET", "projectreportnew")

missing_cos = [
    key for key, value in {
        "COS_API_KEY": COS_API_KEY,
        "COS_SERVICE_INSTANCE_ID": COS_SERVICE_INSTANCE_ID,
        "COS_ENDPOINT": COS_ENDPOINT,
        "COS_BUCKET": COS_BUCKET,
    }.items()
    if not value
]

if missing_cos:
    st.error("❌ Missing COS configuration: " + ", ".join(missing_cos))
    st.stop()


def to_excel(df):
    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        # Write the DataFrame to Excel, starting from row 1 to leave space for the title
        df.to_excel(writer, index=False, sheet_name='Sheet1', startrow=1)
        
        # Get the xlsxwriter workbook and worksheet objects
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']
        
        # Define a format for the title row with yellow background
        title_format = workbook.add_format({
            'bold': True,
            'bg_color': 'yellow',
            'align': 'center',
            'valign': 'vcenter'
        })
        
        # Add a title in the first row (e.g., "Tower Project Status")
        # worksheet.write(0, 0,0, f'Overall Project Report ({date.today()})', title_format)
        
        # Merge cells across the first row for the title (assuming the DataFrame has columns)
        worksheet.merge_range(0, 0, 0, len(df.columns)-1, f'Overall Project Report ({date.today()})', title_format)
        
    return output.getvalue()


cos_client = ibm_boto3.client(
    's3',
    ibm_api_key_id=COS_API_KEY,
    ibm_service_instance_id=COS_SERVICE_INSTANCE_ID,
    config=Config(signature_version='oauth'),
    endpoint_url=COS_ENDPOINT
)


@st.cache_data(ttl=300, show_spinner=False)
def get_cos_files():
    try:
        response = cos_client.list_objects_v2(Bucket=COS_BUCKET)
        files = [obj['Key'] for obj in response.get('Contents', []) if obj['Key'].endswith('.xlsx')]
        if not files:
            print("No .xlsx files found in the configured COS bucket. Please ensure Excel files are uploaded.")
        return files
    except Exception as e:
        print(f"Error fetching COS files: {e}")
        return ["Error fetching COS files",e]
    
files = get_cos_files() or []
# files = ["Error fetching COS files","Something Error"]
# files = ["EWS LIG P4/Structure Work Tracker (31-05-2025).xlsx", "Eden/Structure Work Tracker (31-05-2025).xlsx", "Eligo/Structure Work Tracker (31-05-2025).xlsx", "Eligo/Tower G Finishing Tracker (01-06-2025).xlsx", "Eligo/Tower H Finishing Tracker (01-06-2025).xlsx", "Veridia/Structure Work Tracker (31-05-2025).xlsx", "Veridia/Tower 4 Finishing Tracker (13-05-2025).xlsx", "Veridia/Tower 5 Finishing Tracker (01-06-2025).xlsx", "Veridia/Tower 7 Finishing Tracker (01-06-2025).xlsx", "Wave City Club/Structure Work Tracker Wave City Club all Block (11-06-2025).xlsx"]
# st.write(files)

today = datetime.today()
current_month = today.month
current_year = today.year
previous_month_date = today - relativedelta(months=1)
previous_month = previous_month_date.month
previous_year = previous_month_date.year


def extract_date(file_name):
    match = re.search(r"\((\d{2}-\d{2}-\d{4})\)", file_name)
    if match:
        return datetime.strptime(match.group(1), "%d-%m-%Y")
    return None


@st.cache_data(ttl=300, show_spinner=False)
def fetch_cos_file_bytes(file_key):
    response = cos_client.get_object(Bucket=COS_BUCKET, Key=file_key)
    return response["Body"].read()


def select_report_files(files):
    selected = {
        "veridia_t4_finishing": None,
        "veridia_t5_finishing": None,
        "veridia_t7_finishing": None,
        "eligo_tg_finishing": None,
        "eligo_th_finishing": None,
        "eligo_structure": None,
        "wave_structure": None,
        "ews_lig_structure": None,
        "eden_structure": None,
        "veridia_structure": None,
    }

    for file in files:
        if file.startswith("Veridia") and "Tower 4 Finishing Tracker" in file:
            selected["veridia_t4_finishing"] = file
        elif file.startswith("Veridia") and "Tower 5 Finishing Tracker" in file:
            selected["veridia_t5_finishing"] = file
        elif file.startswith("Veridia") and "Tower 7 Finishing Tracker" in file:
            selected["veridia_t7_finishing"] = file
        elif file.startswith("Eligo") and "Tower G Finishing Tracker" in file:
            selected["eligo_tg_finishing"] = file
        elif file.startswith("Eligo") and "Tower H Finishing Tracker" in file:
            selected["eligo_th_finishing"] = file
        elif file.startswith("Eligo") and "Structure Work Tracker" in file:
            selected["eligo_structure"] = file
        elif file.startswith("Wave City Club") and "Structure Work Tracker Wave City Club all Block" in file:
            selected["wave_structure"] = file
        elif "EWS LIG" in file and "Structure Work Tracker" in file:
            selected["ews_lig_structure"] = file
        elif file.startswith("Eden") and "Structure Work Tracker" in file:
            selected["eden_structure"] = file
        elif file.startswith("Veridia") and "Structure Work Tracker" in file:
            selected["veridia_structure"] = file

    return selected


def fetch_selected_cos_files(file_map):
    bytes_map = {}
    keys_to_fetch = [file_key for file_key in file_map.values() if file_key]

    if not keys_to_fetch:
        return bytes_map

    max_workers = min(6, len(keys_to_fetch))
    with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_to_key = {
            executor.submit(fetch_cos_file_bytes, file_key): file_key
            for file_key in keys_to_fetch
        }
        for future in concurrent.futures.as_completed(future_to_key):
            file_key = future_to_key[future]
            try:
                bytes_map[file_key] = future.result()
            except Exception as exc:
                st.warning(f"Failed to fetch `{file_key}` from COS: {exc}")

    return bytes_map




def GetOverallreport(files):
        ews_lig = {}
        veridia = {}
        eligo = {}
        eden_data = {}
        wave = {}
        file_map = select_report_files(files)
        fetched_files = fetch_selected_cos_files(file_map)

        if file_map["veridia_t4_finishing"] in fetched_files:
            Finishing.GetTower4Finishing(io.BytesIO(fetched_files[file_map["veridia_t4_finishing"]]))

        if file_map["veridia_t5_finishing"] in fetched_files:
            Finishing.GetTower5Finishing(io.BytesIO(fetched_files[file_map["veridia_t5_finishing"]]))

        if file_map["veridia_t7_finishing"] in fetched_files:
            Finishing.GetTower7Finishing(io.BytesIO(fetched_files[file_map["veridia_t7_finishing"]]))

        if file_map["eligo_tg_finishing"] in fetched_files:
            Finishing.GetTowerGFinishing(io.BytesIO(fetched_files[file_map["eligo_tg_finishing"]]))

        if file_map["eligo_th_finishing"] in fetched_files:
            Finishing.GetTowerHFinishing(io.BytesIO(fetched_files[file_map["eligo_th_finishing"]]))

        if file_map["eligo_structure"] in fetched_files:
            eligo = Tower_G_and_H.ProcessGandH(io.BytesIO(fetched_files[file_map["eligo_structure"]]))

        if file_map["wave_structure"] in fetched_files:
            wave = Wavecity.GetWaveCity(io.BytesIO(fetched_files[file_map["wave_structure"]]))

        if file_map["ews_lig_structure"] in fetched_files:
            ews_lig = EWS_LIG.ProcessEWSLIG(io.BytesIO(fetched_files[file_map["ews_lig_structure"]]))

    

        # #ELIGO TOWER STRUCTURE
        # for file in files:
        #     if file.startswith("Eligo") and "Structure Work Tracker" in file:
        #         response = cos_client.get_object(Bucket=COS_BUCKET, Key=file)
        #         eligo = ProcessGandH(io.BytesIO(response['Body'].read()))
        #         st.write(file,"✅")


        if file_map["eden_structure"] in fetched_files:
            eden_data = Eden.get_percentages(io.BytesIO(fetched_files[file_map["eden_structure"]]))


       
        if file_map["veridia_structure"] in fetched_files:
            veridia = Veridia.ProcessVeridia(io.BytesIO(fetched_files[file_map["veridia_structure"]]))

        # for i in files_after_or_on_10th:
        #     st.write(i)

        combined_data = []

        for data in [ews_lig, veridia, eligo, eden_data, wave]:
            if isinstance(data, list):
                # Check if it's a list of dicts (most typical for DataFrame input)
                if all(isinstance(item, dict) for item in data):
                    combined_data.extend(data)
                else:
                    st.warning("Some Files Are Missing")
            elif data is not None:
                st.warning("Some Files Are Missing")

        # Now safely create the DataFrame
        if combined_data:
            df = pd.DataFrame(combined_data)
            return df


# Group by project and subcategory (e.g. "Structure Work Tracker", "Tower G Finishing Tracker")
project_map = defaultdict(lambda: {"current": [], "previous": []})

for file in files:
    parts = file.split('/')
    project = parts[0]
    sub_path = '/'.join(parts[1:])  # handles nested folders
    file_date = extract_date(file)

    if not file_date:
        continue

    key = f"{project}/{sub_path.split('(')[0].strip()}"  # e.g., "Eligo/Tower G Finishing Tracker"
    
    if file_date.year == current_year and file_date.month == current_month:
        project_map[key]["current"].append(file)
    elif file_date.year == previous_year and file_date.month == previous_month:
        project_map[key]["previous"].append(file)

# Final list of selected files
selected_files = []

for key, month_files in project_map.items():
    if month_files["current"]:
        selected_files.extend(month_files["current"])
    elif month_files["previous"]:
        selected_files.extend(month_files["previous"])


st.header("OVERALL PROJECT REPORT")

if files and files[0] == "Error fetching COS files":
    st.warning(files[1])
elif not files:
    st.warning("No Excel files found in COS bucket.")
else:
    st.session_state.overalldf = GetOverallreport(selected_files)
    st.session_state.overalldf = st.session_state.overalldf.drop_duplicates(subset='Tower Name')
    st.session_state.check = True
    if st.session_state.check:
        if st.session_state.overalldf is not None and not st.session_state.overalldf.empty:
            st.title("Tower Project Status Table")
            st.dataframe(st.session_state.overalldf)
        # st.write(df)
            excel_data = to_excel(st.session_state.overalldf)
            st.session_state.overall = excel_data

            # st.dataframe(df)

            st.download_button(
                label="Download as Excel",
                data=excel_data,
                file_name="Overall_Project_Report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
